from __future__ import annotations

# main.py
# ------------------------------------------------------------------------------------
# Render 조회용: PC Collector가 업로드한 데이터만 사용
# 추가 기능:
# - 개인 조회: 오늘 완료 / 거절 / 취소 / 거절+취소율 표시
# - 관리자 대시보드: 이름 옆 거절+취소율 배지 표시
# - 이전등급 플러스 / 예정등급 플러스 엑셀 백업 다운로드
# - 기존 dashboard.xlsx 또는 plus-backup.xlsx 업로드 시 prev_plus / planned_plus 자동 복원
# - 룰렛 기능:
#   * 10건당 1회(기본) 룰렛 가능
#   * 10~100건 구간별 서로 다른 확률표 적용
#   * 운영주차 기준: 수요일 06:00 ~ 다음주 수요일 03:00
#   * 개인 조회 화면에서 룰렛 상태/주간누적/최근당첨내역 표시
#   * 관리자 지급 CSV 다운로드 / 룰렛 API 제공
# ------------------------------------------------------------------------------------



import csv
import json
import os
import re
import threading
import time
import zipfile
from datetime import date, timedelta, datetime, timezone
KST = timezone(timedelta(hours=9))

# Render 서버 시간이 UTC여도 날짜/운영주차/표시시간은 한국시간으로 맞춥니다.
os.environ.setdefault("TZ", "Asia/Seoul")
try:
    time.tzset()
except Exception:
    pass

def now_kst() -> datetime:
    return datetime.now(KST)

def today_kst() -> date:
    return now_kst().date()

def operation_date(now: datetime | None = None) -> date:
    """라웰 운영일 기준 날짜. 매일 오전 06:00에 새 운영일로 전환됩니다.

    예) 6/28 05:59 -> 6/27 운영일, 6/28 06:00 -> 6/28 운영일
    출석체크/룰렛/오늘완료 기준일에 사용합니다.
    """
    if now is None:
        now = now_kst()
    if now.hour < 6:
        return (now - timedelta(days=1)).date()
    return now.date()

def operation_day_bounds(op_date: date | None = None) -> tuple[datetime, datetime]:
    """운영일의 시작/종료 시각(KST): 06:00 ~ 다음날 05:59:59."""
    if op_date is None:
        op_date = operation_date()
    start = datetime(op_date.year, op_date.month, op_date.day, 6, 0, 0, tzinfo=KST)
    end = start + timedelta(days=1) - timedelta(seconds=1)
    return start, end

def kst_from_epoch(ts_value: int | float | str | None) -> datetime:
    try:
        return datetime.fromtimestamp(int(ts_value or 0), KST)
    except Exception:
        return datetime.now(KST)

from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from dateutil.relativedelta import relativedelta
from fastapi import FastAPI, File, Form, Request, UploadFile, Body, HTTPException
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, Response
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from starlette.middleware.sessions import SessionMiddleware
from starlette.responses import StreamingResponse

import roulette_db as roulette_db_module
from roulette_db import (
    RouletteDB,
    normalize_phone,
    get_operational_cycle_bounds,
    get_previous_operational_cycle_bounds,
)

# roulette_db 내부에서 date.today()/datetime.now()를 쓰는 경우에도
# 라웰 운영일(06:00~다음날 05:59) 기준으로 동작하도록 보정합니다.
# 함수 객체는 그대로 사용하되, roulette_db 모듈의 전역 date/datetime 참조만 교체합니다.
class _OperationalDate(date):
    @classmethod
    def today(cls):
        d = operation_date()
        return cls(d.year, d.month, d.day)


class _OperationalDateTime(datetime):
    @classmethod
    def now(cls, tz=None):
        base = now_kst()
        if tz is not None:
            base = base.astimezone(tz)
        return cls.fromtimestamp(base.timestamp(), base.tzinfo)


try:
    roulette_db_module.date = _OperationalDate
    roulette_db_module.datetime = _OperationalDateTime
except Exception:
    pass

# -----------------------------
# Config
# -----------------------------
RIDERS_CACHE_TTL = 10
STATUS_CACHE_TTL = 10
DELIVERY_STATUS_CACHE_TTL = 5

# 수집 오류가 기존 정상 데이터를 초기화하지 못하도록 하는 서버측 안전 기준
MIN_RIDERS_INGEST = max(1, int(os.getenv("MIN_RIDERS_INGEST", "50")))
MIN_DELIVERY_STATUS_INGEST = max(1, int(os.getenv("MIN_DELIVERY_STATUS_INGEST", "30")))
MIN_INGEST_RATIO = min(1.0, max(0.1, float(os.getenv("MIN_INGEST_RATIO", "0.65"))))

RATE_WINDOW_SEC = 60
RATE_MAX_REQ = 30

ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "0315")
SESSION_SECRET = os.getenv("SESSION_SECRET", "rider-welfare-admin-secret")
INGEST_TOKEN = (os.getenv("INGEST_TOKEN") or "").strip()

OVERRIDE_FILE = "join_overrides.json"            # key: "normname|login4" -> "YYYY-MM-DD"
LOGIN4_FILE = "login4_overrides.json"            # key: "normname|real4"  -> "login4"
PREVPLUS_FILE = "prevplus_overrides.json"        # key: "normname|login4" -> int
PLANNEDPLUS_FILE = "plannedplus_overrides.json"  # key: "normname|login4" -> int
ATTENDANCE_FILE = "attendance_log.json"          # key: login_key -> {period_start, period_end, days:[YYYY-MM-DD...]}
STICKER_FILE = "sticker_overrides.json"          # key: login_key -> bool
ADMIN_SETTINGS_FILE = "admin_settings.json"        # 관리자 비밀번호 등 복원 가능한 설정

ATTENDANCE_BONUS_PER_DAY = 5
ATTENDANCE_MIN_TODAY_COMPLETE = 6
MAX_PLUS_INPUT = 999999999

NMAX_START_DATE = date(2026, 5, 28)
NMAX_LABEL = "26.05.28 ~ (어제)"
EVENT_NAME = "NMAX"

_override_lock = threading.Lock()
_login4_lock = threading.Lock()
_prevplus_lock = threading.Lock()
_plannedplus_lock = threading.Lock()
_att_lock = threading.Lock()
_sticker_lock = threading.Lock()

_rate_bucket: Dict[str, List[float]] = {}

# 저장 위치
STORE_DIR = Path(os.getenv("STORE_DIR", "."))
STORE_DIR.mkdir(parents=True, exist_ok=True)

# Render Disk를 쓰는 경우, 모든 운영 데이터 파일을 STORE_DIR에 저장합니다.
# 기존 루트 경로에 남아있던 파일은 최초 실행 시 STORE_DIR로 1회 복사합니다.
_STATE_FILES = [
    OVERRIDE_FILE, LOGIN4_FILE, PREVPLUS_FILE, PLANNEDPLUS_FILE,
    ATTENDANCE_FILE, STICKER_FILE, ADMIN_SETTINGS_FILE,
    "teams.json", "team_weekly_sets.json", "team_period_stats.json",
]
for _fn in _STATE_FILES:
    try:
        _src = Path(_fn)
        _dst = STORE_DIR / _fn
        if _src.exists() and not _dst.exists():
            _dst.write_bytes(_src.read_bytes())
    except Exception:
        pass

OVERRIDE_FILE = str(STORE_DIR / OVERRIDE_FILE)
LOGIN4_FILE = str(STORE_DIR / LOGIN4_FILE)
PREVPLUS_FILE = str(STORE_DIR / PREVPLUS_FILE)
PLANNEDPLUS_FILE = str(STORE_DIR / PLANNEDPLUS_FILE)
ATTENDANCE_FILE = str(STORE_DIR / ATTENDANCE_FILE)
STICKER_FILE = str(STORE_DIR / STICKER_FILE)
ADMIN_SETTINGS_FILE = str(STORE_DIR / ADMIN_SETTINGS_FILE)
RIDERS_STORE = STORE_DIR / "store_riders.json"
STATUS_STORE = STORE_DIR / "store_status.json"
DELIVERY_STATUS_STORE = STORE_DIR / "store_delivery_status.json"

# 룰렛 DB
roulette_db = RouletteDB(str(STORE_DIR))

# 메모리 캐시
_riders_cache: Dict[str, Any] = {"ts": 0.0, "data": None}
_status_cache: Dict[str, Any] = {}
_delivery_status_cache: Dict[str, Any] = {"ts": 0.0, "data": None}

app = FastAPI(title="라웰 등급 조회 (Collector Ingest)")
app.add_middleware(SessionMiddleware, secret_key=SESSION_SECRET)


# -----------------------------
# HTML helpers
# -----------------------------
def html_page(title: str, body: str) -> str:
    return f"""<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>{title}</title>
</head>
<body style="font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif; padding:16px; background:#fafafa;">
  <div style="max-width:1400px; margin:0 auto;">
    {body}
  </div>
</body>
</html>"""


def norm_name(x: str) -> str:
    return re.sub(r"\s+", "", (x or "")).strip().lower()


def last4_from_phone(phone: str) -> str:
    m = re.search(r"(\d{4})\s*$", (phone or "").replace(" ", ""))
    return m.group(1) if m else ""


def mask_phone(phone: str) -> str:
    p = phone or ""
    m = re.search(r"(\d{2,3})-?(\d{3,4})-?(\d{4})", p)
    if not m:
        return p
    a, b, c = m.group(1), m.group(2), m.group(3)
    return f"{a}-****-{c}"


def rate_limit(ip: str) -> bool:
    now = time.time()
    arr = _rate_bucket.get(ip, [])
    arr = [t for t in arr if now - t <= RATE_WINDOW_SEC]
    if len(arr) >= RATE_MAX_REQ:
        _rate_bucket[ip] = arr
        return False
    arr.append(now)
    _rate_bucket[ip] = arr
    return True


def safe_date_parse(s: str) -> Optional[date]:
    s = (s or "").strip()
    try:
        return date.fromisoformat(s)
    except Exception:
        return None


def is_ended_contract(rider: Dict[str, Any]) -> bool:
    st = rider.get("accountStatus") or {}
    code = (st.get("code") or "").upper()
    desc = st.get("desc") or ""
    if "END" in code or "TERMIN" in code or "EXPIRE" in code:
        return True
    if "계약" in desc and "종료" in desc:
        return True
    return False


# -----------------------------
# Ratio helpers
# -----------------------------
def int0(v: Any) -> int:
    try:
        return int(v or 0)
    except Exception:
        return 0


def get_complete_count(dac: Dict[str, Any]) -> int:
    """완료건수: 푸드 + 비마트 + 배민스토어 전체 완료.
    v4는 totalComplete를 우선 사용하고, 없으면 세부 항목 합산, 마지막으로 구버전 complete를 사용합니다.
    """
    dac = dac or {}
    if "totalComplete" in dac:
        return int0(dac.get("totalComplete"))

    if any(k in dac for k in ("foodComplete", "bmartComplete", "storeComplete")):
        return (
            int0(dac.get("foodComplete"))
            + int0(dac.get("bmartComplete"))
            + int0(dac.get("storeComplete"))
        )

    return int0(dac.get("complete"))


def get_reject_count(dac: Dict[str, Any]) -> int:
    """거절: 푸드 거절만 반영. 비마트/배민스토어 거절은 제외합니다."""
    dac = dac or {}
    if "foodReject" in dac:
        return int0(dac.get("foodReject"))
    return int0(dac.get("reject"))


def get_cancel_count(dac: Dict[str, Any]) -> int:
    """취소: 푸드 + 비마트 + 배민스토어 전체 취소."""
    dac = dac or {}
    if "totalCancel" in dac:
        return int0(dac.get("totalCancel"))

    if any(k in dac for k in ("foodCancel", "bmartCancel", "storeCancel")):
        return (
            int0(dac.get("foodCancel"))
            + int0(dac.get("bmartCancel"))
            + int0(dac.get("storeCancel"))
        )

    return int0(dac.get("cancel"))


def calc_bad_ratio(complete: int, reject: int, cancel: int) -> Dict[str, Any]:
    complete = int(complete or 0)
    reject = int(reject or 0)
    cancel = int(cancel or 0)

    # 관제와 동일 기준: (푸드 거절 + 전체 취소) / 전체 완료
    total = complete
    bad = reject + cancel

    if complete > 0:
        ratio = round((bad / complete) * 100, 1)
    else:
        ratio = 0.0

    if ratio <= 19:
        fg = "#15803d"
        bg = "#e8f7ee"
        label = "양호"
    elif ratio < 30:
        fg = "#a16207"
        bg = "#fff8db"
        label = "주의"
    else:
        fg = "#b91c1c"
        bg = "#fdecec"
        label = "위험"

    return {
        "total": total,
        "bad": bad,
        "ratio": ratio,
        "fg": fg,
        "bg": bg,
        "label": label,
    }


def build_today_stats_map(ds: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
    out: Dict[str, Dict[str, Any]] = {}
    rows = ds.get("data") or []
    if not isinstance(rows, list):
        return out

    for row in rows:
        if not isinstance(row, dict):
            continue
        nm = norm_name(row.get("name") or "")
        ph = (row.get("phoneNumber") or "").replace(" ", "")
        real4 = last4_from_phone(ph)
        if not nm or not real4:
            continue

        dac = row.get("deliveryAcceptanceCount") or {}
        complete = get_complete_count(dac)
        reject = get_reject_count(dac)
        cancel = get_cancel_count(dac)
        ratio_info = calc_bad_ratio(complete, reject, cancel)

        peak = row.get("deliveryPeakTimeCount") or {}
        hourly = row.get("hourlyCompleted") or []
        peak_complete = {
            "morning_lunch": int0(peak.get("morning")),
            "afternoon_offpeak": int0(peak.get("afternoon")),
            "dinner_peak": int0(peak.get("evening")),
            "late_night": int0(peak.get("midnight")),
        }

        st = row.get("status") or {}
        out[f"{nm}|{real4}"] = {
            "complete": complete,
            "reject": reject,
            "cancel": cancel,
            "status_desc": st.get("desc") or "-",
            "peak_complete": peak_complete,
            "hourly_completed": hourly if isinstance(hourly, list) else [],
            **ratio_info,
        }

    return out

# -----------------------------
# Grade rules
# -----------------------------
def grade_from_total(total: int) -> str:
    if total <= 479:
        return "무등급"
    if total <= 719:
        return "R5"
    if total <= 959:
        return "R4"
    if total <= 1199:
        return "R3"
    if total <= 1439:
        return "R2"
    return "R1"


def next_grade_target(total: int) -> Tuple[Optional[str], Optional[int]]:
    thresholds = [
        ("무등급", 0),
        ("R5", 480),
        ("R4", 720),
        ("R3", 960),
        ("R2", 1200),
        ("R1", 1440),
    ]
    cur = grade_from_total(total)
    idx = [g for g, _ in thresholds].index(cur)
    if cur == "R1":
        return None, None
    nxt_g, nxt_t = thresholds[idx + 1]
    return nxt_g, max(0, nxt_t - total)


# -----------------------------
# Period logic
# -----------------------------
def clamp_day(year: int, month: int, target_day: int) -> date:
    first = date(year, month, 1)
    last_day = (first + relativedelta(months=1) - timedelta(days=1)).day
    return date(year, month, min(target_day, last_day))


def current_period(join_date: date, today: date) -> Tuple[date, date]:
    join_day = join_date.day
    this_anchor = clamp_day(today.year, today.month, join_day)

    if today >= this_anchor:
        start_d = this_anchor
    else:
        prev = date(today.year, today.month, 1) + relativedelta(months=-1)
        start_d = clamp_day(prev.year, prev.month, join_day)

    next_m = date(start_d.year, start_d.month, 1) + relativedelta(months=1)
    next_anchor = clamp_day(next_m.year, next_m.month, join_day)

    end_inclusive = next_anchor - timedelta(days=1)
    return start_d, end_inclusive


def period_to_from_to(start_d: date, end_inclusive: date) -> Tuple[date, date]:
    api_max = operation_date() - timedelta(days=1)
    from_d = start_d
    to_d = min(end_inclusive, api_max)
    return from_d, to_d


# -----------------------------
# Join-date override
# -----------------------------
def load_overrides() -> Dict[str, str]:
    with _override_lock:
        if not os.path.exists(OVERRIDE_FILE):
            return {}
        try:
            with open(OVERRIDE_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            return data if isinstance(data, dict) else {}
        except Exception:
            return {}


def save_overrides(data: Dict[str, str]) -> None:
    with _override_lock:
        with open(OVERRIDE_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)


# -----------------------------
# Login4 override
# -----------------------------
def load_login4_map() -> Dict[str, str]:
    with _login4_lock:
        if not os.path.exists(LOGIN4_FILE):
            return {}
        try:
            with open(LOGIN4_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                return {str(k): str(v) for k, v in data.items()}
            return {}
        except Exception:
            return {}


def save_login4_map(data: Dict[str, str]) -> None:
    with _login4_lock:
        with open(LOGIN4_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)


def get_login4_for_rider(rider: Dict[str, Any]) -> Tuple[str, str, str]:
    nm = rider.get("name") or ""
    ph = rider.get("phoneNumber") or ""
    real4 = last4_from_phone(ph)
    k_real = f"{norm_name(nm)}|{real4}"

    m = load_login4_map()
    if k_real in m and re.fullmatch(r"\d{4}", (m[k_real] or "").strip()):
        return m[k_real].strip(), real4, "override"
    return real4, real4, "real"


def set_login4_override(name_norm: str, real4: str, login4: str) -> None:
    m = load_login4_map()
    m[f"{name_norm}|{real4}"] = login4
    save_login4_map(m)


def clear_login4_override(name_norm: str, real4: str) -> None:
    m = load_login4_map()
    k = f"{name_norm}|{real4}"
    if k in m:
        del m[k]
    save_login4_map(m)


def get_effective_join_date_by_login_key(rider: Dict[str, Any], login4: str) -> Tuple[date, str]:
    nm = rider.get("name") or ""
    key = f"{norm_name(nm)}|{login4}"

    ov = load_overrides().get(key)
    if ov:
        d = safe_date_parse(ov)
        if d:
            return d, "override"

    created_raw = rider.get("createdDate")
    if isinstance(created_raw, str) and len(created_raw) >= 10:
        try:
            return date.fromisoformat(created_raw[:10]), "createdDate"
        except Exception:
            pass

    return today_kst(), "fallback"


# -----------------------------
# PrevPlus override
# -----------------------------
def load_prevplus_map() -> Dict[str, int]:
    with _prevplus_lock:
        if not os.path.exists(PREVPLUS_FILE):
            return {}
        try:
            with open(PREVPLUS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            out: Dict[str, int] = {}
            if isinstance(data, dict):
                for k, v in data.items():
                    try:
                        out[str(k)] = int(v)
                    except Exception:
                        pass
            return out
        except Exception:
            return {}


def save_prevplus_map(data: Dict[str, int]) -> None:
    with _prevplus_lock:
        with open(PREVPLUS_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)


def get_prevplus(login_key: str) -> int:
    return int(load_prevplus_map().get(login_key, 0) or 0)


def set_prevplus(login_key: str, v: int) -> None:
    m = load_prevplus_map()
    m[login_key] = int(v)
    save_prevplus_map(m)


def clear_prevplus(login_key: str) -> None:
    m = load_prevplus_map()
    if login_key in m:
        del m[login_key]
    save_prevplus_map(m)


# -----------------------------
# PlannedPlus override
# -----------------------------
def load_plannedplus_map() -> Dict[str, int]:
    with _plannedplus_lock:
        if not os.path.exists(PLANNEDPLUS_FILE):
            return {}
        try:
            with open(PLANNEDPLUS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            out: Dict[str, int] = {}
            if isinstance(data, dict):
                for k, v in data.items():
                    try:
                        out[str(k)] = int(v)
                    except Exception:
                        pass
            return out
        except Exception:
            return {}


def save_plannedplus_map(data: Dict[str, int]) -> None:
    with _plannedplus_lock:
        with open(PLANNEDPLUS_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)


def get_plannedplus(login_key: str) -> int:
    return int(load_plannedplus_map().get(login_key, 0) or 0)


def set_plannedplus(login_key: str, v: int) -> None:
    m = load_plannedplus_map()
    m[login_key] = int(v)
    save_plannedplus_map(m)


def clear_plannedplus(login_key: str) -> None:
    m = load_plannedplus_map()
    if login_key in m:
        del m[login_key]
    save_plannedplus_map(m)


# -----------------------------
# Attendance
# -----------------------------
def load_attendance_map() -> Dict[str, Any]:
    with _att_lock:
        if not os.path.exists(ATTENDANCE_FILE):
            return {}
        try:
            with open(ATTENDANCE_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            return data if isinstance(data, dict) else {}
        except Exception:
            return {}


def save_attendance_map(data: Dict[str, Any]) -> None:
    with _att_lock:
        with open(ATTENDANCE_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)


def _att_get_record(att_map: Dict[str, Any], login_key: str) -> Dict[str, Any]:
    rec = att_map.get(login_key)
    if isinstance(rec, dict):
        return rec
    return {"period_start": "", "period_end": "", "days": []}


def attendance_rollover_if_needed(login_key: str, cur_start: date, cur_end_incl: date) -> Tuple[int, int, bool]:
    """
    기간이 바뀌면:
    - 지난 기간의 현재플러스(planned_plus + sticker_bonus)를 이전등급플러스(prev_plus)로 이동
    - 현재플러스(planned_plus) 초기화
    - 출석기록(days) 초기화
    """
    att_map = load_attendance_map()
    rec = _att_get_record(att_map, login_key)

    rec_ps = safe_date_parse(rec.get("period_start", ""))
    rec_pe = safe_date_parse(rec.get("period_end", ""))

    # 최초 생성
    if rec_ps is None or rec_pe is None or not rec.get("period_start"):
        rec["period_start"] = cur_start.isoformat()
        rec["period_end"] = cur_end_incl.isoformat()
        rec["days"] = list(dict.fromkeys(rec.get("days") or []))
        att_map[login_key] = rec
        save_attendance_map(att_map)
        return 0, get_prevplus(login_key), False

    # 같은 기간이면 롤오버 없음
    if rec_ps == cur_start and rec_pe == cur_end_incl:
        return 0, get_prevplus(login_key), False

    # 기간 변경 -> 현재기간 플러스 전체를 이전등급 플러스로 이관
    old_planned_plus = get_plannedplus(login_key)
    old_sticker_bonus = get_sticker_bonus(login_key)
    moved_bonus = int(old_planned_plus or 0) + int(old_sticker_bonus or 0)

    set_prevplus(login_key, moved_bonus)
    set_plannedplus(login_key, 0)

    # 새 기간으로 출석기록 초기화
    rec["period_start"] = cur_start.isoformat()
    rec["period_end"] = cur_end_incl.isoformat()
    rec["days"] = []
    att_map[login_key] = rec
    save_attendance_map(att_map)

    return moved_bonus, get_prevplus(login_key), True


def attendance_is_checked_today(login_key: str, today: date, cur_start: date, cur_end_incl: date) -> bool:
    att_map = load_attendance_map()
    rec = _att_get_record(att_map, login_key)
    if rec.get("period_start") != cur_start.isoformat() or rec.get("period_end") != cur_end_incl.isoformat():
        return False
    days = rec.get("days") or []
    if not isinstance(days, list):
        return False
    return today.isoformat() in set([d for d in days if isinstance(d, str)])


def attendance_mark_today(login_key: str, today: date, cur_start: date, cur_end_incl: date) -> bool:
    att_map = load_attendance_map()
    rec = _att_get_record(att_map, login_key)

    if rec.get("period_start") != cur_start.isoformat() or rec.get("period_end") != cur_end_incl.isoformat():
        rec["period_start"] = cur_start.isoformat()
        rec["period_end"] = cur_end_incl.isoformat()
        rec["days"] = []

    days = rec.get("days") or []
    if not isinstance(days, list):
        days = []
    days_set = set([d for d in days if isinstance(d, str)])

    if today.isoformat() in days_set:
        return False

    days.append(today.isoformat())
    rec["days"] = list(dict.fromkeys(days))
    att_map[login_key] = rec
    save_attendance_map(att_map)
    return True


def attendance_count_in_period(login_key: str, cur_start: date, cur_end_incl: date) -> int:
    att_map = load_attendance_map()
    rec = _att_get_record(att_map, login_key)
    if rec.get("period_start") != cur_start.isoformat() or rec.get("period_end") != cur_end_incl.isoformat():
        return 0
    days = rec.get("days") or []
    if not isinstance(days, list):
        return 0
    return len(set([d for d in days if isinstance(d, str)]))

# -----------------------------
# Sticker
# -----------------------------
def load_sticker_map() -> Dict[str, bool]:
    with _sticker_lock:
        if not os.path.exists(STICKER_FILE):
            return {}
        try:
            with open(STICKER_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                return {str(k): bool(v) for k, v in data.items()}
            return {}
        except Exception:
            return {}


def save_sticker_map(data: Dict[str, bool]) -> None:
    with _sticker_lock:
        with open(STICKER_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)


def is_sticker_attached(login_key: str) -> bool:
    return bool(load_sticker_map().get(login_key, False))


def get_sticker_bonus(login_key: str) -> int:
    return 20 if is_sticker_attached(login_key) else 0


def set_sticker_attached(login_key: str, checked: bool) -> None:
    m = load_sticker_map()
    m[login_key] = bool(checked)
    save_sticker_map(m)


# -----------------------------
# Store helpers
# -----------------------------
def _read_json(path: Path, default):
    try:
        if not path.exists():
            return default
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return default


def _write_json(path: Path, data):
    """임시 파일에 완전히 기록한 뒤 원자적으로 교체합니다."""
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(path.name + ".tmp")
    tmp.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(tmp, path)


def _backup_store_file(path: Path) -> None:
    """정상 데이터 교체 직전에 직전본 1개를 보관합니다."""
    if not path.exists():
        return
    backup = path.with_name(path.name + ".last_good")
    tmp = backup.with_name(backup.name + ".tmp")
    tmp.write_bytes(path.read_bytes())
    os.replace(tmp, backup)


def _valid_rider_identity_count(riders: List[Dict[str, Any]]) -> int:
    keys = set()
    for rider in riders:
        if not isinstance(rider, dict):
            continue
        name = norm_name(rider.get("name") or "")
        real4 = last4_from_phone(rider.get("phoneNumber") or "")
        if name and real4:
            keys.add(f"{name}|{real4}")
    return len(keys)


def _delivery_status_row_count(data: Dict[str, Any]) -> int:
    rows = data.get("data") or []
    if not isinstance(rows, list):
        return 0
    return sum(1 for row in rows if isinstance(row, dict) and norm_name(row.get("name") or ""))


def _require_ingest(request: Request):
    token = request.headers.get("x-ingest-token", "")
    if not INGEST_TOKEN or token != INGEST_TOKEN:
        return JSONResponse({"ok": False, "error": "UNAUTHORIZED"}, status_code=401)
    return None


# -----------------------------
# Data access
# -----------------------------
def fetch_riders_cached() -> List[Dict[str, Any]]:
    now = time.time()
    if _riders_cache["data"] is not None and now - _riders_cache["ts"] <= RIDERS_CACHE_TTL:
        return _riders_cache["data"]

    store = _read_json(RIDERS_STORE, {"riders": []})
    riders = store.get("riders") or []
    if not isinstance(riders, list):
        riders = []
    riders = [r for r in riders if isinstance(r, dict)]
    riders = [r for r in riders if not is_ended_contract(r)]

    _riders_cache["ts"] = now
    _riders_cache["data"] = riders
    return riders


def fetch_status_complete_map_cached(from_d: date, to_d: date) -> Dict[str, int]:
    key = f"{from_d.isoformat()}_{to_d.isoformat()}"
    now = time.time()
    cached = _status_cache.get(key)
    if cached and now - cached["ts"] <= STATUS_CACHE_TTL:
        return cached["data"]

    all_status = _read_json(STATUS_STORE, {})
    raw = all_status.get(key) or {}
    # 신형: {completeMap, historyMap}; 구형: completeMap 자체
    m = raw.get("completeMap") if isinstance(raw, dict) and isinstance(raw.get("completeMap"), dict) else raw
    out: Dict[str, int] = {}
    if isinstance(m, dict):
        for k, v in m.items():
            try:
                out[str(k)] = int(v)
            except Exception:
                pass

    _status_cache[key] = {"ts": now, "data": out}
    return out


def fetch_status_history_map(from_d: date, to_d: date) -> Dict[str, Dict[str, Any]]:
    key = f"{from_d.isoformat()}_{to_d.isoformat()}"
    all_status = _read_json(STATUS_STORE, {}) or {}
    raw = all_status.get(key) or {}
    if not isinstance(raw, dict):
        return {}
    hm = raw.get("historyMap") or {}
    return hm if isinstance(hm, dict) else {}


def has_status_range(from_d: date, to_d: date) -> bool:
    key = f"{from_d.isoformat()}_{to_d.isoformat()}"
    all_status = _read_json(STATUS_STORE, {}) or {}
    return isinstance(all_status, dict) and (key in all_status)


def fetch_delivery_status_cached() -> Dict[str, Any]:
    now = time.time()
    if _delivery_status_cache["data"] is not None and now - _delivery_status_cache["ts"] <= DELIVERY_STATUS_CACHE_TTL:
        return _delivery_status_cache["data"]

    st = _read_json(DELIVERY_STATUS_STORE, {"data": {}, "ts": 0})
    data = st.get("data") if isinstance(st, dict) else {}
    if not isinstance(data, dict):
        data = {}

    _delivery_status_cache["ts"] = now
    _delivery_status_cache["data"] = data
    return data


def store_ready() -> bool:
    st = _read_json(RIDERS_STORE, {})
    riders = st.get("riders") if isinstance(st, dict) else None
    return isinstance(riders, list) and len(riders) > 0


def not_ready_page() -> HTMLResponse:
    body = """
    <div style="background:#fff;border:1px solid #e8e8e8;border-radius:16px;padding:16px;max-width:720px;margin:0 auto;">
      <h3 style="margin-top:0;">아직 데이터가 없습니다</h3>
      <div style="color:#666;line-height:1.6;">
        Render는 배민 API를 직접 호출하지 않습니다.<br/>
        PC에서 collector.py를 실행해서 <b>/ingest</b>로 데이터를 먼저 업로드해야 조회가 됩니다.
      </div>
      <div style="margin-top:12px;">
        <a href="/health" style="text-decoration:none;color:#111;">Health 보기</a>
      </div>
    </div>
    """
    return HTMLResponse(html_page("데이터 없음", body), status_code=200)


# -----------------------------
# Roulette helpers
# -----------------------------
def find_rider_by_name_login4(name: str, login4: str) -> Optional[Dict[str, Any]]:
    name_in = norm_name(name)
    login4 = (login4 or "").strip()

    riders = fetch_riders_cached()
    candidates = [r for r in riders if norm_name(r.get("name", "")) == name_in]
    matches: List[Dict[str, Any]] = []
    for r in candidates:
        l4, _, _ = get_login4_for_rider(r)
        if l4 == login4:
            matches.append(r)

    if len(matches) == 1:
        return matches[0]
    return None


def get_today_completed_for_phone(phone: str) -> tuple[int, str]:
    phone = normalize_phone(phone)
    riders = fetch_riders_cached()
    ds = fetch_delivery_status_cached()
    stats_map = build_today_stats_map(ds)

    for rider in riders:
        rider_phone = normalize_phone(rider.get("phoneNumber", ""))
        if rider_phone != phone:
            continue

        rider_name = rider.get("name", "") or ""
        real4 = last4_from_phone(rider.get("phoneNumber", "") or "")
        real_key = f"{norm_name(rider_name)}|{real4}"

        today_info = stats_map.get(real_key) or {}
        today_completed = int(today_info.get("complete") or 0)
        return today_completed, rider_name

    return 0, ""


def get_rider_roulette_status_by_name_login4(name: str, login4: str) -> Dict[str, Any]:
    rider = find_rider_by_name_login4(name, login4)
    if not rider:
        raise ValueError("조회 대상 라이더를 찾을 수 없습니다.")

    rider_name = rider.get("name", "") or ""
    rider_phone = normalize_phone(rider.get("phoneNumber", "") or "")

    ds = fetch_delivery_status_cached()
    stats_map = build_today_stats_map(ds)
    real4 = last4_from_phone(rider.get("phoneNumber", "") or "")
    api_key = f"{norm_name(rider_name)}|{real4}"
    today_completed = int((stats_map.get(api_key) or {}).get("complete") or 0)

    return roulette_db.get_status(rider_phone, rider_name, today_completed)


# -----------------------------
# Admin settings
# -----------------------------
def load_admin_settings() -> Dict[str, Any]:
    try:
        if not os.path.exists(ADMIN_SETTINGS_FILE):
            return {}
        with open(ADMIN_SETTINGS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def save_admin_settings(data: Dict[str, Any]) -> None:
    with open(ADMIN_SETTINGS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def get_admin_password() -> str:
    # 백업 복원된 비밀번호가 있으면 우선 적용, 없으면 Render 환경변수 ADMIN_PASSWORD 사용
    settings = load_admin_settings()
    pw = str(settings.get("ADMIN_PASSWORD") or "").strip()
    return pw or ADMIN_PASSWORD


def get_backup_payload() -> Dict[str, Any]:
    files = {}
    paths = [
        RIDERS_STORE, STATUS_STORE, DELIVERY_STATUS_STORE,
        Path(OVERRIDE_FILE), Path(LOGIN4_FILE), Path(PREVPLUS_FILE),
        Path(PLANNEDPLUS_FILE), Path(ATTENDANCE_FILE), Path(STICKER_FILE),
        Path(ADMIN_SETTINGS_FILE),
        STORE_DIR / "teams.json",
        STORE_DIR / "team_weekly_sets.json",
        STORE_DIR / "team_period_stats.json",
    ]

    for path in paths:
        try:
            if path.exists():
                files[path.name] = json.loads(path.read_text(encoding="utf-8"))
            else:
                files[path.name] = None
        except Exception as e:
            files[path.name] = {"_backup_error": str(e)}

    try:
        roulette_payload = roulette_db.export_json()
    except Exception as e:
        roulette_payload = {"_backup_error": str(e)}

    return {
        "ok": True,
        "service": "riderwelfare-render-backup",
        "backup_format": 2,
        "created_at": int(time.time()),
        "event_name": EVENT_NAME,
        "event_start_date": NMAX_START_DATE.isoformat(),
        "event_label": NMAX_LABEL,
        "store_dir": str(STORE_DIR),
        "admin": {
            # Render 환경변수 자체를 바꾸는 것은 아니고, 복원용 admin_settings.json에 저장됩니다.
            "ADMIN_PASSWORD": get_admin_password(),
        },
        "files": files,
        "roulette": roulette_payload,
    }


def restore_backup_payload(payload: Dict[str, Any]) -> Dict[str, Any]:
    if not isinstance(payload, dict):
        raise ValueError("백업 JSON 형식이 올바르지 않습니다.")

    # 안전장치: 현재 Render 데이터를 먼저 백업 파일로 남김
    backup_before = get_backup_payload()
    backup_dir = STORE_DIR / "restore_backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_name = f"before_restore_{int(time.time())}.json"
    (backup_dir / backup_name).write_text(json.dumps(backup_before, ensure_ascii=False, indent=2), encoding="utf-8")

    restored_files = []
    file_map = {
        "store_riders.json": RIDERS_STORE,
        "store_status.json": STATUS_STORE,
        "store_delivery_status.json": DELIVERY_STATUS_STORE,
        Path(OVERRIDE_FILE).name: Path(OVERRIDE_FILE),
        Path(LOGIN4_FILE).name: Path(LOGIN4_FILE),
        Path(PREVPLUS_FILE).name: Path(PREVPLUS_FILE),
        Path(PLANNEDPLUS_FILE).name: Path(PLANNEDPLUS_FILE),
        Path(ATTENDANCE_FILE).name: Path(ATTENDANCE_FILE),
        Path(STICKER_FILE).name: Path(STICKER_FILE),
        Path(ADMIN_SETTINGS_FILE).name: Path(ADMIN_SETTINGS_FILE),
        "teams.json": STORE_DIR / "teams.json",
        "team_weekly_sets.json": STORE_DIR / "team_weekly_sets.json",
        "team_period_stats.json": STORE_DIR / "team_period_stats.json",
    }

    files = payload.get("files") or {}
    if isinstance(files, dict):
        for name, data in files.items():
            if name not in file_map or data is None:
                continue
            # 백업 오류 객체는 복원하지 않음
            if isinstance(data, dict) and "_backup_error" in data:
                continue
            path = file_map[name]
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
            restored_files.append(name)

    # 예전/새 백업 모두 대응: admin 섹션이 있으면 관리자 비밀번호 복원 파일에 저장
    admin = payload.get("admin") or {}
    if isinstance(admin, dict) and str(admin.get("ADMIN_PASSWORD") or "").strip():
        settings = load_admin_settings()
        settings["ADMIN_PASSWORD"] = str(admin.get("ADMIN_PASSWORD")).strip()
        save_admin_settings(settings)
        if "admin_settings.json" not in restored_files:
            restored_files.append("admin_settings.json")

    roulette_result = None
    roulette_payload = payload.get("roulette")
    if isinstance(roulette_payload, dict) and "_backup_error" not in roulette_payload:
        try:
            roulette_result = roulette_db.import_json(roulette_payload)
        except Exception as e:
            roulette_result = {"ok": False, "error": str(e)}

    _riders_cache["ts"] = 0.0
    _riders_cache["data"] = None
    _status_cache.clear()
    _delivery_status_cache["ts"] = 0.0
    _delivery_status_cache["data"] = None

    return {
        "ok": True,
        "backup_before_restore": backup_name,
        "restored_files": restored_files,
        "roulette_result": roulette_result,
    }


# -----------------------------
# Admin helpers
# -----------------------------
def require_admin(request: Request) -> Optional[RedirectResponse]:
    if not request.session.get("is_admin"):
        return RedirectResponse("/admin-login", status_code=303)
    return None


def _safe_int(v: Any, default: int = 0) -> int:
    try:
        if v is None:
            return default
        if isinstance(v, float) and v != v:
            return default
        s = str(v).strip().replace(",", "")
        if s == "":
            return default
        return int(float(s))
    except Exception:
        return default


def restore_plus_from_workbook(file_bytes: bytes) -> Tuple[int, int]:
    """
    플러스/스티커 복원.
    우선순위:
    1) 엑셀의 login_key가 현재 라이더 목록에 존재하면 그대로 사용
    2) name + real4(배민 실제 뒷4자리)로 현재 라이더를 찾아 현재 login_key로 변환
    3) name + login4로 현재 login_key 확인

    이렇게 해야 로그인용 뒷4가 바뀌었거나 정렬이 달라도 다른 기사에게 플러스/스티커가 들어가지 않습니다.
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("업로드된 엑셀에 데이터가 없습니다.")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(headers)}

    if "name" not in idx:
        raise ValueError("필수 컬럼 없음: name")
    if "prev_plus" not in idx or "planned_plus" not in idx:
        raise ValueError("필수 컬럼 없음: prev_plus / planned_plus")

    riders = fetch_riders_cached()
    valid_login_keys: set[str] = set()
    by_name_real4: Dict[str, str] = {}
    by_name_login4: Dict[str, str] = {}

    for rr in riders:
        nm = rr.get("name") or ""
        nn = norm_name(nm)
        real4 = last4_from_phone(rr.get("phoneNumber") or "")
        login4, _, _ = get_login4_for_rider(rr)
        if not nn or not real4:
            continue
        login_key = f"{nn}|{login4}"
        valid_login_keys.add(login_key)
        by_name_real4[f"{nn}|{real4}"] = login_key
        by_name_login4[f"{nn}|{login4}"] = login_key

    prev_map = load_prevplus_map()
    planned_map = load_plannedplus_map()
    sticker_map = load_sticker_map()

    restored_rows = 0
    skipped_rows = 0

    def cell(row, name, default=""):
        if name not in idx:
            return default
        i = idx[name]
        return row[i] if i < len(row) else default

    for row in rows[1:]:
        if row is None:
            continue

        name_s = str(cell(row, "name", "") or "").strip()
        if not name_s:
            continue
        nn = norm_name(name_s)

        login_key = str(cell(row, "login_key", "") or "").strip()
        real4_s = str(cell(row, "real4", "") or "").strip()
        login4_s = str(cell(row, "login4", "") or "").strip()

        resolved_key = ""
        if login_key and login_key in valid_login_keys:
            resolved_key = login_key
        elif real4_s and re.fullmatch(r"\d{4}", real4_s):
            resolved_key = by_name_real4.get(f"{nn}|{real4_s}", "")
        if not resolved_key and login4_s and re.fullmatch(r"\d{4}", login4_s):
            resolved_key = by_name_login4.get(f"{nn}|{login4_s}", "")

        if not resolved_key:
            skipped_rows += 1
            continue

        prev_map[resolved_key] = max(-MAX_PLUS_INPUT, min(MAX_PLUS_INPUT, _safe_int(cell(row, "prev_plus", 0), 0)))
        planned_map[resolved_key] = max(-MAX_PLUS_INPUT, min(MAX_PLUS_INPUT, _safe_int(cell(row, "planned_plus", 0), 0)))

        if "sticker_attached" in idx:
            raw = str(cell(row, "sticker_attached", "") or "").strip().lower()
            sticker_map[resolved_key] = raw in ("1", "true", "y", "yes", "o", "ok", "부착", "부착자")

        restored_rows += 1

    save_prevplus_map(prev_map)
    save_plannedplus_map(planned_map)
    save_sticker_map(sticker_map)
    return restored_rows, skipped_rows


# -----------------------------
# Ingest endpoints
# -----------------------------
@app.post("/ingest/riders")
async def ingest_riders(request: Request):
    auth = _require_ingest(request)
    if auth:
        return auth

    payload = await request.json()
    riders = payload.get("riders")
    if not isinstance(riders, list):
        return JSONResponse({"ok": False, "error": "INVALID_PAYLOAD"}, status_code=400)

    valid_count = _valid_rider_identity_count(riders)
    current = _read_json(RIDERS_STORE, {"riders": []})
    old_riders = current.get("riders") or [] if isinstance(current, dict) else []
    old_count = _valid_rider_identity_count(old_riders) if isinstance(old_riders, list) else 0
    required = MIN_RIDERS_INGEST
    if old_count > 0:
        required = max(required, int(old_count * MIN_INGEST_RATIO))

    if valid_count < required:
        print(f"INGEST RIDERS BLOCKED new={valid_count} old={old_count} required={required}")
        return JSONResponse({
            "ok": False,
            "error": "RIDERS_SAFETY_BLOCK",
            "new_count": valid_count,
            "old_count": old_count,
            "required": required,
        }, status_code=409)

    _backup_store_file(RIDERS_STORE)
    _write_json(RIDERS_STORE, {"ts": time.time(), "riders": riders})
    _riders_cache["ts"] = 0.0
    _riders_cache["data"] = None
    return {"ok": True, "count": valid_count, "previous_count": old_count}


@app.post("/ingest/delivery-status")
async def ingest_delivery_status(request: Request):
    auth = _require_ingest(request)
    if auth:
        return auth

    payload = await request.json()
    data = payload.get("data")
    if not isinstance(data, dict):
        return JSONResponse({"ok": False, "error": "INVALID_PAYLOAD"}, status_code=400)

    new_count = _delivery_status_row_count(data)
    current = _read_json(DELIVERY_STATUS_STORE, {"data": {}})
    old_data = current.get("data") if isinstance(current, dict) else {}
    old_count = _delivery_status_row_count(old_data) if isinstance(old_data, dict) else 0
    required = MIN_DELIVERY_STATUS_INGEST
    if old_count > 0:
        required = max(required, int(old_count * MIN_INGEST_RATIO))

    if new_count < required:
        print(f"INGEST DELIVERY STATUS BLOCKED new={new_count} old={old_count} required={required}")
        return JSONResponse({
            "ok": False,
            "error": "DELIVERY_STATUS_SAFETY_BLOCK",
            "new_count": new_count,
            "old_count": old_count,
            "required": required,
        }, status_code=409)

    _backup_store_file(DELIVERY_STATUS_STORE)
    _write_json(DELIVERY_STATUS_STORE, {"ts": time.time(), "data": data})
    _delivery_status_cache["ts"] = 0.0
    _delivery_status_cache["data"] = None
    try:
        capture_team_snapshot(data)
    except Exception as e:
        print("TEAM SNAPSHOT INGEST ERROR:", e)
    return {"ok": True, "count": new_count, "previous_count": old_count}


@app.get("/ingest/ranges")
def ingest_ranges(request: Request):
    auth = _require_ingest(request)
    if auth:
        return auth

    store = _read_json(RIDERS_STORE, {"riders": []})
    riders = store.get("riders") or []
    if not isinstance(riders, list) or not riders:
        return {"ok": False, "error": "NO_RIDERS"}

    today = operation_date()
    requested_ranges = set()

    # 등급 조회 범위: 해당 정확한 키가 저장돼 있지 않을 때만 요청합니다.
    for rr in riders:
        if not isinstance(rr, dict):
            continue
        ph = rr.get("phoneNumber") or ""
        real4 = last4_from_phone(ph)
        if not real4:
            continue

        login4, _, _ = get_login4_for_rider(rr)
        eff_join, _ = get_effective_join_date_by_login_key(rr, login4)
        cur_start, cur_end_incl = current_period(eff_join, today)
        cur_from, cur_to = period_to_from_to(cur_start, cur_end_incl)

        prev_end_incl = cur_start - timedelta(days=1)
        prev_m = date(cur_start.year, cur_start.month, 1) + relativedelta(months=-1)
        prev_start = clamp_day(prev_m.year, prev_m.month, eff_join.day)
        prev_from, prev_to = period_to_from_to(prev_start, prev_end_incl)

        if cur_from <= cur_to:
            requested_ranges.add((cur_from.isoformat(), cur_to.isoformat(), False))
        if prev_from <= prev_to:
            requested_ranges.add((prev_from.isoformat(), prev_to.isoformat(), False))

    # NMAX 누적 범위도 날짜 키가 바뀔 때 하루 한 번만 새로 수집됩니다.
    nmax_from = NMAX_START_DATE
    nmax_to = operation_date() - timedelta(days=1)
    if nmax_from <= nmax_to:
        requested_ranges.add((nmax_from.isoformat(), nmax_to.isoformat(), False))

    # 팀 과거 조회용 하루 단위 기록입니다. historyMap이 없을 때만 요청합니다.
    try:
        yesterday = operation_date() - timedelta(days=1)
        starts = []
        for team in load_teams():
            if not team.get("active", True):
                continue
            td = safe_date_parse(str(team.get("team_start_date") or ""))
            if td is None:
                td = safe_date_parse(str(team.get("created_at") or "")[:10])
            if td is not None:
                starts.append(td)
        if starts:
            d = max(min(starts), yesterday - timedelta(days=119))
            while d <= yesterday:
                requested_ranges.add((d.isoformat(), d.isoformat(), True))
                d += timedelta(days=1)
    except Exception as e:
        print("TEAM RANGE BUILD ERROR:", e)

    status_store = _read_json(STATUS_STORE, {}) or {}
    out = []
    for from_d, to_d, require_history in sorted(requested_ranges):
        key = f"{from_d}_{to_d}"
        saved = status_store.get(key) if isinstance(status_store, dict) else None

        if require_history:
            is_saved = (
                isinstance(saved, dict)
                and isinstance(saved.get("historyMap"), dict)
                and len(saved.get("historyMap") or {}) > 0
            )
        else:
            if isinstance(saved, dict) and "completeMap" in saved:
                is_saved = isinstance(saved.get("completeMap"), dict)
            else:
                # 구버전 completeMap 자체 저장 형식도 완료된 범위로 인정합니다.
                is_saved = isinstance(saved, dict) and len(saved) > 0

            # 전날로 끝나는 등급 범위는 새 운영일(06:00) 이후 최소 한 번 다시 수집합니다.
            # 배민 통계가 마감되기 전에 저장된 값이 남아 마지막 날 건수가 빠지는 문제를 방지합니다.
            try:
                range_to = date.fromisoformat(to_d)
                yesterday = operation_date() - timedelta(days=1)
                updated_raw = str(saved.get("updated_at") or "") if isinstance(saved, dict) else ""
                updated_dt = datetime.fromisoformat(updated_raw) if updated_raw else None
                refreshed_today = bool(updated_dt and updated_dt.astimezone(KST).date() == now_kst().date())
                if range_to == yesterday and not refreshed_today:
                    is_saved = False
            except Exception:
                # 갱신시각을 판별할 수 없으면 최근 마감 범위를 안전하게 다시 받습니다.
                try:
                    if date.fromisoformat(to_d) == operation_date() - timedelta(days=1):
                        is_saved = False
                except Exception:
                    pass

        if not is_saved:
            out.append({"fromDate": from_d, "toDate": to_d})

    return {"ok": True, "ranges": out, "count": len(out)}


@app.post("/ingest/status")
async def ingest_status(request: Request):
    auth = _require_ingest(request)
    if auth:
        return auth

    payload = await request.json()
    from_d = payload.get("fromDate")
    to_d = payload.get("toDate")
    complete_map = payload.get("completeMap")
    history_map = payload.get("historyMap") or {}

    if not (isinstance(from_d, str) and isinstance(to_d, str) and isinstance(complete_map, dict)):
        return JSONResponse({"ok": False, "error": "INVALID_PAYLOAD"}, status_code=400)
    if not isinstance(history_map, dict):
        history_map = {}

    all_status = _read_json(STATUS_STORE, {})
    key = f"{from_d}_{to_d}"
    existing = all_status.get(key) if isinstance(all_status, dict) else None
    existing_map = existing.get("completeMap") if isinstance(existing, dict) else None
    old_count = len(existing_map) if isinstance(existing_map, dict) else 0
    new_count = len(complete_map)
    required = MIN_RIDERS_INGEST
    if old_count > 0:
        required = max(required, int(old_count * MIN_INGEST_RATIO))
    if new_count < required:
        print(f"INGEST STATUS BLOCKED key={key} new={new_count} old={old_count} required={required}")
        return JSONResponse({
            "ok": False, "error": "STATUS_SAFETY_BLOCK",
            "key": key, "new_count": new_count, "old_count": old_count, "required": required,
        }, status_code=409)

    _backup_store_file(STATUS_STORE)
    all_status[key] = {
        "completeMap": complete_map,
        "historyMap": history_map,
        "updated_at": now_kst().isoformat(),
    }
    _write_json(STATUS_STORE, all_status)

    _status_cache.pop(key, None)
    return {
        "ok": True,
        "key": key,
        "count": len(complete_map),
        "history_count": len(history_map),
    }


# -----------------------------
# Main routes
# -----------------------------
@app.get("/", response_class=HTMLResponse)
def home():
    if not store_ready():
        return not_ready_page()

    body = """
    <div style="background:#fff; border:1px solid #e8e8e8; border-radius:16px; padding:16px; max-width:520px; margin:0 auto;">
      <h2 style="margin:0 0 6px 0;">라웰 등급 조회</h2>
      <div style="color:#666; margin-bottom:14px;">이름 + <b>로그인용 뒷4자리</b>로 조회합니다. (관리자가 설정)</div>

      <form method="post" action="/check">
        <div style="margin-bottom:12px;">
          <label style="display:block; margin-bottom:6px;">이름</label>
          <input name="name" autocomplete="name"
                 style="font-size:18px; padding:12px; width:100%; box-sizing:border-box; border:1px solid #ddd; border-radius:12px;"
                 required />
        </div>

        <div style="margin-bottom:14px;">
          <label style="display:block; margin-bottom:6px;">로그인용 뒷 4자리</label>
          <input name="login4" inputmode="numeric" pattern="\\d{4}" maxlength="4"
                 style="font-size:18px; padding:12px; width:180px; border:1px solid #ddd; border-radius:12px;"
                 required />
        </div>

        <button type="submit"
                style="font-size:18px; padding:12px 16px; border:none; border-radius:12px; background:#111; color:#fff; width:100%;">
          조회
        </button>
      </form>

      <div style="display:flex; justify-content:space-between; margin-top:14px; font-size:14px;">
        <a href="/dashboard" style="text-decoration:none; color:#111;">관리자: 전체현황</a>
        <a href="/admin" style="text-decoration:none; color:#666;">관리자 도움말</a>
      </div>

      <div style="color:#888; margin-top:12px; font-size:13px;">
        * 완료건수는 ‘어제까지’ 반영됩니다. (등급용)<br/>
        * 출석체크는 ‘오늘 완료건수’ 기준입니다. (실시간)<br/>
        * 룰렛/출석은 라웰 운영일 기준(06:00~다음날 05:59)으로 초기화됩니다.
      </div>
    </div>
    """
    return html_page("라웰 등급 조회", body)


@app.get("/check")
def check_get_redirect():
    return RedirectResponse(url="/", status_code=302)


@app.get("/admin", response_class=HTMLResponse)
def admin_help():
    cycle = get_operational_cycle_bounds()
    prev_cycle = get_previous_operational_cycle_bounds()

    body = f"""
    <div style="background:#fff; border:1px solid #e8e8e8; border-radius:16px; padding:16px; max-width:900px; margin:0 auto;">
      <h2 style="margin:0 0 6px 0;">관리자</h2>
      <div style="color:#666; margin-bottom:12px;">
        이 서버(Render)는 배민 API를 직접 호출하지 않습니다.<br/>
        PC에서 collector.py가 /ingest 로 데이터를 업로드하면 조회가 됩니다.
      </div>

      <div style="background:#f7f7f7; border-radius:12px; padding:12px; font-size:14px; line-height:1.6;">
        <div><b>Render 설정</b></div>
        <div>- Settings → Environment → <b>INGEST_TOKEN</b> 등록</div>
        <div>- Start Command: <span style="font-family: ui-monospace;">uvicorn main:app --host 0.0.0.0 --port $PORT</span></div>
      </div>

      <div style="margin-top:14px; background:#fff8e8; border:1px solid #ffe1a8; border-radius:12px; padding:12px; font-size:14px; line-height:1.6;">
        <div><b>NMAX 누적 집계</b></div>
        <div>- 개인조회 화면에 <b>{NMAX_LABEL}</b> 누적 완료건수를 보여줍니다.</div>
        <div>- 단, PC에서 그 기간({NMAX_START_DATE.isoformat()} ~ 어제) 범위를 <b>한 번이라도</b> 업로드해야 숫자가 뜹니다.</div>
      </div>

      <div style="margin-top:14px; background:#eef6ff; border:1px solid #cfe3ff; border-radius:12px; padding:12px; font-size:14px; line-height:1.6;">
        <div><b>플러스 백업 / 복원</b></div>
        <div>- 관리자 대시보드에서 <b>플러스 백업 엑셀</b> 다운로드 가능</div>
        <div>- 코드 업데이트 후 기존 <b>dashboard.xlsx</b> 또는 <b>plus-backup.xlsx</b> 업로드 시</div>
        <div>&nbsp;&nbsp;→ <b>이전등급 플러스(prev_plus)</b>, <b>예정등급 플러스(planned_plus)</b> 자동 복원</div>
      </div>

      <div style="margin-top:14px; background:#f5f7ff; border:1px solid #d9e1ff; border-radius:12px; padding:12px; font-size:14px; line-height:1.7;">
        <div><b>룰렛 운영주차</b></div>
        <div>- 현재 운영주차: <b>{cycle["display_start"]}</b> ~ <b>{cycle["display_end"]}</b></div>
        <div>- 직전 마감 주차: <b>{prev_cycle["display_start"]}</b> ~ <b>{prev_cycle["display_end"]}</b></div>
        <div>- 지급 CSV: <b>/admin/api/roulette/payout.csv</b> (기본은 직전 마감 주차)</div>
      </div>

      <div style="margin-top:14px; background:#f7fff5; border:1px solid #cfe8c8; border-radius:12px; padding:12px; font-size:14px; line-height:1.7;">
        <div><b>전체 백업 / 전체 복원</b></div>
        <div>- 룰렛 확률, 룰렛 당첨 전체내역, 예정플러스, 이전플러스, 스티커, 입사일, 로그인용 뒷4자리, 관리자 비밀번호까지 포함합니다.</div>
        <div style="margin-top:10px; display:flex; gap:8px; flex-wrap:wrap; align-items:center;">
          <a href="/admin/backup/export" style="text-decoration:none; color:#111; font-weight:900; padding:10px 12px; border:1px solid #ddd; border-radius:10px; background:#fff;">전체 백업 다운로드</a>
          <form method="post" action="/admin/backup/import" enctype="multipart/form-data" style="display:flex; gap:8px; flex-wrap:wrap; align-items:center; margin:0;">
            <input type="file" name="file" accept=".json" required />
            <button type="submit" style="padding:10px 12px; border:none; border-radius:10px; background:#111; color:#fff; font-weight:900; cursor:pointer;">전체 백업 업로드 복원</button>
          </form>
        </div>
      </div>

      <div style="margin-top:14px;">
        <a href="/" style="text-decoration:none; color:#111;">← 조회 화면으로</a>
      </div>
    </div>
    """
    return html_page("관리자", body)


@app.post("/attendance-check")
def attendance_check(request: Request, name: str = Form(...), login4: str = Form(...)):
    if not store_ready():
        return not_ready_page()

    ip = request.client.host if request.client else "unknown"
    if not rate_limit(ip):
        return RedirectResponse("/", status_code=303)

    name_in = norm_name(name)
    login4 = (login4 or "").strip()

    if not re.fullmatch(r"\d{4}", login4):
        return RedirectResponse("/", status_code=303)

    riders = fetch_riders_cached()
    candidates = [r for r in riders if norm_name(r.get("name", "")) == name_in]
    matches: List[Dict[str, Any]] = []
    for r in candidates:
        l4, _, _ = get_login4_for_rider(r)
        if l4 == login4:
            matches.append(r)

    if len(matches) != 1:
        return RedirectResponse("/", status_code=303)

    rider = matches[0]
    rider_login4, rider_real4, _ = get_login4_for_rider(rider)
    eff_join_date, _ = get_effective_join_date_by_login_key(rider, rider_login4)

    today = operation_date()
    cur_start, cur_end_incl = current_period(eff_join_date, today)
    login_key = f"{name_in}|{rider_login4}"

    attendance_rollover_if_needed(login_key, cur_start, cur_end_incl)

    ds = fetch_delivery_status_cached()
    stats_map = build_today_stats_map(ds)
    today_completed = int((stats_map.get(f"{name_in}|{rider_real4}") or {}).get("complete") or 0)

    if today_completed < ATTENDANCE_MIN_TODAY_COMPLETE:
        return RedirectResponse(f"/check-result?name={name}&login4={login4}", status_code=303)

    if attendance_is_checked_today(login_key, today, cur_start, cur_end_incl):
        return RedirectResponse(f"/check-result?name={name}&login4={login4}", status_code=303)

    ok = attendance_mark_today(login_key, today, cur_start, cur_end_incl)
    if not ok:
        return RedirectResponse(f"/check-result?name={name}&login4={login4}", status_code=303)

    planned_plus = get_plannedplus(login_key)
    set_plannedplus(login_key, planned_plus + ATTENDANCE_BONUS_PER_DAY)

    return RedirectResponse(f"/check-result?name={name}&login4={login4}", status_code=303)


@app.post("/roulette-spin")
def roulette_spin_page(request: Request, name: str = Form(...), login4: str = Form(...)):
    if not store_ready():
        return not_ready_page()

    ip = request.client.host if request.client else "unknown"
    if not rate_limit(ip):
        return RedirectResponse("/", status_code=303)

    rider = find_rider_by_name_login4(name, login4)
    if not rider:
        return RedirectResponse("/", status_code=303)

    rider_name = rider.get("name", "") or ""
    rider_phone = normalize_phone(rider.get("phoneNumber", "") or "")
    real4 = last4_from_phone(rider.get("phoneNumber", "") or "")
    ds = fetch_delivery_status_cached()
    stats_map = build_today_stats_map(ds)
    api_key = f"{norm_name(rider_name)}|{real4}"
    today_completed = int((stats_map.get(api_key) or {}).get("complete") or 0)

    try:
        result = roulette_db.spin(rider_phone, rider_name, today_completed)
        reward = int(result.get("reward") or 0)
        return RedirectResponse(
            f"/check-result?name={name}&login4={login4}&roulette_reward={reward}",
            status_code=303
        )
    except Exception:
        return RedirectResponse(
            f"/check-result?name={name}&login4={login4}&roulette_error=1",
            status_code=303
        )


@app.get("/check-result", response_class=HTMLResponse)
def check_result_get(request: Request, name: str, login4: str, roulette_reward: int = 0, roulette_error: int = 0):
    return check(request, name=name, login4=login4, roulette_reward=roulette_reward, roulette_error=roulette_error)


@app.post("/check", response_class=HTMLResponse)
def check(
    request: Request,
    name: str = Form(...),
    login4: str = Form(...),
    roulette_reward: int = 0,
    roulette_error: int = 0,
):
    if not store_ready():
        return not_ready_page()

    ip = request.client.host if request.client else "unknown"
    if not rate_limit(ip):
        body = """
        <div style="background:#fff; border:1px solid #e8e8e8; border-radius:16px; padding:16px; max-width:520px; margin:0 auto;">
          <h3 style="margin-top:0;">요청이 너무 많습니다</h3>
          <div style="color:#666;">잠시 후 다시 시도해주세요.</div>
          <div style="margin-top:12px;"><a href="/" style="text-decoration:none; color:#111;">← 뒤로</a></div>
        </div>
        """
        return html_page("제한됨", body)

    name_in = norm_name(name)
    login4 = (login4 or "").strip()

    if not re.fullmatch(r"\d{4}", login4):
        body = """
        <div style="background:#fff; border:1px solid #e8e8e8; border-radius:16px; padding:16px; max-width:520px; margin:0 auto;">
          <h3 style="margin-top:0;">입력 오류</h3>
          <div style="color:#666;">뒷 4자리는 숫자 4자리로 입력해주세요.</div>
          <div style="margin-top:12px;"><a href="/" style="text-decoration:none; color:#111;">← 뒤로</a></div>
        </div>
        """
        return html_page("입력 오류", body)

    riders = fetch_riders_cached()
    candidates = [r for r in riders if norm_name(r.get("name", "")) == name_in]
    matches: List[Dict[str, Any]] = []
    for r in candidates:
        l4, _, _ = get_login4_for_rider(r)
        if l4 == login4:
            matches.append(r)

    if not matches:
        body = f"""
        <div style="background:#fff; border:1px solid #e8e8e8; border-radius:16px; padding:16px; max-width:520px; margin:0 auto;">
          <h3 style="margin-top:0;">조회 결과 없음</h3>
          <div style="color:#666;">입력: <b>{name}</b> / <b>{login4}</b></div>
          <div style="color:#888; margin-top:8px; font-size:13px;">이름(띄어쓰기/철자) 또는 로그인용 뒷4를 확인해주세요.</div>
          <div style="margin-top:12px;"><a href="/" style="text-decoration:none; color:#111;">← 다시 조회</a></div>
        </div>
        """
        return html_page("조회 결과 없음", body)

    if len(matches) >= 2:
        body = """
        <div style="background:#fff; border:1px solid #e8e8e8; border-radius:16px; padding:16px; max-width:520px; margin:0 auto;">
          <h3 style="margin-top:0;">동일 정보 다수</h3>
          <div style="color:#666;">동일 이름/로그인용뒷4가 여러 명입니다. 관리자에게 문의해주세요.</div>
          <div style="margin-top:12px;"><a href="/" style="text-decoration:none; color:#111;">← 뒤로</a></div>
        </div>
        """
        return html_page("동일 정보 다수", body)

    rider = matches[0]
    rider_login4, rider_real4, _ = get_login4_for_rider(rider)
    eff_join_date, join_src = get_effective_join_date_by_login_key(rider, rider_login4)

    today = operation_date()
    cur_start, cur_end_incl = current_period(eff_join_date, today)
    cur_from, cur_to = period_to_from_to(cur_start, cur_end_incl)

    prev_end_incl = cur_start - timedelta(days=1)
    prev_m = date(cur_start.year, cur_start.month, 1) + relativedelta(months=-1)
    prev_start = clamp_day(prev_m.year, prev_m.month, eff_join_date.day)
    prev_from, prev_to = period_to_from_to(prev_start, prev_end_incl)

    api_key = f"{name_in}|{rider_real4}"

    if cur_from > cur_to:
        cur_completed_raw = 0
    else:
        cmap_cur = fetch_status_complete_map_cached(cur_from, cur_to)
        cur_completed_raw = int(cmap_cur.get(api_key, 0))

    if prev_from > prev_to:
        prev_completed_raw = 0
    else:
        cmap_prev = fetch_status_complete_map_cached(prev_from, prev_to)
        prev_completed_raw = int(cmap_prev.get(api_key, 0))

    login_key = f"{name_in}|{rider_login4}"
    request.session["rider_login_key"] = login_key
    rolled_bonus, prevplus_after, rolled = attendance_rollover_if_needed(login_key, cur_start, cur_end_incl)

    prev_plus = get_prevplus(login_key)
    planned_plus_base = get_plannedplus(login_key)
    sticker_bonus = get_sticker_bonus(login_key)
    planned_plus = planned_plus_base + sticker_bonus

    ds = fetch_delivery_status_cached()
    today_stats_map = build_today_stats_map(ds)
    today_info = today_stats_map.get(api_key) or {
        "complete": 0,
        "reject": 0,
        "cancel": 0,
        "status_desc": "-",
        **calc_bad_ratio(0, 0, 0),
    }

    today_completed = int(today_info["complete"])
    today_rejected = int(today_info["reject"])
    today_canceled = int(today_info["cancel"])
    today_status_desc = str(today_info["status_desc"])
    today_total_attempt = int(today_info["total"])
    today_bad_ratio = float(today_info["ratio"])
    ratio_bg = str(today_info["bg"])
    ratio_fg = str(today_info["fg"])
    ratio_label = str(today_info["label"])

    att_count = attendance_count_in_period(login_key, cur_start, cur_end_incl)
    att_checked_today = attendance_is_checked_today(login_key, today, cur_start, cur_end_incl)
    attendance_enabled = (today_completed >= ATTENDANCE_MIN_TODAY_COMPLETE) and (not att_checked_today)

    planned_total = cur_completed_raw + planned_plus
    prev_total = prev_completed_raw + prev_plus

    planned_grade = grade_from_total(planned_total)
    current_grade = grade_from_total(prev_total)

    nxt, remain = next_grade_target(planned_total)

    join_note = "관리자 설정" if join_src == "override" else "배민 입사일"

    pcx_from = NMAX_START_DATE
    pcx_to = operation_date() - timedelta(days=1)
    pcx_text = "데이터 없음(업로드 필요)"
    if pcx_from <= pcx_to and has_status_range(pcx_from, pcx_to):
        cmap_pcx = fetch_status_complete_map_cached(pcx_from, pcx_to)
        pcx_text = f"{int(cmap_pcx.get(api_key, 0))}건"

    rollover_note = ""
    if rolled_bonus:
        rollover_note = f"<div style='margin-top:8px;color:#111;font-size:13px;'><b>기간 이관:</b> 이전기간 현재플러스 {rolled_bonus}건이 현재등급(이전) 플러스로 자동 이관되고 현재플러스는 초기화되었습니다.</div>"

    # 룰렛 상태
    roulette_status = roulette_db.get_status(
        normalize_phone(rider.get("phoneNumber", "") or ""),
        rider.get("name", "") or "",
        today_completed,
    )

    roulette_alert_html = ""
    if roulette_reward:
        roulette_alert_html = f"""
        <div style="margin-top:12px; padding:14px; border-radius:14px; background:#fff8db; border:1px solid #f1d27a;">
          <div style="font-size:18px; font-weight:900; color:#111;">🎉 룰렛 당첨 {int(roulette_reward):,}원</div>
        </div>
        """
    elif roulette_error:
        roulette_alert_html = """
        <div style="margin-top:12px; padding:14px; border-radius:14px; background:#fdecec; border:1px solid #f2b8b8;">
          <div style="font-size:15px; font-weight:900; color:#b91c1c;">룰렛 처리 중 오류가 발생했거나 가능 횟수가 없습니다.</div>
        </div>
        """

    recent_li = ""
    for item in roulette_status.get("recent_spins", []) or []:
        created_at = int(item.get("created_at") or 0)
        ts_text = "-"
        if created_at:
            try:
                ts_text = kst_from_epoch(created_at).strftime("%m/%d %H:%M")
            except Exception:
                ts_text = "-"
        recent_li += f"<li style='margin-bottom:4px;'>{ts_text} / {int(item.get('reward_amount') or 0):,}원 / {int(item.get('segment_value') or 0)}건 구간</li>"

    if not recent_li:
        recent_li = "<li>최근 당첨 내역 없음</li>"

    cycle = roulette_status.get("cycle") or {}
    can_spin = bool(roulette_status.get("can_spin"))
    roulette_btn_bg = "#111" if can_spin else "#bbb"
    roulette_btn_cursor = "pointer" if can_spin else "not-allowed"

    body = f"""
    <div style="background:#fff; border:1px solid #e8e8e8; border-radius:16px; padding:16px; max-width:920px; margin:0 auto;">
      <h2 style="margin:0 0 6px 0;">등급 조회 결과</h2>

      <div style="color:#888; font-size:13px; margin-top:6px;">
        기준일(입사일): <b>{eff_join_date}</b> ({join_note})
      </div>

      <div style="margin-top:12px; padding:12px; border:1px solid #eee; border-radius:14px; background:#fcfcfc;">
        <div style="font-size:18px;"><b>{rider.get('name','')}</b> 님</div>
        <div style="color:#777; margin-top:6px;">휴대폰: {mask_phone(rider.get('phoneNumber',''))}</div>
      </div>

      {roulette_alert_html}

      <div style="display:flex; gap:12px; margin-top:12px; flex-wrap:wrap;">
        <div style="flex:1; min-width:250px; padding:12px; border-radius:12px; border:1px solid #eee; background:#fff;">
          <div style="color:#777; font-size:13px;">현재등급(이전기간)</div>
          <div style="font-size:32px; font-weight:900; line-height:1.1;">{current_grade}</div>
          <div style="font-size:12px; color:#999; margin-top:6px;">
            정책기간: {prev_start} ~ {prev_end_incl}<br/>
            반영기간(업로드): {prev_from} ~ {prev_to}<br/>
            완료 {prev_completed_raw}건 + 이전등급플러스 {prev_plus}건 = <b>{prev_total}</b>건
          </div>
        </div>

        <div style="flex:1; min-width:250px; padding:12px; border-radius:12px; border:1px solid #eee; background:#fff;">
          <div style="color:#777; font-size:13px;">예정등급(현재기간)</div>
          <div style="font-size:32px; font-weight:900; line-height:1.1;">{planned_grade}</div>
          <div style="font-size:12px; color:#999; margin-top:6px;">
            정책기간: {cur_start} ~ {cur_end_incl}<br/>
            반영기간(업로드): {cur_from} ~ {cur_to}<br/>
            완료 {cur_completed_raw}건 + 예정등급플러스 {planned_plus}건 = <b>{planned_total}</b>건
          </div>
        </div>
      </div>

      <div style="margin-top:12px; padding:12px; border:1px solid #eee; border-radius:14px; background:#fff;">
        <div style="display:flex; justify-content:space-between; gap:16px; flex-wrap:wrap;">
          <div style="color:#666;">
            다음등급: <b>{(nxt or '-')}</b> / 남은건수: <b>{(remain if remain is not None else '-')}</b>
            <div style="color:#999; font-size:12px; margin-top:6px;">* 다음등급/남은건수는 “현재기간 완료건수(어제까지)” 기준입니다.</div>
          </div>

          <div style="color:#111; font-weight:900; min-width:300px;">
            <div>오늘완료(실시간): {today_completed}건</div>
            <div style="margin-top:4px; font-size:14px; font-weight:700; color:#666;">
              거절: {today_rejected}건 / 취소: {today_canceled}건
            </div>
            <div style="margin-top:8px;">
              <span style="
                display:inline-block;
                padding:6px 10px;
                border-radius:999px;
                background:{ratio_bg};
                color:{ratio_fg};
                font-size:13px;
                font-weight:900;
                border:1px solid rgba(0,0,0,0.06);
              ">
                거절+취소율 {today_bad_ratio}% ({ratio_label})
              </span>
            </div>
            <div style="color:#777; font-weight:600; font-size:12px; margin-top:6px;">
              계산기준: 완료 {today_completed}건 / 반영 거절+취소 {today_rejected + today_canceled}건
            </div>
            <div style="color:#777; font-weight:600; font-size:12px; margin-top:4px;">운행상태: {today_status_desc}</div>
          </div>
        </div>

        <div style="margin-top:10px; border-top:1px dashed #eee; padding-top:10px;">
          <div style="font-weight:900; margin-bottom:6px;">출석체크</div>
          <div style="color:#666; font-size:13px; line-height:1.6;">
            - 조건: <b>오늘 완료 {ATTENDANCE_MIN_TODAY_COMPLETE}건 이상</b>이면 출석체크 가능 (운행중/운행종료 무관)<br/>
            - 출석 1회당 <b>+{ATTENDANCE_BONUS_PER_DAY}건</b> (예정등급에 즉시 반영)<br/>
            - 기간 종료 시 <b>현재기간 플러스 전체</b>가 현재등급(이전) 플러스로 자동 이관되고 현재플러스는 초기화됩니다
          </div>

          <div style="margin-top:10px; display:flex; align-items:center; gap:12px; flex-wrap:wrap;">
            <div style="color:#111; font-weight:900;">
                이번기간 출석: {att_count}회 (출석보너스 {att_count * ATTENDANCE_BONUS_PER_DAY}건)
                / 스티커보너스 {sticker_bonus}건
                / 현재플러스 합계 {planned_plus}건
            </div>
            <form method="post" action="/attendance-check" style="margin:0;">
              <input type="hidden" name="name" value="{name}" />
              <input type="hidden" name="login4" value="{login4}" />
              <button type="submit"
                {'disabled' if not attendance_enabled else ''}
                style="padding:10px 14px; border:none; border-radius:12px; background:{'#111' if attendance_enabled else '#bbb'}; color:#fff; font-weight:900; cursor:{'pointer' if attendance_enabled else 'not-allowed'};">
                {'오늘 출석체크 완료' if att_checked_today else '출석체크 (+5)'}
              </button>
            </form>
            <div style="color:#999; font-size:12px;">
              {'' if attendance_enabled else ('오늘완료가 부족합니다' if today_completed < ATTENDANCE_MIN_TODAY_COMPLETE else '이미 오늘 출석체크 완료')}
            </div>
          </div>
          {rollover_note}
        </div>
      </div>

    <div style="margin-top:12px; padding:14px; border:1px solid #eee; border-radius:18px; background:linear-gradient(180deg,#ffffff 0%,#fcfcff 100%);">
  <div style="display:flex; justify-content:space-between; align-items:flex-start; gap:12px; flex-wrap:wrap;">
    <div>
      <div style="font-weight:900; font-size:18px; color:#111;">룰렛</div>
      <div style="color:#666; font-size:13px; line-height:1.7; margin-top:6px;">
        운영주차: <b>{cycle.get("display_start", "-")}</b> ~ <b>{cycle.get("display_end", "-")}</b><br/>
        오늘 완료 {roulette_status.get("spin_unit", 10)}건당 1회<br/>
        다음 룰렛 구간: <b><span id="nextSegmentText">{int(roulette_status.get("next_segment_value") or 0)}</span>건</b>
      </div>
    </div>

    <div style="min-width:180px; padding:10px 12px; border-radius:14px; background:#f8fafc; border:1px solid #edf2f7;">
      <div style="font-size:12px; color:#666;">이번 운영주차 누적당첨금</div>
      <div id="weeklyTotalText" style="font-size:24px; font-weight:900; color:#111; margin-top:4px;">
        {int(roulette_status.get("weekly_total") or 0):,}원
      </div>
    </div>
  </div>

  <div style="display:flex; gap:10px; flex-wrap:wrap; margin-top:14px;">
    <div style="flex:1; min-width:120px; padding:10px 12px; border-radius:14px; background:#fafafa; border:1px solid #eee;">
      <div style="font-size:12px; color:#777;">가능 횟수</div>
      <div id="remainCountText" style="font-size:22px; font-weight:900; color:#111;">{roulette_status.get("remain_count", 0)}회</div>
    </div>

    <div style="flex:1; min-width:120px; padding:10px 12px; border-radius:14px; background:#fafafa; border:1px solid #eee;">
      <div style="font-size:12px; color:#777;">사용 횟수</div>
      <div id="usedCountText" style="font-size:22px; font-weight:900; color:#111;">{roulette_status.get("used_count", 0)}회</div>
    </div>

    <div style="flex:1; min-width:120px; padding:10px 12px; border-radius:14px; background:#fafafa; border:1px solid #eee;">
      <div style="font-size:12px; color:#777;">총 가능</div>
      <div id="eligibleCountText" style="font-size:22px; font-weight:900; color:#111;">{roulette_status.get("eligible_count", 0)}회</div>
    </div>
  </div>

  <div style="margin-top:14px; display:flex; gap:12px; flex-wrap:wrap; align-items:flex-start;">
    <button
      type="button"
      id="openRouletteBtn"
      {'disabled' if not can_spin else ''}
      onclick="openRouletteModal()"
      style="
        padding:14px 18px;
        border:none;
        border-radius:16px;
        background:{roulette_btn_bg};
        color:#fff;
        font-weight:900;
        font-size:17px;
        cursor:{roulette_btn_cursor};
        min-width:150px;
        box-shadow:0 8px 18px rgba(17,17,17,0.18);
      ">
      {'룰렛 돌리기' if can_spin else '룰렛 불가'}
    </button>

    <div style="min-width:280px; flex:1;">
      <div style="font-weight:900; margin-bottom:6px;">최근 당첨 내역</div>
      <ul id="recentRouletteList" style="margin:0; padding-left:18px; color:#666; font-size:13px; line-height:1.7;">
        {recent_li}
      </ul>
    </div>
  </div>
</div>

<div id="rouletteModal" style="
  display:none;
  position:fixed;
  inset:0;
  background:rgba(15,23,42,0.62);
  z-index:9999;
  padding:14px;
  box-sizing:border-box;
  backdrop-filter:blur(2px);
">
  <div style="
    max-width:430px;
    margin:0 auto;
    min-height:100%;
    display:flex;
    align-items:center;
    justify-content:center;
  ">
    <div style="
      width:100%;
      border-radius:28px;
      padding:18px 16px 20px 16px;
      box-sizing:border-box;
      position:relative;
      overflow:hidden;
      background:
        radial-gradient(circle at top, rgba(255,255,255,0.98) 0%, rgba(250,250,255,0.98) 55%, rgba(245,247,255,0.98) 100%);
      box-shadow:0 24px 60px rgba(0,0,0,0.28);
      border:1px solid rgba(255,255,255,0.5);
    ">
      <div style="
        position:absolute;
        inset:auto -60px -60px auto;
        width:180px;
        height:180px;
        border-radius:50%;
        background:radial-gradient(circle, rgba(99,102,241,0.18), rgba(99,102,241,0));
      "></div>

      <button type="button" onclick="closeRouletteModal()" style="
        position:absolute;
        right:12px;
        top:12px;
        border:none;
        background:#eef2f7;
        width:38px;
        height:38px;
        border-radius:999px;
        font-size:22px;
        cursor:pointer;
        color:#111;
      ">×</button>

      <div style="font-size:28px; font-weight:900; color:#111; letter-spacing:-0.02em;">행운 룰렛</div>
      <div style="color:#667085; font-size:14px; margin-top:6px;">
        서버 당첨 결과에 맞춰 실제 회전 후 멈춥니다
      </div>

      <div id="rouletteConfetti" style="position:relative; height:0; overflow:visible;"></div>

      <div style="margin-top:18px; display:flex; justify-content:center;">
        <div style="position:relative; width:310px; height:310px;">
          <div style="
            position:absolute;
            left:50%;
            top:-8px;
            transform:translateX(-50%);
            width:0;
            height:0;
            border-left:18px solid transparent;
            border-right:18px solid transparent;
            border-top:34px solid #ef4444;
            z-index:4;
            filter:drop-shadow(0 3px 3px rgba(0,0,0,0.16));
          "></div>

          <div style="
            position:absolute;
            inset:0;
            border-radius:50%;
            background:radial-gradient(circle, rgba(255,255,255,0.6), rgba(255,255,255,0));
            z-index:0;
            transform:scale(1.07);
          "></div>

          <div id="rouletteWheel" style="
            width:310px;
            height:310px;
            border-radius:50%;
            border:12px solid #f6e9e9;
            box-sizing:border-box;
            position:relative;
            overflow:hidden;
            transform:rotate(0deg);
            transition:transform 5.8s cubic-bezier(0.12, 0.82, 0.16, 1);
            background:
              conic-gradient(
                #e7ecf8 0deg 72deg,
                #f0dede 72deg 144deg,
                #e7ecf8 144deg 216deg,
                #f0dede 216deg 288deg,
                #e7ecf8 288deg 360deg
              );
            box-shadow:
              inset 0 0 0 8px rgba(0,0,0,0.05),
              0 16px 36px rgba(15,23,42,0.18);
          ">
            <div style="position:absolute; inset:0; border-radius:50%;">
              <div style="position:absolute; left:50%; top:18px; transform:translateX(-50%) rotate(36deg); transform-origin:center 138px; font-weight:900; font-size:18px; color:#344054;">1,000원</div>
              <div style="position:absolute; left:50%; top:18px; transform:translateX(-50%) rotate(108deg); transform-origin:center 138px; font-weight:900; font-size:18px; color:#344054;">2,000원</div>
              <div style="position:absolute; left:50%; top:18px; transform:translateX(-50%) rotate(180deg); transform-origin:center 138px; font-weight:900; font-size:18px; color:#344054;">3,000원</div>
              <div style="position:absolute; left:50%; top:18px; transform:translateX(-50%) rotate(252deg); transform-origin:center 138px; font-weight:900; font-size:18px; color:#344054;">5,000원</div>
              <div style="position:absolute; left:50%; top:18px; transform:translateX(-50%) rotate(324deg); transform-origin:center 138px; font-weight:900; font-size:18px; color:#344054;">10,000원</div>
            </div>

            <div style="
              position:absolute;
              left:50%;
              top:50%;
              transform:translate(-50%,-50%);
              width:84px;
              height:84px;
              border-radius:50%;
              background:linear-gradient(180deg,#ffffff 0%,#f8fafc 100%);
              border:7px solid #e5e7eb;
              display:flex;
              align-items:center;
              justify-content:center;
              font-weight:900;
              font-size:18px;
              color:#111;
              z-index:3;
              box-shadow:0 8px 18px rgba(0,0,0,0.14);
            ">
              SPIN
            </div>
          </div>
        </div>
      </div>

      <div id="rouletteResultText" style="
        margin-top:20px;
        text-align:center;
        font-size:20px;
        font-weight:900;
        color:#111;
        min-height:32px;
      ">
        버튼을 누르면 회전합니다
      </div>

      <div style="margin-top:16px; display:flex; gap:10px;">
        <button
          type="button"
          id="realSpinBtn"
          {'disabled' if not can_spin else ''}
          onclick="startRouletteSpin()"
          style="
            flex:1;
            padding:15px 16px;
            border:none;
            border-radius:16px;
            background:{roulette_btn_bg};
            color:#fff;
            font-size:18px;
            font-weight:900;
            cursor:{roulette_btn_cursor};
            box-shadow:0 10px 20px rgba(17,17,17,0.18);
          ">
          {'돌리기' if can_spin else '불가'}
        </button>

        <button
          type="button"
          onclick="closeRouletteModal()"
          style="
            padding:15px 16px;
            border:none;
            border-radius:16px;
            background:#eef2f7;
            color:#111;
            font-size:16px;
            font-weight:900;
          ">
          닫기
        </button>
      </div>
    </div>
  </div>
</div>

<script>
  let rouletteSpinning = false;
  let rouletteCurrentRotation = 0;

  function openRouletteModal() {{
    document.getElementById('rouletteModal').style.display = 'block';
  }}

  function closeRouletteModal() {{
    if (rouletteSpinning) return;
    document.getElementById('rouletteModal').style.display = 'none';
  }}

  function rewardToStopAngle(reward) {{
    const map = {{
      1000: 324,
      2000: 252,
      3000: 180,
      5000: 108,
      10000: 36
    }};
    return map[reward] ?? 324;
  }}

  function makeConfetti() {{
    const wrap = document.getElementById('rouletteConfetti');
    wrap.innerHTML = '';
    const colors = ['#ef4444', '#f59e0b', '#10b981', '#3b82f6', '#8b5cf6'];

    for (let i = 0; i < 24; i++) {{
      const div = document.createElement('div');
      const left = Math.random() * 100;
      const delay = Math.random() * 0.4;
      const duration = 1.2 + Math.random() * 0.7;
      const rotate = Math.random() * 360;
      const color = colors[i % colors.length];

      div.style.cssText = `
        position:absolute;
        left:${{left}}%;
        top:0;
        width:10px;
        height:16px;
        background:${{color}};
        opacity:0.95;
        border-radius:2px;
        transform:rotate(${{rotate}}deg);
        animation: confettiDrop ${{duration}}s ease-in forwards;
        animation-delay:${{delay}}s;
      `;
      wrap.appendChild(div);
    }}

    if (!document.getElementById('confettiStyle')) {{
      const style = document.createElement('style');
      style.id = 'confettiStyle';
      style.innerHTML = `
        @keyframes confettiDrop {{
          0% {{ transform: translateY(0) rotate(0deg); opacity:1; }}
          100% {{ transform: translateY(180px) rotate(360deg); opacity:0; }}
        }}
      `;
      document.head.appendChild(style);
    }}
  }}

  function prependRecentHistory(reward, segmentValue) {{
    const ul = document.getElementById('recentRouletteList');
    const li = document.createElement('li');

    const now = new Date();
    const mm = String(now.getMonth() + 1).padStart(2, '0');
    const dd = String(now.getDate()).padStart(2, '0');
    const hh = String(now.getHours()).padStart(2, '0');
    const mi = String(now.getMinutes()).padStart(2, '0');

    li.style.marginBottom = '4px';
    li.textContent = `${{mm}}/${{dd}} ${{hh}}:${{mi}} / ${{Number(reward).toLocaleString()}}원 / ${{segmentValue}}건 구간`;

    if (ul.firstChild && ul.firstChild.textContent === '최근 당첨 내역 없음') {{
      ul.innerHTML = '';
    }}

    ul.prepend(li);

    while (ul.children.length > 5) {{
      ul.removeChild(ul.lastChild);
    }}
  }}

  function updateRouletteSummary(data) {{
    document.getElementById('weeklyTotalText').innerText =
      Number(data.weekly_total || 0).toLocaleString() + '원';

    document.getElementById('remainCountText').innerText =
      String(data.remain_count || 0) + '회';

    document.getElementById('usedCountText').innerText =
      String(data.used_count || 0) + '회';

    document.getElementById('eligibleCountText').innerText =
      String(data.eligible_count || 0) + '회';

    const nextSegment = Math.min(100, (Number(data.used_count || 0) + 1) * {roulette_status.get("spin_unit", 10)});
    document.getElementById('nextSegmentText').innerText = String(data.remain_count > 0 ? nextSegment : 0);

    const openBtn = document.getElementById('openRouletteBtn');
    const spinBtn = document.getElementById('realSpinBtn');

    if (Number(data.remain_count || 0) <= 0) {{
      openBtn.disabled = true;
      openBtn.innerText = '룰렛 불가';
      openBtn.style.background = '#bbb';
      openBtn.style.cursor = 'not-allowed';

      if (spinBtn) {{
        spinBtn.disabled = true;
        spinBtn.innerText = '불가';
        spinBtn.style.background = '#bbb';
        spinBtn.style.cursor = 'not-allowed';
      }}
    }} else {{
      openBtn.disabled = false;
      openBtn.innerText = '룰렛 돌리기';
      openBtn.style.background = '#111';
      openBtn.style.cursor = 'pointer';

      if (spinBtn && !rouletteSpinning) {{
        spinBtn.disabled = false;
        spinBtn.innerText = '다시 돌리기';
        spinBtn.style.background = '#111';
        spinBtn.style.cursor = 'pointer';
      }}
    }}
  }}

  async function refreshRouletteStatus() {{
    const res = await fetch('/api/roulette/status', {{
      method: 'POST',
      headers: {{ 'Content-Type': 'application/json' }},
      body: JSON.stringify({{
        phone: "{normalize_phone(rider.get('phoneNumber', '') or '')}"
      }})
    }});
    const data = await res.json();
    if (res.ok) {{
      updateRouletteSummary(data);
    }}
  }}

  async function startRouletteSpin() {{
    if (rouletteSpinning) return;

    rouletteSpinning = true;
    const btn = document.getElementById('realSpinBtn');
    const wheel = document.getElementById('rouletteWheel');
    const resultText = document.getElementById('rouletteResultText');

    btn.disabled = true;
    btn.innerText = '회전 중...';
    resultText.innerText = '당첨 결과를 확인하고 있어요...';

    try {{
      const res = await fetch('/api/roulette/spin', {{
        method: 'POST',
        headers: {{ 'Content-Type': 'application/json' }},
        body: JSON.stringify({{
          phone: "{normalize_phone(rider.get('phoneNumber', '') or '')}"
        }})
      }});

      const data = await res.json();

      if (!res.ok || !data.ok) {{
        throw new Error(data.detail || '룰렛 처리 실패');
      }}

      const reward = Number(data.reward || 0);
      const targetAngle = rewardToStopAngle(reward);
      const extraSpins = 360 * 7;

      const currentAngle = ((rouletteCurrentRotation % 360) + 360) % 360;
      let delta = (targetAngle - currentAngle + 360) % 360;
      if (delta < 180) {{
        delta += 360;
      }}

      rouletteCurrentRotation = rouletteCurrentRotation + extraSpins + delta;
      wheel.style.transform = `rotate(${{rouletteCurrentRotation}}deg)`;

      setTimeout(() => {{
        makeConfetti();
        resultText.innerText = `🎉 ${{reward.toLocaleString()}}원 당첨`;
      }}, 5200);

      setTimeout(async () => {{
        prependRecentHistory(reward, Number(data.segment_value || 0));
        updateRouletteSummary(data);
        btn.innerText = '완료';
        rouletteSpinning = false;
        await refreshRouletteStatus();
      }}, 5900);

    }} catch (e) {{
      resultText.innerText = e.message || '룰렛 처리 중 오류';
      btn.disabled = false;
      btn.innerText = '다시 시도';
      rouletteSpinning = false;
    }}
  }}
</script>

      <div style="margin-top:12px; padding:12px; border:1px solid #eee; border-radius:14px; background:#fff;">
        <div style="font-weight:900; margin-bottom:6px;">NMAX 이벤트 누적 완료건수</div>
        <div style="color:#666;">기간: <b>{NMAX_LABEL}</b></div>
        <div style="font-size:22px; font-weight:900; margin-top:6px;">{pcx_text}</div>
        <div style="color:#999; font-size:12px; margin-top:6px;">
          * 이 숫자는 PC에서 <b>{NMAX_START_DATE.isoformat()} ~ 어제</b> 범위를 업로드했을 때만 표시됩니다.
        </div>
      </div>

      <div style="margin-top:14px; display:flex; gap:10px; flex-wrap:wrap;">
        <a href="/team" style="text-decoration:none; color:#fff; background:#111; padding:10px 14px; border-radius:10px; font-weight:900;">내 팀 페이지</a>
        <a href="/" style="text-decoration:none; color:#111; padding:10px 0;">← 다시 조회</a>
      </div>
    </div>
    """
    return html_page("등급 조회 결과", body)


# -----------------------------
# Admin login
# -----------------------------
@app.get("/admin-login", response_class=HTMLResponse)
def admin_login_page():
    body = """
    <div style="max-width:420px; margin:80px auto; background:#fff; border:1px solid #e8e8e8; border-radius:16px; padding:20px;">
      <h2 style="margin-top:0;">관리자 로그인</h2>
      <form method="post" action="/admin-login">
        <input type="password" name="password" placeholder="비밀번호"
               style="width:100%; font-size:18px; padding:12px; border:1px solid #ddd; border-radius:12px;" required />
        <button type="submit"
                style="width:100%; margin-top:12px; font-size:18px; padding:12px;
                       border:none; border-radius:12px; background:#111; color:#fff;">
          로그인
        </button>
      </form>
      <div style="margin-top:12px;">
        <a href="/" style="color:#666; text-decoration:none;">← 메인으로</a>
      </div>
    </div>
    """
    return html_page("관리자 로그인", body)


@app.post("/admin-login")
def admin_login_action(request: Request, password: str = Form(...)):
    if password == get_admin_password():
        request.session["is_admin"] = True
        return RedirectResponse("/dashboard", status_code=303)

    body = """
    <div style="max-width:420px; margin:80px auto; background:#fff; border:1px solid #e8e8e8; border-radius:16px; padding:20px;">
      <h3 style="margin-top:0;">비밀번호가 틀렸습니다</h3>
      <div style="color:#666;">다시 시도해주세요.</div>
      <div style="margin-top:12px;"><a href="/admin-login" style="text-decoration:none; color:#111;">다시 로그인</a></div>
      <div style="margin-top:10px;"><a href="/" style="text-decoration:none; color:#666;">← 메인으로</a></div>
    </div>
    """
    return HTMLResponse(html_page("로그인 실패", body))


@app.get("/admin-logout")
def admin_logout(request: Request):
    request.session.clear()
    return RedirectResponse("/", status_code=303)


# -----------------------------
# Admin set / clear
# -----------------------------
@app.post("/admin/set-join")
def admin_set_join(request: Request, key: str = Form(...), join_date: str = Form(...), redirect_q: str = Form(default="")):
    r = require_admin(request)
    if r:
        return r

    key = (key or "").strip()
    jd = safe_date_parse(join_date)
    if not key or jd is None:
        return RedirectResponse(f"/dashboard?q={redirect_q}", status_code=303)

    data = load_overrides()
    data[key] = jd.isoformat()
    save_overrides(data)
    return RedirectResponse(f"/dashboard?q={redirect_q}", status_code=303)


@app.post("/admin/clear-join")
def admin_clear_join(request: Request, key: str = Form(...), redirect_q: str = Form(default="")):
    r = require_admin(request)
    if r:
        return r

    key = (key or "").strip()
    data = load_overrides()
    if key in data:
        del data[key]
        save_overrides(data)
    return RedirectResponse(f"/dashboard?q={redirect_q}", status_code=303)


@app.post("/admin/set-login4")
def admin_set_login4(request: Request, name_norm: str = Form(...), real4: str = Form(...), login4: str = Form(...), redirect_q: str = Form(default="")):
    r = require_admin(request)
    if r:
        return r

    name_norm = (name_norm or "").strip()
    real4 = (real4 or "").strip()
    login4 = (login4 or "").strip()

    if not name_norm or not re.fullmatch(r"\d{4}", real4) or not re.fullmatch(r"\d{4}", login4):
        return RedirectResponse(f"/dashboard?q={redirect_q}", status_code=303)

    set_login4_override(name_norm, real4, login4)
    return RedirectResponse(f"/dashboard?q={redirect_q}", status_code=303)


@app.post("/admin/clear-login4")
def admin_clear_login4(request: Request, name_norm: str = Form(...), real4: str = Form(...), redirect_q: str = Form(default="")):
    r = require_admin(request)
    if r:
        return r

    name_norm = (name_norm or "").strip()
    real4 = (real4 or "").strip()
    if not name_norm or not re.fullmatch(r"\d{4}", real4):
        return RedirectResponse(f"/dashboard?q={redirect_q}", status_code=303)

    clear_login4_override(name_norm, real4)
    return RedirectResponse(f"/dashboard?q={redirect_q}", status_code=303)


@app.post("/admin/set-prevplus")
def admin_set_prevplus(request: Request, key: str = Form(...), prevplus: str = Form(...), redirect_q: str = Form(default="")):
    r = require_admin(request)
    if r:
        return r

    key = (key or "").strip()
    v = _safe_int(prevplus, 0)
    if not key:
        return RedirectResponse(f"/dashboard?q={redirect_q}", status_code=303)

    v = max(-MAX_PLUS_INPUT, min(MAX_PLUS_INPUT, v))
    set_prevplus(key, v)
    return RedirectResponse(f"/dashboard?q={redirect_q}", status_code=303)


@app.post("/admin/clear-prevplus")
def admin_clear_prevplus(request: Request, key: str = Form(...), redirect_q: str = Form(default="")):
    r = require_admin(request)
    if r:
        return r

    key = (key or "").strip()
    if key:
        clear_prevplus(key)
    return RedirectResponse(f"/dashboard?q={redirect_q}", status_code=303)


@app.post("/admin/set-plannedplus")
def admin_set_plannedplus(request: Request, key: str = Form(...), plannedplus: str = Form(...), redirect_q: str = Form(default="")):
    r = require_admin(request)
    if r:
        return r

    key = (key or "").strip()
    v = _safe_int(plannedplus, 0)
    if not key:
        return RedirectResponse(f"/dashboard?q={redirect_q}", status_code=303)

    v = max(-MAX_PLUS_INPUT, min(MAX_PLUS_INPUT, v))
    set_plannedplus(key, v)
    return RedirectResponse(f"/dashboard?q={redirect_q}", status_code=303)

@app.post("/admin/set-sticker")
def admin_set_sticker(
    request: Request,
    key: str = Form(...),
    sticker_attached: str = Form(default="0"),
    redirect_q: str = Form(default="")
):
    r = require_admin(request)
    if r:
        return r

    key = (key or "").strip()
    if not key:
        return RedirectResponse(f"/dashboard?q={redirect_q}", status_code=303)

    checked = (sticker_attached == "1")
    set_sticker_attached(key, checked)
    return RedirectResponse(f"/dashboard?q={redirect_q}", status_code=303)


@app.post("/admin/bulk-save-plus-sticker")
async def admin_bulk_save_plus_sticker(request: Request):
    r = require_admin(request)
    if r:
        return JSONResponse({"ok": False, "error": "UNAUTHORIZED"}, status_code=401)

    payload = await request.json()
    rows = payload.get("rows") or []
    if not isinstance(rows, list):
        return JSONResponse({"ok": False, "error": "INVALID_PAYLOAD"}, status_code=400)

    valid_keys: set[str] = set()
    for rr in fetch_riders_cached():
        nm = rr.get("name") or ""
        login4, _, _ = get_login4_for_rider(rr)
        if nm and login4:
            valid_keys.add(f"{norm_name(nm)}|{login4}")

    saved = 0
    skipped = 0
    for row in rows:
        if not isinstance(row, dict):
            skipped += 1
            continue
        key = str(row.get("key") or "").strip()
        if key not in valid_keys:
            skipped += 1
            continue

        prev_v = max(-MAX_PLUS_INPUT, min(MAX_PLUS_INPUT, _safe_int(row.get("prevplus"), 0)))
        planned_v = max(-MAX_PLUS_INPUT, min(MAX_PLUS_INPUT, _safe_int(row.get("plannedplus"), 0)))
        sticker_v = bool(row.get("sticker"))

        set_prevplus(key, prev_v)
        set_plannedplus(key, planned_v)
        set_sticker_attached(key, sticker_v)
        saved += 1

    return {"ok": True, "saved": saved, "skipped": skipped}

@app.post("/admin/clear-plannedplus")
def admin_clear_plannedplus(request: Request, key: str = Form(...), redirect_q: str = Form(default="")):
    r = require_admin(request)
    if r:
        return r

    key = (key or "").strip()
    if key:
        clear_plannedplus(key)
    return RedirectResponse(f"/dashboard?q={redirect_q}", status_code=303)


# -----------------------------
# Plus backup / restore
# -----------------------------
@app.get("/plus-backup.xlsx")
def plus_backup_excel(request: Request):
    r = require_admin(request)
    if r:
        return r

    if not store_ready():
        return not_ready_page()

    riders = fetch_riders_cached()
    prevplus_map = load_prevplus_map()
    plannedplus_map = load_plannedplus_map()
    sticker_map = load_sticker_map()

    rows = []
    for rr in riders:
        name = rr.get("name") or ""
        if not name:
            continue
        login4, real4, _ = get_login4_for_rider(rr)
        key = f"{norm_name(name)}|{login4}"
        rows.append({
            "name": name,
            "login_key": key,
            "login4": login4,
            "real4": real4,
            "prev_plus": int(prevplus_map.get(key, 0) or 0),
            "planned_plus": int(plannedplus_map.get(key, 0) or 0),
            "sticker_attached": 1 if bool(sticker_map.get(key, False)) else 0,
            "sticker_bonus": 20 if bool(sticker_map.get(key, False)) else 0,
        })

    rows.sort(key=lambda x: x["name"])

    wb = Workbook()
    ws = wb.active
    ws.title = "plus_backup"

    headers = ["name", "login_key", "login4", "real4", "prev_plus", "planned_plus", "sticker_attached", "sticker_bonus"]
    ws.append(headers)
    for r0 in rows:
        ws.append([r0.get(h, "") for h in headers])

    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(h) + 4)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"riderwelfare_plus_backup_{today_kst().isoformat()}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/admin/restore-plus-xlsx")
async def admin_restore_plus_xlsx(request: Request, file: UploadFile = File(...)):
    r = require_admin(request)
    if r:
        return r

    if not file.filename.lower().endswith(".xlsx"):
        body = """
        <div style="max-width:640px; margin:40px auto; background:#fff; border:1px solid #e8e8e8; border-radius:16px; padding:20px;">
          <h3 style="margin-top:0;">업로드 실패</h3>
          <div style="color:#666;">xlsx 파일만 업로드할 수 있습니다.</div>
          <div style="margin-top:12px;"><a href="/dashboard" style="text-decoration:none; color:#111;">← 대시보드로 돌아가기</a></div>
        </div>
        """
        return HTMLResponse(html_page("업로드 실패", body))

    try:
        file_bytes = await file.read()
        restored_rows, saved_count = restore_plus_from_workbook(file_bytes)
        body = f"""
        <div style="max-width:720px; margin:40px auto; background:#fff; border:1px solid #e8e8e8; border-radius:16px; padding:20px;">
          <h3 style="margin-top:0;">플러스 복원 완료</h3>
          <div style="color:#666; line-height:1.7;">
            업로드 파일: <b>{file.filename}</b><br/>
            처리 행 수: <b>{restored_rows}</b><br/>
            건너뛴 행 수: <b>{saved_count}</b><br/>
            이전등급 플러스 / 예정등급 플러스 / 스티커가 현재 라이더 기준으로 갱신되었습니다.
          </div>
          <div style="margin-top:14px;"><a href="/dashboard" style="text-decoration:none; color:#111;">← 대시보드로 돌아가기</a></div>
        </div>
        """
        return HTMLResponse(html_page("플러스 복원 완료", body))
    except Exception as e:
        body = f"""
        <div style="max-width:720px; margin:40px auto; background:#fff; border:1px solid #e8e8e8; border-radius:16px; padding:20px;">
          <h3 style="margin-top:0;">플러스 복원 실패</h3>
          <div style="color:#666; line-height:1.7;">
            파일: <b>{file.filename}</b><br/>
            오류: <b>{str(e)}</b><br/>
            업로드 엑셀에 <b>name / prev_plus / planned_plus</b> 컬럼이 필요합니다. real4 또는 login_key가 있으면 더 정확히 복원됩니다.
          </div>
          <div style="margin-top:14px;"><a href="/dashboard" style="text-decoration:none; color:#111;">← 대시보드로 돌아가기</a></div>
        </div>
        """
        return HTMLResponse(html_page("플러스 복원 실패", body))


# -----------------------------
# Dashboard
# -----------------------------
@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(request: Request, q: str = ""):
    r = require_admin(request)
    if r:
        return r

    if not store_ready():
        return not_ready_page()

    riders = fetch_riders_cached()
    # 관리자 페이지 성능 최적화: 설정 파일은 요청당 한 번만 읽습니다.
    join_overrides = load_overrides()
    login4_map = load_login4_map()
    prevplus_map = load_prevplus_map()
    plannedplus_map = load_plannedplus_map()
    sticker_map = load_sticker_map()
    qn = norm_name(q)

    rider_rows = []
    for rr in riders:
        nm = rr.get("name") or ""
        ph = rr.get("phoneNumber") or ""
        real4 = last4_from_phone(ph)
        if not real4:
            continue
        if qn and (qn not in (norm_name(nm) + real4)):
            continue
        rider_rows.append(rr)

    today = operation_date()
    ds = fetch_delivery_status_cached()
    today_stats_map = build_today_stats_map(ds)

    cur_group: Dict[Tuple[date, date], List[Dict[str, Any]]] = {}
    prev_group: Dict[Tuple[date, date], List[Dict[str, Any]]] = {}

    for rr in rider_rows:
        nm = rr.get("name") or ""
        ph = rr.get("phoneNumber") or ""
        real4 = last4_from_phone(ph)
        nn = norm_name(nm)

        real_lookup_key = f"{nn}|{real4}"
        mapped_login4 = str(login4_map.get(real_lookup_key) or "").strip()
        if re.fullmatch(r"\d{4}", mapped_login4):
            login4, login_src = mapped_login4, "override"
        else:
            login4, login_src = real4, "real"

        real_key = f"{nn}|{real4}"
        login_key = f"{nn}|{login4}"

        ov_join = join_overrides.get(login_key)
        eff_join = safe_date_parse(ov_join) if ov_join else None
        if eff_join is not None:
            join_src = "override"
        else:
            created_raw = rr.get("createdDate")
            eff_join = safe_date_parse(created_raw[:10]) if isinstance(created_raw, str) and len(created_raw) >= 10 else None
            if eff_join is not None:
                join_src = "createdDate"
            else:
                eff_join = today
                join_src = "fallback"

        cur_start, cur_end_incl = current_period(eff_join, today)
        cur_from, cur_to = period_to_from_to(cur_start, cur_end_incl)

        prev_end_incl = cur_start - timedelta(days=1)
        prev_m = date(cur_start.year, cur_start.month, 1) + relativedelta(months=-1)
        prev_start = clamp_day(prev_m.year, prev_m.month, eff_join.day)
        prev_from, prev_to = period_to_from_to(prev_start, prev_end_incl)

        item = {
            "rider": rr,
            "nn": nn,
            "real4": real4,
            "login4": login4,
            "login_src": login_src,
            "real_key": real_key,
            "login_key": login_key,
            "eff_join": eff_join,
            "join_src": join_src,
            "cur_start": cur_start,
            "cur_end_incl": cur_end_incl,
            "cur_from": cur_from,
            "cur_to": cur_to,
            "prev_start": prev_start,
            "prev_end_incl": prev_end_incl,
            "prev_from": prev_from,
            "prev_to": prev_to,
        }

        cur_group.setdefault((cur_from, cur_to), []).append(item)
        prev_group.setdefault((prev_from, prev_to), []).append(item)

    prev_completed_map: Dict[str, int] = {}
    for (from_d, to_d), items in prev_group.items():
        if from_d > to_d:
            for it in items:
                prev_completed_map[it["real_key"]] = 0
            continue

        cmap = fetch_status_complete_map_cached(from_d, to_d)
        for it in items:
            prev_completed_map[it["real_key"]] = int(cmap.get(it["real_key"], 0))

    final_rows = []
    for (from_d, to_d), items in cur_group.items():
        if from_d > to_d:
            cmap = {}
        else:
            cmap = fetch_status_complete_map_cached(from_d, to_d)

        for it in items:
            rr = it["rider"]
            nm = rr.get("name") or ""
            created_raw = rr.get("createdDate")
            created_d = created_raw[:10] if isinstance(created_raw, str) and len(created_raw) >= 10 else "-"

            cur_completed_raw = int(cmap.get(it["real_key"], 0))
            prev_completed_raw = int(prev_completed_map.get(it["real_key"], 0))

            prev_plus = int(prevplus_map.get(it["login_key"], 0) or 0)
            planned_plus_base = int(plannedplus_map.get(it["login_key"], 0) or 0)
            sticker_attached = bool(sticker_map.get(it["login_key"], False))
            sticker_bonus = 20 if sticker_attached else 0
            planned_plus = planned_plus_base + sticker_bonus

            planned_total = cur_completed_raw + planned_plus
            prev_total = prev_completed_raw + prev_plus

            planned_grade = grade_from_total(planned_total)
            current_grade = grade_from_total(prev_total)

            nxt, remain = next_grade_target(planned_total)

            ov = join_overrides.get(it["login_key"])
            join_default_val = ov if ov else it["eff_join"].isoformat()

            login_badge = "가상뒷4" if it["login_src"] == "override" else "실제뒷4"
            login_badge_color = "#111" if it["login_src"] == "override" else "#888"

            join_badge = "관리자설정" if it["join_src"] == "override" else "배민입사"
            join_badge_color = "#111" if it["join_src"] == "override" else "#888"

            today_info = today_stats_map.get(it["real_key"]) or {
                "complete": 0,
                "reject": 0,
                "cancel": 0,
                "status_desc": "-",
                **calc_bad_ratio(0, 0, 0),
            }

            # 룰렛 상태는 관리자 표에서 사용하지 않으므로 전원 DB 조회를 생략합니다.
            final_rows.append({
                "name": nm,
                "created": created_d,
                "real4": it["real4"],
                "login4": it["login4"],
                "login_badge": login_badge,
                "login_badge_color": login_badge_color,
                "join_effective": it["eff_join"].isoformat(),
                "join_badge": join_badge,
                "join_badge_color": join_badge_color,
                "join_default_val": join_default_val,
                "policy_from": it["cur_start"].isoformat(),
                "policy_to": it["cur_end_incl"].isoformat(),
                "api_from": it["cur_from"].isoformat(),
                "api_to": it["cur_to"].isoformat(),
                "cur_completed_raw": cur_completed_raw,
                "prev_completed_raw": prev_completed_raw,
                "prev_plus": prev_plus,
                "planned_plus": planned_plus,
                "current_grade": current_grade,
                "planned_grade": planned_grade,
                "next": nxt or "-",
                "remain": remain if remain is not None else "-",
                "login_key": it["login_key"],
                "name_norm": it["nn"],
                "today_complete": int(today_info["complete"]),
                "today_reject": int(today_info["reject"]),
                "today_cancel": int(today_info["cancel"]),
                "today_total": int(today_info["total"]),
                "today_ratio": float(today_info["ratio"]),
                "today_ratio_fg": str(today_info["fg"]),
                "today_ratio_bg": str(today_info["bg"]),
                "today_ratio_label": str(today_info["label"]),
                "today_status_desc": str(today_info["status_desc"]),
                "planned_plus_base": planned_plus_base,
                "sticker_bonus": sticker_bonus,
                "sticker_attached": sticker_attached,
            })

    final_rows.sort(key=lambda x: (x["cur_completed_raw"], -x["today_ratio"]), reverse=True)

    tr_html = ""
    for i, it in enumerate(final_rows, start=1):
        tr_html += f"""
        <tr class="bulk-row" data-key="{it['login_key']}">
          <td style="padding:10px; border-bottom:1px solid #eee; text-align:right; color:#999;">{i}</td>

          <td style="padding:10px; border-bottom:1px solid #eee;">
            <div style="display:flex; align-items:center; gap:8px; flex-wrap:wrap;">
              <div style="font-weight:900;">{it['name']}</div>
              <span style="
                display:inline-block;
                padding:3px 8px;
                border-radius:999px;
                background:{it['today_ratio_bg']};
                color:{it['today_ratio_fg']};
                font-size:12px;
                font-weight:900;
                border:1px solid rgba(0,0,0,0.06);
              ">
                거절+취소 {it['today_ratio']}%
              </span>
            </div>
            <div style="font-size:12px; color:#999; margin-top:4px;">
              오늘 완료 {it['today_complete']} / 거절 {it['today_reject']} / 취소 {it['today_cancel']}
            </div>
          </td>

          <td style="padding:10px; border-bottom:1px solid #eee; color:#666;">
            배민뒷4: {it['real4']}<br/>
            <b>로그인뒷4: {it['login4']}</b>
            <div style="margin-top:6px;">
              <span style="font-size:12px; color:{it['login_badge_color']}; border:1px solid #ddd; padding:2px 8px; border-radius:999px; background:#fafafa;">
                {it['login_badge']}
              </span>
            </div>

            <div style="margin-top:8px; display:flex; gap:8px; flex-wrap:wrap;">
              <form method="post" action="/admin/set-login4" style="display:flex; gap:6px; align-items:center;">
                <input type="hidden" name="name_norm" value="{it['name_norm']}" />
                <input type="hidden" name="real4" value="{it['real4']}" />
                <input type="hidden" name="redirect_q" value="{q}" />
                <input name="login4" value="{it['login4']}" placeholder="4자리"
                       style="width:90px; padding:8px 10px; border:1px solid #ddd; border-radius:10px;" />
                <button type="submit" style="padding:8px 10px; border:none; border-radius:10px; background:#111; color:#fff;">변경</button>
              </form>

              <form method="post" action="/admin/clear-login4">
                <input type="hidden" name="name_norm" value="{it['name_norm']}" />
                <input type="hidden" name="real4" value="{it['real4']}" />
                <input type="hidden" name="redirect_q" value="{q}" />
                <button type="submit" style="padding:8px 10px; border:1px solid #ddd; border-radius:10px; background:#fff; color:#111;">초기화</button>
              </form>
            </div>
          </td>

          <td style="padding:10px; border-bottom:1px solid #eee; color:#666;">{it['created']}</td>

          <td style="padding:10px; border-bottom:1px solid #eee;">
            <div style="display:flex; align-items:center; gap:10px; flex-wrap:wrap;">
              <div style="font-weight:900;">{it['join_effective']}</div>
              <span style="font-size:12px; color:{it['join_badge_color']}; border:1px solid #ddd; padding:2px 8px; border-radius:999px; background:#fafafa;">
                {it['join_badge']}
              </span>
            </div>

            <div style="margin-top:8px; display:flex; gap:8px; flex-wrap:wrap;">
              <form method="post" action="/admin/set-join" style="display:flex; gap:6px; align-items:center;">
                <input type="hidden" name="key" value="{it['login_key']}" />
                <input type="hidden" name="redirect_q" value="{q}" />
                <input name="join_date" value="{it['join_default_val']}" placeholder="YYYY-MM-DD"
                       style="width:120px; padding:8px 10px; border:1px solid #ddd; border-radius:10px;" />
                <button type="submit" style="padding:8px 10px; border:none; border-radius:10px; background:#111; color:#fff;">저장</button>
              </form>

              <form method="post" action="/admin/clear-join">
                <input type="hidden" name="key" value="{it['login_key']}" />
                <input type="hidden" name="redirect_q" value="{q}" />
                <button type="submit" style="padding:8px 10px; border:1px solid #ddd; border-radius:10px; background:#fff; color:#111;">초기화</button>
              </form>
            </div>
          </td>

          <td style="padding:10px; border-bottom:1px solid #eee; color:#666;">
            <div style="font-weight:700;">정책: {it['policy_from']} ~ {it['policy_to']}</div>
            <div style="font-size:12px; color:#999; margin-top:4px;">업로드 반영: {it['api_from']} ~ {it['api_to']}</div>
          </td>

          <td style="padding:10px; border-bottom:1px solid #eee; text-align:center;">
            <div style="font-weight:900;">{it['today_complete']}</div>
            <div style="font-size:12px; color:#999;">거절 {it['today_reject']} / 취소 {it['today_cancel']}</div>
            <div style="font-size:12px; color:{it['today_ratio_fg']}; font-weight:900; margin-top:4px;">{it['today_ratio']}%</div>
          </td>

          <td style="padding:10px; border-bottom:1px solid #eee; text-align:center;">
              <div style="font-weight:900; font-size:20px;">{it['cur_completed_raw'] + it['planned_plus']}</div>
              <div style="font-size:12px; color:#999; margin-top:4px;">
                  현재 {it['cur_completed_raw']} + 플러스 {it['planned_plus']}
              </div>
          </td>

          <td style="padding:10px; border-bottom:1px solid #eee; text-align:center;">
            <div style="font-weight:900;">{it['current_grade']}</div>
            <div style="font-size:12px; color:#999;">이전 {it['prev_completed_raw']} + 플러스 {it['prev_plus']}</div>
          </td>

          <td style="padding:10px; border-bottom:1px solid #eee; text-align:center;">
            <div style="font-weight:900;">{it['planned_grade']}</div>
            <div style="font-size:12px; color:#999;">현재 {it['cur_completed_raw']} + 플러스 {it['planned_plus']}</div>
          </td>

          <td style="padding:10px; border-bottom:1px solid #eee; text-align:center; color:#666;">{it['next']}</td>
          <td style="padding:10px; border-bottom:1px solid #eee; text-align:right; color:#666;">{it['remain']}</td>

          <td style="padding:10px; border-bottom:1px solid #eee; color:#666; min-width:260px;">
            <div style="font-weight:700;">현재등급(이전) 플러스</div>
            <div style="font-size:12px; color:#999;">이전등급 계산에만 적용</div>
            <div style="margin-top:8px; display:flex; gap:8px; flex-wrap:wrap;">
              <form method="post" action="/admin/set-prevplus" style="display:flex; gap:6px; align-items:center;">
                <input type="hidden" name="key" value="{it['login_key']}" />
                <input type="hidden" name="redirect_q" value="{q}" />
                <input name="prevplus" class="bulk-prevplus" data-key="{it['login_key']}" value="{it['prev_plus']}" placeholder="예: 20"
                       style="width:100px; padding:8px 10px; border:1px solid #ddd; border-radius:10px;" />
                <button type="submit" style="padding:8px 10px; border:none; border-radius:10px; background:#111; color:#fff;">저장</button>
              </form>

              <form method="post" action="/admin/clear-prevplus">
                <input type="hidden" name="key" value="{it['login_key']}" />
                <input type="hidden" name="redirect_q" value="{q}" />
                <button type="submit" style="padding:8px 10px; border:1px solid #ddd; border-radius:10px; background:#fff; color:#111;">초기화</button>
              </form>
            </div>
          </td>

          <td style="padding:10px; border-bottom:1px solid #eee; color:#666; min-width:260px;">
            <div style="font-weight:700;">예정등급(현재) 플러스</div>
            <div style="font-size:12px; color:#999;">수동플러스 + 스티커(+20)가 예정등급 계산에 적용</div>
            <div style="font-size:12px; color:#666; margin-top:4px;">
                수동 {it['planned_plus_base']} + 스티커 {it['sticker_bonus']} = 합계 {it['planned_plus']}
            </div>
            <div style="margin-top:8px; display:flex; gap:8px; flex-wrap:wrap;">
              <form method="post" action="/admin/set-plannedplus" style="display:flex; gap:6px; align-items:center;">
                <input type="hidden" name="key" value="{it['login_key']}" />
                <input type="hidden" name="redirect_q" value="{q}" />
                <input name="plannedplus" class="bulk-plannedplus" data-key="{it['login_key']}" value="{it['planned_plus_base']}" placeholder="예: 20"
                       style="width:100px; padding:8px 10px; border:1px solid #ddd; border-radius:10px;" />
                <button type="submit" style="padding:8px 10px; border:none; border-radius:10px; background:#111; color:#fff;">저장</button>
              </form>

              <form method="post" action="/admin/clear-plannedplus">
                <input type="hidden" name="key" value="{it['login_key']}" />
                <input type="hidden" name="redirect_q" value="{q}" />
                <button type="submit" style="padding:8px 10px; border:1px solid #ddd; border-radius:10px; background:#fff; color:#111;">초기화</button>
              </form>
            </div>

            <form method="post" action="/admin/set-sticker" style="margin-top:8px; display:flex; align-items:center; gap:8px; flex-wrap:wrap;">
              <input type="hidden" name="key" value="{it['login_key']}" />
              <input type="hidden" name="redirect_q" value="{q}" />
              <label style="display:flex; align-items:center; gap:6px; font-size:13px; color:#111;">
                <input type="checkbox" class="bulk-sticker" data-key="{it['login_key']}" name="sticker_attached" value="1" {'checked' if it['sticker_attached'] else ''} />
                스티커부착자 (+20)
              </label>
              <button type="submit" style="padding:8px 10px; border:none; border-radius:10px; background:#111; color:#fff;">저장</button>
            </form>
          </td>
        </tr>
        """

    cycle = get_previous_operational_cycle_bounds()

    body = f"""
    <div style="background:#fff; border:1px solid #e8e8e8; border-radius:16px; padding:16px;">
      <div style="display:flex; align-items:flex-start; justify-content:space-between; gap:10px; flex-wrap:wrap;">
        <div>
          <h2 style="margin:0 0 6px 0;">전체 등급 현황</h2>
          <div style="color:#666; line-height:1.6;">
            - <b>계약종료</b> 라이더는 자동 제외<br/>
            - <b>오늘운행</b>은 실시간 기준: 완료 / 거절 / 취소 / 거절+취소율 표시<br/>
            - <b>현재등급(이전)</b>은 “이전기간 완료건수 + 플러스(개인별)”로 계산<br/>
            - <b>예정등급(현재)</b>은 “현재기간 완료건수 + 예정플러스(개인별)”로 계산
          </div>
          <div style="color:#888; font-size:13px; margin-top:6px;">
            * 이름 옆 배지는 <b>거절+취소율</b>입니다. 19% 이하 초록 / 20~29.9% 노랑 / 30% 이상 빨강
          </div>
        </div>

        <div style="display:flex; gap:10px; align-items:center; flex-wrap:wrap;">
            <a href="/dashboard.xlsx" style="text-decoration:none; color:#111;">전체 엑셀 다운로드</a>
            <a href="/plus-backup.xlsx" style="text-decoration:none; color:#111;">플러스 백업 엑셀</a>
            <a href="/admin/roulette" style="text-decoration:none; color:#111; font-weight:900;">룰렛 관리자</a>
            <a href="/admin/api/roulette/payout.csv" style="text-decoration:none; color:#111;">직전주 룰렛 지급CSV</a>
            <a href="/admin/api/roulette/export" style="text-decoration:none; color:#111;">룰렛 JSON</a>
            <a href="/" style="text-decoration:none; color:#111;">개인 조회</a>
            <a href="/admin/teams" style="text-decoration:none; color:#111; font-weight:900;">팀 구성 관리</a>
        <a href="/admin/team-sets" style="text-decoration:none; color:#111; font-weight:900;">팀 세트 관리</a>
        <a href="/admin-logout" style="text-decoration:none; color:#666;">로그아웃</a>
        </div>
      </div>

      <div style="margin-top:10px; color:#555; font-size:13px;">
        직전 지급 주차: <b>{cycle["display_start"]}</b> ~ <b>{cycle["display_end"]}</b>
      </div>

      <div style="margin-top:14px; display:flex; gap:12px; flex-wrap:wrap;">
        <form method="get" action="/dashboard" style="display:flex; gap:8px; flex:1 1 480px;">
          <input name="q" value="{q}"
                 placeholder="이름 또는 배민뒷4 검색 (예: 이정 / 1898)"
                 style="flex:1; font-size:16px; padding:10px 12px; border:1px solid #ddd; border-radius:12px;" />
          <button type="submit"
                  style="font-size:16px; padding:10px 14px; border:none; border-radius:12px; background:#111; color:#fff;">
            검색
          </button>
        </form>

        <form method="post" action="/admin/restore-plus-xlsx" enctype="multipart/form-data"
              style="display:flex; gap:8px; align-items:center; flex:1 1 420px; background:#f7f9fc; padding:10px 12px; border:1px solid #e4ebf5; border-radius:12px;">
          <input type="file" name="file" accept=".xlsx"
                 style="flex:1; font-size:14px;" required />
          <button type="submit"
                  style="font-size:14px; padding:10px 14px; border:none; border-radius:12px; background:#111; color:#fff;">
            플러스 복원 업로드
          </button>
        </form>
      </div>

      <div style="margin-top:8px; color:#777; font-size:13px;">
        * 기존 <b>dashboard.xlsx</b> 또는 <b>플러스 백업 엑셀</b> 업로드 가능 (권장 컬럼: name / login_key / real4 / prev_plus / planned_plus / sticker_attached)<br/>
        * 예정등급 플러스 입력칸은 <b>수동 플러스</b>만 입력합니다. 스티커 +20은 자동으로 별도 합산됩니다.
      </div>

      <div style="margin-top:12px; display:flex; gap:8px; align-items:center; flex-wrap:wrap; background:#fff8e8; border:1px solid #ffe1a8; border-radius:12px; padding:10px 12px;">
        <button type="button" onclick="bulkSavePlusSticker()"
                style="font-size:15px; padding:10px 14px; border:none; border-radius:12px; background:#111; color:#fff; font-weight:900; cursor:pointer;">
          플러스/스티커 전체 저장
        </button>
        <span id="bulkSaveResult" style="font-size:13px; color:#666;">여러 명 수정 후 이 버튼 한 번만 누르면 됩니다.</span>
      </div>

      <script>
      async function bulkSavePlusSticker() {{
        // 행 단위로 읽습니다.
        // CSS selector로 data-key를 다시 찾으면 한글/특수문자/중복 상황에서 잘못 매칭될 수 있어
        // 같은 <tr> 안의 입력값만 묶어 저장합니다.
        const rows = Array.from(document.querySelectorAll('tr.bulk-row')).map(tr => {{
          const key = tr.dataset.key || '';
          const prevEl = tr.querySelector('.bulk-prevplus');
          const planEl = tr.querySelector('.bulk-plannedplus');
          const stEl = tr.querySelector('.bulk-sticker');
          return {{
            key: key,
            prevplus: prevEl ? prevEl.value : 0,
            plannedplus: planEl ? planEl.value : 0,
            sticker: stEl ? stEl.checked : false
          }};
        }}).filter(row => row.key);
        const box = document.getElementById('bulkSaveResult');
        box.textContent = '저장 중...';
        try {{
          const res = await fetch('/admin/bulk-save-plus-sticker', {{
            method: 'POST',
            headers: {{'Content-Type': 'application/json'}},
            body: JSON.stringify({{rows: rows}})
          }});
          const data = await res.json();
          if (!res.ok || !data.ok) throw new Error(data.error || '저장 실패');
          box.textContent = '저장 완료: ' + data.saved + '명 / 건너뜀 ' + data.skipped + '명. 새로고침 중...';
          setTimeout(() => location.reload(), 600);
        }} catch (e) {{
          box.textContent = '저장 실패: ' + e.message;
        }}
      }}
      </script>

      <div style="margin-top:14px; overflow:auto; border:1px solid #eee; border-radius:12px;">
        <table style="border-collapse:collapse; width:100%; min-width:2350px;">
          <thead>
            <tr style="background:#fafafa;">
              <th style="padding:10px; border-bottom:1px solid #eee; text-align:right; color:#999;">#</th>
              <th style="padding:10px; border-bottom:1px solid #eee; text-align:left;">이름 / 오늘비율</th>
              <th style="padding:10px; border-bottom:1px solid #eee; text-align:left;">로그인 설정</th>
              <th style="padding:10px; border-bottom:1px solid #eee; text-align:left;">배민 입사일</th>
              <th style="padding:10px; border-bottom:1px solid #eee; text-align:left;">기준일(수정가능)</th>
              <th style="padding:10px; border-bottom:1px solid #eee; text-align:left;">평가기간</th>
              <th style="padding:10px; border-bottom:1px solid #eee; text-align:center;">오늘운행</th>
              <th style="padding:10px; border-bottom:1px solid #eee; text-align:right;">완료(현재)</th>
              <th style="padding:10px; border-bottom:1px solid #eee; text-align:center;">현재등급(이전)</th>
              <th style="padding:10px; border-bottom:1px solid #eee; text-align:center;">예정등급(현재)</th>
              <th style="padding:10px; border-bottom:1px solid #eee; text-align:center;">다음등급</th>
              <th style="padding:10px; border-bottom:1px solid #eee; text-align:right;">남은건수</th>
              <th style="padding:10px; border-bottom:1px solid #eee; text-align:left;">이전등급 플러스</th>
              <th style="padding:10px; border-bottom:1px solid #eee; text-align:left;">예정등급 플러스</th>
            </tr>
          </thead>
          <tbody>
            {tr_html if tr_html else '<tr><td colspan="14" style="padding:14px; color:#777;">조회 결과가 없습니다.</td></tr>'}
          </tbody>
        </table>
      </div>
    </div>
    """
    return html_page("전체 등급 현황", body)


# -----------------------------
# Excel export
# -----------------------------
@app.get("/dashboard.xlsx")
def dashboard_excel(request: Request):
    r = require_admin(request)
    if r:
        return r

    if not store_ready():
        return not_ready_page()

    riders = fetch_riders_cached()
    prevplus_map = load_prevplus_map()
    plannedplus_map = load_plannedplus_map()
    sticker_map = load_sticker_map()
    today_stats_map = build_today_stats_map(fetch_delivery_status_cached())

    today = operation_date()
    cur_group: Dict[Tuple[date, date], List[Dict[str, Any]]] = {}
    prev_group: Dict[Tuple[date, date], List[Dict[str, Any]]] = {}

    for rr in riders:
        nm = rr.get("name") or ""
        ph = rr.get("phoneNumber") or ""
        real4 = last4_from_phone(ph)
        if not real4:
            continue

        nn = norm_name(nm)
        login4, _, _ = get_login4_for_rider(rr)
        real_key = f"{nn}|{real4}"
        login_key = f"{nn}|{login4}"

        eff_join, _ = get_effective_join_date_by_login_key(rr, login4)

        cur_start, cur_end_incl = current_period(eff_join, today)
        cur_from, cur_to = period_to_from_to(cur_start, cur_end_incl)

        prev_end_incl = cur_start - timedelta(days=1)
        prev_m = date(cur_start.year, cur_start.month, 1) + relativedelta(months=-1)
        prev_start = clamp_day(prev_m.year, prev_m.month, eff_join.day)
        prev_from, prev_to = period_to_from_to(prev_start, prev_end_incl)

        it = {
            "name": nm,
            "login4": login4,
            "real4": real4,
            "real_key": real_key,
            "login_key": login_key,
            "join_effective": eff_join.isoformat(),
            "policy_from": cur_start.isoformat(),
            "policy_to": cur_end_incl.isoformat(),
            "api_from": cur_from.isoformat(),
            "api_to": cur_to.isoformat(),
            "prev_api_from": prev_from.isoformat(),
            "prev_api_to": prev_to.isoformat(),
            "prev_plus": int(prevplus_map.get(login_key, 0) or 0),
            "planned_plus": int(plannedplus_map.get(login_key, 0) or 0),
            "sticker_attached": 1 if bool(sticker_map.get(login_key, False)) else 0,
            "sticker_bonus": 20 if bool(sticker_map.get(login_key, False)) else 0,
        }

        cur_group.setdefault((cur_from, cur_to), []).append(it)
        prev_group.setdefault((prev_from, prev_to), []).append(it)

    prev_completed_map: Dict[str, int] = {}
    for (from_d, to_d), items in prev_group.items():
        if from_d > to_d:
            for it in items:
                prev_completed_map[it["real_key"]] = 0
            continue

        cmap = fetch_status_complete_map_cached(from_d, to_d)
        for it in items:
            prev_completed_map[it["real_key"]] = int(cmap.get(it["real_key"], 0))

    rows = []
    for (from_d, to_d), items in cur_group.items():
        if from_d > to_d:
            cmap = {}
        else:
            cmap = fetch_status_complete_map_cached(from_d, to_d)

        for it in items:
            cur_raw = int(cmap.get(it["real_key"], 0))
            prev_raw = int(prev_completed_map.get(it["real_key"], 0))
            prev_plus = int(it["prev_plus"])
            planned_plus = int(it["planned_plus"])

            prev_total_for_grade = prev_raw + prev_plus
            planned_total_for_grade = cur_raw + planned_plus

            planned_grade = grade_from_total(planned_total_for_grade)
            current_grade = grade_from_total(prev_total_for_grade)
            nxt, remain = next_grade_target(planned_total_for_grade)

            today_info = today_stats_map.get(it["real_key"]) or {
                "complete": 0,
                "reject": 0,
                "cancel": 0,
                **calc_bad_ratio(0, 0, 0),
            }

            rows.append({
                "name": it["name"],
                "login4": it["login4"],
                "real4": it["real4"],
                "join_effective": it["join_effective"],
                "policy_from": it["policy_from"],
                "policy_to": it["policy_to"],
                "api_from": it["api_from"],
                "api_to": it["api_to"],
                "today_complete": int(today_info["complete"]),
                "today_reject": int(today_info["reject"]),
                "today_cancel": int(today_info["cancel"]),
                "today_ratio": float(today_info["ratio"]),
                "cur_completed": cur_raw,
                "prev_plus": prev_plus,
                "planned_plus": planned_plus,
                "planned_total_for_grade": planned_total_for_grade,
                "planned_grade_cur": planned_grade,
                "prev_completed": prev_raw,
                "prev_total_for_grade": prev_total_for_grade,
                "current_grade_prev": current_grade,
                "next": nxt or "",
                "remain": remain if remain is not None else "",
            })

    rows.sort(key=lambda x: x["cur_completed"], reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "dashboard"

    headers = [
        "name", "login4", "real4", "join_effective",
        "policy_from", "policy_to", "api_from", "api_to",
        "today_complete", "today_reject", "today_cancel", "today_ratio",
        "cur_completed",
        "prev_plus", "planned_plus",
        "planned_total_for_grade", "planned_grade_cur",
        "prev_completed", "prev_total_for_grade", "current_grade_prev",
        "next", "remain",
    ]
    ws.append(headers)
    for r0 in rows:
        ws.append([r0.get(h, "") for h in headers])

    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, min(28, len(h) + 4))

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"riderwelfare_dashboard_{today_kst().isoformat()}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# -----------------------------
# Roulette API
# -----------------------------
@app.post("/api/roulette/status")
def api_roulette_status(payload: dict = Body(...)):
    phone = normalize_phone(payload.get("phone", ""))
    if not phone:
        raise HTTPException(status_code=400, detail="phone required")

    today_completed, rider_name = get_today_completed_for_phone(phone)
    return roulette_db.get_status(phone, rider_name, today_completed)


@app.post("/api/roulette/spin")
def api_roulette_spin(payload: dict = Body(...)):
    phone = normalize_phone(payload.get("phone", ""))
    if not phone:
        raise HTTPException(status_code=400, detail="phone required")

    today_completed, rider_name = get_today_completed_for_phone(phone)

    try:
        return roulette_db.spin(phone, rider_name, today_completed)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/admin/api/roulette/settings")
def admin_get_roulette_settings(request: Request):
    r = require_admin(request)
    if r:
        raise HTTPException(status_code=401, detail="admin required")

    return {
        "enabled": roulette_db.get_enabled(),
        "spin_unit": roulette_db.get_spin_unit(),
        "max_segment_value": roulette_db.get_max_segment_value(),
        "segment_rewards": roulette_db.get_all_segment_rewards(),
        "current_cycle": get_operational_cycle_bounds(),
        "previous_cycle": get_previous_operational_cycle_bounds(),
    }


@app.post("/admin/api/roulette/settings")
def admin_save_roulette_settings(request: Request, payload: dict = Body(...)):
    r = require_admin(request)
    if r:
        raise HTTPException(status_code=401, detail="admin required")

    try:
        enabled = bool(payload.get("enabled", True))
        spin_unit = int(payload.get("spin_unit", 10))
        max_segment_value = int(payload.get("max_segment_value", 100))
        segment_rewards = payload.get("segment_rewards", [])

        roulette_db.set_enabled(enabled)
        roulette_db.set_spin_unit(spin_unit)
        roulette_db.set_max_segment_value(max_segment_value)
        roulette_db.set_segment_rewards(segment_rewards)

        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/admin/roulette", response_class=HTMLResponse)
def admin_roulette_page(request: Request):
    r = require_admin(request)
    if r:
        return r

    current_cycle = get_operational_cycle_bounds()
    prev_cycle = get_previous_operational_cycle_bounds()

    body = f"""
    <div style="background:#fff; border:1px solid #e8e8e8; border-radius:16px; padding:16px;">
      <div style="display:flex; justify-content:space-between; align-items:flex-start; gap:12px; flex-wrap:wrap;">
        <div>
          <h2 style="margin:0 0 6px 0;">룰렛 관리자</h2>
          <div style="color:#666; line-height:1.7;">
            현재 운영주차: <b>{current_cycle["display_start"]}</b> ~ <b>{current_cycle["display_end"]}</b><br/>
            직전 마감 주차: <b>{prev_cycle["display_start"]}</b> ~ <b>{prev_cycle["display_end"]}</b>
          </div>
        </div>

        <div style="display:flex; gap:8px; flex-wrap:wrap;">
          <a href="/dashboard" style="text-decoration:none; color:#111;">전체 대시보드</a>
          <a href="/admin/api/roulette/payout.csv" style="text-decoration:none; color:#111;">직전주 지급 CSV</a>
          <a href="/admin/api/roulette/export" style="text-decoration:none; color:#111;">JSON 내보내기</a>
          <a href="/admin-logout" style="text-decoration:none; color:#666;">로그아웃</a>
        </div>
      </div>

      <div style="margin-top:16px; border:1px solid #eee; border-radius:14px; padding:14px; background:#fcfcfc;">
          <div style="font-weight:900; margin-bottom:10px;">룰렛 복원 업로드</div>
          <div style="color:#666; font-size:13px; line-height:1.6; margin-bottom:10px;">
            export JSON 또는 백업 JSON 파일을 업로드해서 룰렛 설정/당첨내역을 복원합니다.
          </div>

        <form method="post" action="/admin/api/roulette/import" enctype="multipart/form-data"
                     style="display:flex; gap:8px; flex-wrap:wrap; align-items:center;">
            <input type="file" id="rouletteImportFile" name="file" accept=".json"
                       style="font-size:14px;" required />
            <button type="submit"
                         style="padding:10px 14px; border:none; border-radius:12px; background:#111; color:#fff; font-weight:900; cursor:pointer;">
                업로드 복원
             </button>
         </form>

  <div id="rouletteImportMsg" style="margin-top:10px; color:#666; font-size:13px;"></div>
</div>

      <div style="margin-top:16px; display:grid; grid-template-columns:1fr 1fr; gap:12px;">
        <div style="border:1px solid #eee; border-radius:14px; padding:14px; background:#fcfcfc;">
          <div style="font-weight:900; margin-bottom:10px;">기본 설정</div>
          <div style="display:flex; gap:12px; flex-wrap:wrap; align-items:center;">
            <label style="display:flex; gap:6px; align-items:center;">
              <input type="checkbox" id="rouletteEnabled" />
              <span>룰렛 사용</span>
            </label>

            <label style="display:flex; gap:6px; align-items:center;">
              <span>기준</span>
              <input type="number" id="spinUnit" min="1" value="10"
                     style="width:90px; padding:8px 10px; border:1px solid #ddd; border-radius:10px;" />
              <span>건당 1회</span>
            </label>

            <label style="display:flex; gap:6px; align-items:center;">
              <span>최대 구간</span>
              <input type="number" id="maxSegmentValue" min="10" max="100" step="10" value="100"
                     style="width:90px; padding:8px 10px; border:1px solid #ddd; border-radius:10px;" />
            </label>

            <button onclick="downloadRouletteHistoryCsv()"
                    style="padding:10px 14px; border:1px solid #ddd; border-radius:12px; background:#fff; color:#111; font-weight:900;">
                CSV 다운로드
            </button>

            <button onclick="saveRouletteSettings()"
                    style="padding:10px 14px; border:none; border-radius:12px; background:#111; color:#fff; font-weight:900;">
              설정 저장
            </button>

            <button onclick="backupRouletteDb()"
                    style="padding:10px 14px; border:1px solid #ddd; border-radius:12px; background:#fff; color:#111; font-weight:900;">
              DB 백업
            </button>
          </div>
          <div id="saveMsg" style="margin-top:10px; color:#666; font-size:13px;"></div>
        </div>

        <div style="border:1px solid #eee; border-radius:14px; padding:14px; background:#fcfcfc;">
          <div style="font-weight:900; margin-bottom:10px;">지급 다운로드</div>
          <div style="display:flex; gap:8px; flex-wrap:wrap; align-items:center;">
            <input type="text" id="weekKeyInput" placeholder="예: 2026-03-04"
                   style="width:180px; padding:8px 10px; border:1px solid #ddd; border-radius:10px;" />
            <button onclick="downloadPayoutCsv()"
                    style="padding:10px 14px; border:none; border-radius:12px; background:#111; color:#fff; font-weight:900;">
              주차 지정 CSV
            </button>
            <button onclick="downloadPrevPayoutCsv()"
                    style="padding:10px 14px; border:1px solid #ddd; border-radius:12px; background:#fff; color:#111; font-weight:900;">
              직전주 CSV
            </button>
          </div>
          <div style="margin-top:8px; color:#888; font-size:12px;">
            비워두면 직전 마감 운영주차 기준으로 다운로드됩니다.
          </div>
        </div>
      </div>

      <div style="margin-top:16px; border:1px solid #eee; border-radius:14px; padding:14px; background:#fff;">
        <div style="display:flex; justify-content:space-between; align-items:center; gap:10px; flex-wrap:wrap;">
          <div style="font-weight:900;">구간별 확률 설정</div>
          <div style="color:#777; font-size:13px;">10건 ~ 100건 구간별로 각각 수정 가능</div>
        </div>
        <div id="segmentRewardWrap" style="margin-top:12px; overflow:auto;"></div>
      </div>

      <div style="margin-top:16px; border:1px solid #eee; border-radius:14px; padding:14px; background:#fff;">
        <div style="display:flex; justify-content:space-between; align-items:center; gap:10px; flex-wrap:wrap;">
          <div style="font-weight:900;">당첨 내역</div>
          <div id="historySummary" style="color:#666; font-size:13px;"></div>
        </div>

        <div style="margin-top:12px; display:flex; gap:8px; flex-wrap:wrap; align-items:center;">
          <input type="date" id="startDate"
                 style="padding:8px 10px; border:1px solid #ddd; border-radius:10px;" />
          <input type="date" id="endDate"
                 style="padding:8px 10px; border:1px solid #ddd; border-radius:10px;" />
          <input type="text" id="keyword" placeholder="이름 또는 전화번호"
                 style="width:220px; padding:8px 10px; border:1px solid #ddd; border-radius:10px;" />
          <button onclick="loadRouletteHistory()"
                  style="padding:10px 14px; border:none; border-radius:12px; background:#111; color:#fff; font-weight:900;">
            조회
          </button>
        </div>

        <div style="margin-top:12px; overflow:auto; border:1px solid #eee; border-radius:12px;">
          <table style="border-collapse:collapse; width:100%; min-width:1100px;">
            <thead>
              <tr style="background:#fafafa;">
                <th style="padding:10px; border-bottom:1px solid #eee;">ID</th>
                <th style="padding:10px; border-bottom:1px solid #eee;">날짜</th>
                <th style="padding:10px; border-bottom:1px solid #eee;">시간</th>
                <th style="padding:10px; border-bottom:1px solid #eee;">이름</th>
                <th style="padding:10px; border-bottom:1px solid #eee;">전화번호</th>
                <th style="padding:10px; border-bottom:1px solid #eee;">오늘완료</th>
                <th style="padding:10px; border-bottom:1px solid #eee;">회차</th>
                <th style="padding:10px; border-bottom:1px solid #eee;">구간</th>
                <th style="padding:10px; border-bottom:1px solid #eee;">당첨금</th>
                <th style="padding:10px; border-bottom:1px solid #eee;">비고</th>
                <th style="padding:10px; border-bottom:1px solid #eee;">수정</th>
                <th style="padding:10px; border-bottom:1px solid #eee;">삭제</th>
              </tr>
            </thead>
            <tbody id="historyBody">
              <tr><td colspan="12" style="padding:14px; color:#777;">불러오는 중...</td></tr>
            </tbody>
          </table>
        </div>
      </div>
    </div>

    <script>
      function fmtMoney(v) {{
        return Number(v || 0).toLocaleString() + '원';
      }}

      function fmtTs(ts) {{
        if (!ts) return '-';
        const d = new Date(Number(ts) * 1000);
        const hh = String(d.getHours()).padStart(2, '0');
        const mm = String(d.getMinutes()).padStart(2, '0');
        const ss = String(d.getSeconds()).padStart(2, '0');
        return `${{hh}}:${{mm}}:${{ss}}`;
      }}

      function segmentList() {{
        const arr = [];
        for (let i = 10; i <= 100; i += 10) arr.push(i);
        return arr;
      }}

      async function loadRouletteSettings() {{
        const res = await fetch('/admin/api/roulette/settings');
        const data = await res.json();

        document.getElementById('rouletteEnabled').checked = !!data.enabled;
        document.getElementById('spinUnit').value = data.spin_unit || 10;
        document.getElementById('maxSegmentValue').value = data.max_segment_value || 100;

        const grouped = {{}};
        (data.segment_rewards || []).forEach(r => {{
          const seg = Number(r.segment_value || 0);
          if (!grouped[seg]) grouped[seg] = [];
          grouped[seg].push(r);
        }});

        const wrap = document.getElementById('segmentRewardWrap');
        let html = `<table style="border-collapse:collapse; width:100%; min-width:900px;">
          <thead>
            <tr style="background:#fafafa;">
              <th style="padding:10px; border-bottom:1px solid #eee;">구간</th>
              <th style="padding:10px; border-bottom:1px solid #eee;">1000원</th>
              <th style="padding:10px; border-bottom:1px solid #eee;">2000원</th>
              <th style="padding:10px; border-bottom:1px solid #eee;">3000원</th>
              <th style="padding:10px; border-bottom:1px solid #eee;">5000원</th>
              <th style="padding:10px; border-bottom:1px solid #eee;">10000원</th>
            </tr>
          </thead>
          <tbody>`;

        const amounts = [1000, 2000, 3000, 5000, 10000];

        segmentList().forEach(seg => {{
          const rows = grouped[seg] || [];
          const byAmount = {{}};
          rows.forEach(r => byAmount[Number(r.amount)] = r);

          html += `<tr>
            <td style="padding:10px; border-bottom:1px solid #eee; font-weight:900;">${{seg}}건</td>`;

          amounts.forEach(amount => {{
            const row = byAmount[amount] || {{ weight: 0, active: 1 }};
            html += `
              <td style="padding:10px; border-bottom:1px solid #eee; text-align:center;">
                <div style="display:flex; flex-direction:column; gap:6px; align-items:center;">
                  <input type="number"
                         data-seg="${{seg}}"
                         data-amount="${{amount}}"
                         class="weight-input"
                         value="${{row.weight || 0}}"
                         style="width:90px; padding:8px 10px; border:1px solid #ddd; border-radius:10px;" />
                  <label style="font-size:12px; color:#666; display:flex; gap:4px; align-items:center;">
                    <input type="checkbox"
                           data-seg="${{seg}}"
                           data-amount="${{amount}}"
                           class="active-input"
                           ${{row.active ? 'checked' : ''}} />
                    사용
                  </label>
                </div>
              </td>`;
          }});

          html += `</tr>`;
        }});

        html += `</tbody></table>`;
        wrap.innerHTML = html;
      }}

      async function saveRouletteSettings() {{
        const enabled = document.getElementById('rouletteEnabled').checked;
        const spin_unit = Number(document.getElementById('spinUnit').value || 10);
        const max_segment_value = Number(document.getElementById('maxSegmentValue').value || 100);

        const payloadRows = [];
        const amounts = [1000, 2000, 3000, 5000, 10000];

        segmentList().forEach(seg => {{
          amounts.forEach((amount, idx) => {{
            const w = document.querySelector(`.weight-input[data-seg="${{seg}}"][data-amount="${{amount}}"]`);
            const a = document.querySelector(`.active-input[data-seg="${{seg}}"][data-amount="${{amount}}"]`);
            payloadRows.push({{
              segment_value: seg,
              amount: amount,
              weight: Number(w?.value || 0),
              active: !!a?.checked,
              sort_order: idx + 1
            }});
          }});
        }});

        const res = await fetch('/admin/api/roulette/settings', {{
          method: 'POST',
          headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{
            enabled,
            spin_unit,
            max_segment_value,
            segment_rewards: payloadRows
          }})
        }});

        const data = await res.json();
        document.getElementById('saveMsg').innerText = data.ok ? '저장 완료' : (data.detail || '저장 실패');
        if (data.ok) loadRouletteSettings();
      }}

      async function loadRouletteHistory() {{
        const start_date = document.getElementById('startDate').value;
        const end_date = document.getElementById('endDate').value;
        const keyword = document.getElementById('keyword').value.trim();

        const qs = new URLSearchParams();
        if (start_date) qs.set('start_date', start_date);
        if (end_date) qs.set('end_date', end_date);
        if (keyword) qs.set('keyword', keyword);
        qs.set('limit', '300');

        const res = await fetch('/admin/api/roulette/history?' + qs.toString());
        const data = await res.json();

        const tbody = document.getElementById('historyBody');
        const summary = document.getElementById('historySummary');
        summary.innerText = `조회건수 ${{Number(data.summary?.count || 0)}}건 / 총 당첨금 ${{fmtMoney(data.summary?.total_amount || 0)}}`;

        const rows = data.rows || [];
        if (!rows.length) {{
          tbody.innerHTML = '<tr><td colspan="12" style="padding:14px; color:#777;">조회 결과가 없습니다.</td></tr>';
          return;
        }}

        let html = '';
        rows.forEach(row => {{
          html += `
            <tr>
              <td style="padding:10px; border-bottom:1px solid #eee; text-align:center;">${{row.id}}</td>
              <td style="padding:10px; border-bottom:1px solid #eee; text-align:center;">${{row.spin_date || '-'}}</td>
              <td style="padding:10px; border-bottom:1px solid #eee; text-align:center;">${{fmtTs(row.created_at)}}</td>
              <td style="padding:10px; border-bottom:1px solid #eee; text-align:center;">${{row.rider_name || ''}}</td>
              <td style="padding:10px; border-bottom:1px solid #eee; text-align:center;">${{row.phone || ''}}</td>
              <td style="padding:10px; border-bottom:1px solid #eee; text-align:center;">${{row.today_completed || 0}}</td>
              <td style="padding:10px; border-bottom:1px solid #eee; text-align:center;">${{row.spin_index || 0}}회차</td>
              <td style="padding:10px; border-bottom:1px solid #eee; text-align:center;">${{row.segment_value || 0}}건</td>
              <td style="padding:10px; border-bottom:1px solid #eee; text-align:center;">
                <input type="number" id="amt_${{row.id}}" value="${{row.reward_amount || 0}}"
                       style="width:100px; padding:8px 10px; border:1px solid #ddd; border-radius:10px;" />
              </td>
              <td style="padding:10px; border-bottom:1px solid #eee; text-align:center;">
                <input type="text" id="note_${{row.id}}" value="${{row.note || ''}}"
                       style="width:180px; padding:8px 10px; border:1px solid #ddd; border-radius:10px;" />
              </td>
              <td style="padding:10px; border-bottom:1px solid #eee; text-align:center;">
                <button onclick="updateRouletteSpin(${{row.id}})"
                        style="padding:8px 12px; border:none; border-radius:10px; background:#111; color:#fff;">
                  수정
                </button>
              </td>
              <td style="padding:10px; border-bottom:1px solid #eee; text-align:center;">
                <button onclick="deleteRouletteSpin(${{row.id}})"
                        style="padding:8px 12px; border:1px solid #ddd; border-radius:10px; background:#fff; color:#111;">
                  삭제
                </button>
              </td>
            </tr>`;
        }});

        tbody.innerHTML = html;
      }}

      async function updateRouletteSpin(id) {{
        const reward_amount = Number(document.getElementById('amt_' + id).value || 0);
        const note = document.getElementById('note_' + id).value || '';

        const res = await fetch('/admin/api/roulette/spin/update', {{
          method: 'POST',
          headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{ spin_id: id, reward_amount, note }})
        }});
        const data = await res.json();
        alert(data.ok ? '수정 완료' : (data.detail || '수정 실패'));
        if (data.ok) loadRouletteHistory();
      }}

      async function deleteRouletteSpin(id) {{
        if (!confirm('정말 삭제하시겠습니까?')) return;

        const res = await fetch('/admin/api/roulette/spin/delete', {{
          method: 'POST',
          headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{ spin_id: id }})
        }});
        const data = await res.json();
        alert(data.ok ? '삭제 완료' : (data.detail || '삭제 실패'));
        if (data.ok) loadRouletteHistory();
      }}

      async function backupRouletteDb() {{
        const res = await fetch('/admin/api/roulette/backup', {{ method: 'POST' }});
        const data = await res.json();
        alert(data.ok ? ('백업 완료: ' + data.backup_name) : (data.detail || '백업 실패'));
      }}

      function downloadPrevPayoutCsv() {{
        window.location.href = '/admin/api/roulette/payout.csv';
      }}

      function downloadPayoutCsv() {{
        const weekKey = document.getElementById('weekKeyInput').value.trim();
        if (!weekKey) {{
          window.location.href = '/admin/api/roulette/payout.csv';
          return;
        }}
        window.location.href = '/admin/api/roulette/payout.csv?week_key=' + encodeURIComponent(weekKey);
      }}

      loadRouletteSettings();
      loadRouletteHistory();
    </script>
    """
    return html_page("룰렛 관리자", body)

@app.get("/admin/api/roulette/history")
def admin_get_roulette_history(
    request: Request,
    start_date: str = "",
    end_date: str = "",
    keyword: str = "",
    limit: int = 300,
):
    r = require_admin(request)
    if r:
        raise HTTPException(status_code=401, detail="admin required")

    rows = roulette_db.get_history(
        start_date=start_date or None,
        end_date=end_date or None,
        keyword=keyword or "",
        limit=limit,
    )
    summary = roulette_db.get_history_summary(
        start_date=start_date or None,
        end_date=end_date or None,
        keyword=keyword or "",
    )
    return {"rows": rows, "summary": summary}


@app.post("/admin/api/roulette/spin/update")
def admin_update_spin(request: Request, payload: dict = Body(...)):
    r = require_admin(request)
    if r:
        raise HTTPException(status_code=401, detail="admin required")

    try:
        spin_id = int(payload.get("spin_id"))
        reward_amount = int(payload.get("reward_amount"))
        note = str(payload.get("note", ""))
        row = roulette_db.update_spin(spin_id, reward_amount, note)
        return {"ok": True, "row": row}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/admin/api/roulette/spin/delete")
def admin_delete_spin(request: Request, payload: dict = Body(...)):
    r = require_admin(request)
    if r:
        raise HTTPException(status_code=401, detail="admin required")

    try:
        spin_id = int(payload.get("spin_id"))
        roulette_db.delete_spin(spin_id)
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/admin/api/roulette/backup")
def admin_roulette_backup(request: Request):
    r = require_admin(request)
    if r:
        raise HTTPException(status_code=401, detail="admin required")

    try:
        return roulette_db.create_backup()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/admin/api/roulette/import")
async def admin_roulette_import(request: Request, file: UploadFile = File(...)):
    r = require_admin(request)
    if r:
        raise HTTPException(status_code=401, detail="admin required")

    try:
        if not file.filename.lower().endswith(".json"):
            raise HTTPException(status_code=400, detail="json 파일만 업로드 가능합니다.")

        backup_info = roulette_db.create_backup()

        raw = await file.read()
        payload = json.loads(raw.decode("utf-8"))

        result = roulette_db.import_json(payload)
        return {
            "ok": True,
            "filename": file.filename,
            "backup_name": backup_info.get("backup_name", ""),
            **result,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/admin/api/roulette/payout.csv")
@app.get("/admin/api/roulette/history.csv")
def admin_roulette_history_csv(
    request: Request,
    start_date: str = "",
    end_date: str = "",
    keyword: str = "",
    limit: int = 5000,
):
    r = require_admin(request)
    if r:
        raise HTTPException(status_code=401, detail="admin required")

    try:
        rows = roulette_db.get_history(
            start_date=start_date or None,
            end_date=end_date or None,
            keyword=keyword or "",
            limit=limit,
        )

        output = StringIO()
        writer = csv.writer(output)
        writer.writerow([
            "id",
            "spin_date",
            "time",
            "rider_name",
            "phone",
            "today_completed",
            "spin_index",
            "segment_value",
            "reward_amount",
            "note",
            "created_at",
        ])

        for row in rows:
            created_at = int(row.get("created_at") or 0)
            time_text = "-"
            if created_at:
                try:
                    time_text = kst_from_epoch(created_at).strftime("%H:%M:%S")
                except Exception:
                    time_text = "-"

            writer.writerow([
                row.get("id", ""),
                row.get("spin_date", ""),
                time_text,
                row.get("rider_name", ""),
                row.get("phone", ""),
                row.get("today_completed", 0),
                row.get("spin_index", 0),
                row.get("segment_value", 0),
                row.get("reward_amount", 0),
                row.get("note", ""),
                row.get("created_at", 0),
            ])

        s = start_date or "all"
        e = end_date or "all"
        filename = f"roulette_history_{s}_{e}.csv"

        return Response(
            content=output.getvalue(),
            media_type="text/csv; charset=utf-8-sig",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# -----------------------------
# Full backup export/import
# -----------------------------
# DISABLED_BY_FORCE_REAL_ZIP_EXPORT
# # DISABLED_BY_FORCE_REAL_ZIP_EXPORT
# @app.get("/backup/export")
def backup_export(request: Request):
    # 로컬 PC 자동백업용: x-ingest-token 필요
    auth = _require_ingest(request)
    if auth:
        return auth
    return get_backup_payload()


# DISABLED_BY_ZIP_IMPORT_FIX
# # DISABLED_BY_ROULETTE_ZIP_IMPORT_FIX
# @app.post("/backup/import")
async def backup_import(request: Request, file: UploadFile = File(...)):
    # 로컬/긴급 복원용: x-ingest-token 필요
    auth = _require_ingest(request)
    if auth:
        return auth
    raw = await file.read()
    payload = json.loads(raw.decode("utf-8"))
    return restore_backup_payload(payload)


# DISABLED_BY_FORCE_REAL_ZIP_EXPORT
# # DISABLED_BY_FORCE_REAL_ZIP_EXPORT
# @app.get("/admin/backup/export")
def admin_backup_export(request: Request):
    r = require_admin(request)
    if r:
        return r
    payload = get_backup_payload()
    filename = f"riderwelfare_full_backup_{today_kst().isoformat()}.json"
    return Response(
        content=json.dumps(payload, ensure_ascii=False, indent=2),
        media_type="application/json; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# DISABLED_BY_ZIP_IMPORT_FIX
# # DISABLED_BY_ROULETTE_ZIP_IMPORT_FIX
# @app.post("/admin/backup/import")
async def admin_backup_import(request: Request, file: UploadFile = File(...)):
    r = require_admin(request)
    if r:
        return r
    if not file.filename.lower().endswith(".json"):
        raise HTTPException(status_code=400, detail="json 백업 파일만 업로드 가능합니다.")
    raw = await file.read()
    payload = json.loads(raw.decode("utf-8"))
    result = restore_backup_payload(payload)
    body = f"""
    <div style="background:#fff;border:1px solid #e8e8e8;border-radius:16px;padding:16px;max-width:820px;margin:0 auto;">
      <h3 style="margin-top:0;">전체 백업 복원 완료</h3>
      <div style="color:#666;line-height:1.7;">
        복원 전 안전백업: <b>{result.get('backup_before_restore')}</b><br/>
        복원 파일: <b>{', '.join(result.get('restored_files') or [])}</b><br/>
        룰렛 복원: <b>{result.get('roulette_result')}</b>
      </div>
      <div style="margin-top:12px;"><a href="/dashboard" style="text-decoration:none;color:#111;">대시보드로 이동</a></div>
    </div>
    """
    return HTMLResponse(html_page("전체 백업 복원 완료", body))



# -----------------------------
# Diagnostics
# -----------------------------
@app.get("/health", response_class=HTMLResponse)
def health():
    ok_token = bool(INGEST_TOKEN)
    riders_ok = store_ready()
    riders_ts = _read_json(RIDERS_STORE, {}).get("ts")
    status_keys = list((_read_json(STATUS_STORE, {}) or {}).keys())[:5]
    delivery_ts = _read_json(DELIVERY_STATUS_STORE, {}).get("ts")

    pcx_from = NMAX_START_DATE
    pcx_to = operation_date() - timedelta(days=1)
    pcx_ready = (pcx_from <= pcx_to) and has_status_range(pcx_from, pcx_to)

    roulette_export = roulette_db.export_json()
    body = f"""
    <div style="background:#fff; border:1px solid #e8e8e8; border-radius:16px; padding:16px; max-width:820px; margin:0 auto;">
      <h3 style="margin-top:0;">Health</h3>
      <div>ingest_token_set: <b>{ok_token}</b></div>
      <div>riders_uploaded: <b>{riders_ok}</b></div>
      <div>riders_ts: <b>{riders_ts}</b></div>
      <div>delivery_status_ts: <b>{delivery_ts}</b></div>
      <div style="margin-top:8px;">status_keys(sample): <b>{status_keys}</b></div>
      <div style="margin-top:8px;">pcx_range_ready({NMAX_START_DATE.isoformat()}~어제): <b>{pcx_ready}</b></div>
      <div style="margin-top:8px;">roulette_enabled: <b>{roulette_db.get_enabled()}</b></div>
      <div style="margin-top:8px;">roulette_segment_rows: <b>{len(roulette_export.get('segment_rewards', []))}</b></div>
      <div style="margin-top:12px;"><a href="/" style="text-decoration:none; color:#111;">&larr; 홈</a></div>
    </div>
    """
    return html_page("Health", body)


@app.get("/version")
def version():
    return {
        "ok": True,
        "ts": int(time.time()),
        "render_git": os.getenv("RENDER_GIT_COMMIT", ""),
    }



# ============================================================
# ZIP FULL BACKUP IMPORT FIX
# ZIP 파일 그대로 업로드 복원 가능
# ============================================================

def _restore_zip_backup_bytes(data: bytes):
    import zipfile
    from io import BytesIO
    from pathlib import Path

    restored = 0
    skipped = 0
    allowed_ext = {".json", ".db", ".sqlite", ".sqlite3"}

    target_root = Path(STORE_DIR)
    target_root.mkdir(parents=True, exist_ok=True)

    with zipfile.ZipFile(BytesIO(data), "r") as z:
        for info in z.infolist():
            if info.is_dir():
                continue

            raw_name = info.filename.replace("\\", "/")
            filename = Path(raw_name).name

            if filename == "backup_info.json":
                skipped += 1
                continue

            suffix = Path(filename).suffix.lower()
            if suffix not in allowed_ext:
                skipped += 1
                continue

            out_path = target_root / filename
            out_path.write_bytes(z.read(info.filename))
            restored += 1

    try:
        _riders_cache["ts"] = 0.0
        _riders_cache["data"] = None
    except Exception:
        pass

    try:
        _status_cache.clear()
    except Exception:
        pass

    try:
        _delivery_status_cache["ts"] = 0.0
        _delivery_status_cache["data"] = None
    except Exception:
        pass

    return {"ok": True, "restored": restored, "skipped": skipped, "type": "zip"}


# DISABLED_BY_ROULETTE_ZIP_IMPORT_FIX
# @app.post("/backup/import")
async def backup_import_zip(request: Request, file: UploadFile = File(...)):
    auth = _require_ingest(request)
    if auth:
        return auth

    filename = (file.filename or "").lower()
    data = await file.read()

    if not filename.endswith(".zip"):
        return JSONResponse(
            {"ok": False, "error": "ZIP 파일만 업로드 가능합니다."},
            status_code=400
        )

    try:
        return _restore_zip_backup_bytes(data)
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=400)


# DISABLED_BY_ROULETTE_ZIP_IMPORT_FIX
# @app.post("/admin/backup/import")
async def admin_backup_import_zip(request: Request, file: UploadFile = File(...)):
    admin_redirect = require_admin(request)
    if admin_redirect:
        return admin_redirect

    filename = (file.filename or "").lower()
    data = await file.read()

    if not filename.endswith(".zip"):
        return JSONResponse(
            {"ok": False, "error": "ZIP 파일만 업로드 가능합니다."},
            status_code=400
        )

    try:
        result = _restore_zip_backup_bytes(data)
        body = f"""
        <div style="background:#fff;border:1px solid #e8e8e8;border-radius:16px;padding:16px;max-width:720px;margin:0 auto;">
          <h3 style="margin-top:0;">백업 복원 완료</h3>
          <div style="line-height:1.7;color:#333;">
            복원 파일 수: <b>{result.get('restored')}</b><br/>
            제외 파일 수: <b>{result.get('skipped')}</b><br/>
            형식: <b>ZIP</b>
          </div>
          <div style="margin-top:14px;">
            <a href="/health" style="color:#111;text-decoration:none;">Health 확인</a>
            &nbsp; | &nbsp;
            <a href="/dashboard" style="color:#111;text-decoration:none;">대시보드 확인</a>
          </div>
        </div>
        """
        return HTMLResponse(html_page("백업 복원 완료", body))
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=400)



# ============================================================
# ROULETTE ZIP IMPORT FIX
# - ZIP 복원 시 룰렛 당첨내역까지 roulette_db에 직접 import
# - 자동백업 파일이 .zip 확장자지만 실제 JSON이어도 복원 가능
# ============================================================

def _restore_full_backup_any_bytes(data: bytes, filename: str = ""):
    import json
    import zipfile
    from io import BytesIO
    from pathlib import Path

    restored = 0
    skipped = 0
    roulette_result = None

    try:
        as_text = data.decode("utf-8")
        payload = json.loads(as_text)
        if isinstance(payload, dict) and ("files" in payload or "roulette" in payload):
            try:
                result = restore_backup_payload(payload)
                result["type"] = "json_payload"
                return result
            except NameError:
                pass
    except Exception:
        pass

    try:
        zf = zipfile.ZipFile(BytesIO(data), "r")
    except Exception as e:
        raise ValueError(f"ZIP 또는 JSON 백업 파일이 아닙니다: {e}")

    allowed_ext = {".json", ".db", ".sqlite", ".sqlite3"}
    target_root = Path(STORE_DIR)
    target_root.mkdir(parents=True, exist_ok=True)

    roulette_spins = None
    roulette_payload = None

    with zf as z:
        for info in z.infolist():
            if info.is_dir():
                continue

            raw_name = info.filename.replace("\\", "/")
            file_name = Path(raw_name).name
            suffix = Path(file_name).suffix.lower()
            raw = z.read(info.filename)

            if file_name == "backup_info.json":
                skipped += 1
                continue

            if file_name in {
                "roulette_spins.json",
                "roulette_history_all.json",
                "source_roulette_history.json",
                "roulette_history.json",
            }:
                try:
                    loaded = json.loads(raw.decode("utf-8-sig"))
                    if isinstance(loaded, list):
                        roulette_spins = loaded
                    elif isinstance(loaded, dict) and isinstance(loaded.get("spins"), list):
                        roulette_spins = loaded.get("spins")
                    elif isinstance(loaded, dict) and isinstance(loaded.get("roulette"), dict):
                        roulette_payload = loaded.get("roulette")
                except Exception:
                    pass

            if suffix == ".json":
                try:
                    loaded = json.loads(raw.decode("utf-8-sig"))
                    if isinstance(loaded, dict) and isinstance(loaded.get("roulette"), dict):
                        roulette_payload = loaded.get("roulette")
                except Exception:
                    pass

            if suffix not in allowed_ext:
                skipped += 1
                continue

            try:
                out_path = target_root / file_name
                out_path.write_bytes(raw)
                restored += 1
            except Exception:
                skipped += 1

    try:
        if roulette_payload is not None and isinstance(roulette_payload, dict):
            roulette_result = roulette_db.import_json(roulette_payload)
        elif roulette_spins is not None and isinstance(roulette_spins, list):
            try:
                current = roulette_db.export_json()
            except Exception:
                current = {}
            if not isinstance(current, dict):
                current = {}
            current["spins"] = roulette_spins
            roulette_result = roulette_db.import_json(current)
    except Exception as e:
        roulette_result = {"ok": False, "error": str(e)}

    try:
        _riders_cache["ts"] = 0.0
        _riders_cache["data"] = None
    except Exception:
        pass

    try:
        _status_cache.clear()
    except Exception:
        pass

    try:
        _delivery_status_cache["ts"] = 0.0
        _delivery_status_cache["data"] = None
    except Exception:
        pass

    return {
        "ok": True,
        "type": "zip",
        "restored": restored,
        "skipped": skipped,
        "roulette_result": roulette_result,
    }


@app.post("/backup/import")
async def backup_import_any(request: Request, file: UploadFile = File(...)):
    auth = _require_ingest(request)
    if auth:
        return auth

    data = await file.read()
    try:
        return _restore_full_backup_any_bytes(data, file.filename or "")
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=400)


@app.post("/admin/backup/import")
async def admin_backup_import_any(request: Request, file: UploadFile = File(...)):
    r = require_admin(request)
    if r:
        return r

    data = await file.read()

    try:
        result = _restore_full_backup_any_bytes(data, file.filename or "")
        rr = result.get("roulette_result")
        body = (
            "<div style='background:#fff;border:1px solid #e8e8e8;border-radius:16px;padding:16px;max-width:820px;margin:0 auto;'>"
            "<h3 style='margin-top:0;'>전체 백업 복원 완료</h3>"
            "<div style='color:#666;line-height:1.7;'>"
            "복원 파일 수: <b>" + str(result.get("restored", "-")) + "</b><br/>"
            "제외 파일 수: <b>" + str(result.get("skipped", "-")) + "</b><br/>"
            "룰렛 복원 결과: <b>" + str(rr) + "</b>"
            "</div>"
            "<div style='margin-top:12px;'>"
            "<a href='/admin/roulette' style='text-decoration:none;color:#111;'>룰렛 관리자 확인</a>"
            "&nbsp; | &nbsp;"
            "<a href='/dashboard' style='text-decoration:none;color:#111;'>대시보드 확인</a>"
            "&nbsp; | &nbsp;"
            "<a href='/health' style='text-decoration:none;color:#111;'>Health 확인</a>"
            "</div></div>"
        )
        return HTMLResponse(html_page("전체 백업 복원 완료", body))
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=400)



# ============================================================
# FORCE REAL ZIP BACKUP EXPORT
# /backup/export 가 JSON이 아니라 진짜 ZIP을 반환하게 강제
# ============================================================

def _make_force_full_backup_zip():
    import json
    import zipfile
    from io import BytesIO
    from pathlib import Path
    from datetime import timezone, timedelta,  datetime

    mem = BytesIO()
    ts = datetime.now(KST).strftime("%Y%m%d_%H%M%S")

    roots = []
    try:
        roots.append(Path(STORE_DIR))
    except Exception:
        pass
    roots.append(Path("."))

    include_names = {
        "store_riders.json",
        "store_status.json",
        "store_delivery_status.json",
        "join_overrides.json",
        "login4_overrides.json",
        "prevplus_overrides.json",
        "plannedplus_overrides.json",
        "attendance_log.json",
        "sticker_overrides.json",
        "roulette.json",
        "roulette_spins.json",
        "roulette_config.json",
        "roulette_settings.json",
        "roulette.db",
        "app.db",
        "grade.db",
    }

    added = set()

    with zipfile.ZipFile(mem, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("backup_info.json", json.dumps({
            "ok": True,
            "created_at": ts,
            "type": "riderwelfare_full_zip_backup",
            "store_dir": str(STORE_DIR),
        }, ensure_ascii=False, indent=2))

        for root in roots:
            if not root.exists():
                continue

            for path in root.rglob("*"):
                if not path.is_file():
                    continue

                name = path.name
                rel = str(path.relative_to(root)).replace("\\", "/")
                lower_rel = rel.lower()

                should_add = False
                if name in include_names:
                    should_add = True
                if name.endswith(".json"):
                    should_add = True
                if "roulette" in lower_rel:
                    should_add = True

                if not should_add:
                    continue

                key = str(path.resolve())
                if key in added:
                    continue

                arcname = f"{root.name}/{rel}" if root.name else rel

                try:
                    z.write(path, arcname)
                    added.add(key)
                except Exception:
                    pass

        try:
            rj = roulette_db.export_json()
            z.writestr("roulette_export.json", json.dumps(rj, ensure_ascii=False, indent=2))
        except Exception as e:
            z.writestr("roulette_export_error.txt", str(e))

    mem.seek(0)
    return mem.getvalue()


# DISABLED_BY_FORCE_REAL_ZIP_EXPORT
# @app.get("/backup/export")
def backup_export_force_real_zip(request: Request):
    auth = _require_ingest(request)
    if auth:
        return auth

    from datetime import timezone, timedelta,  datetime

    data = _make_force_full_backup_zip()
    filename = "backup_" + datetime.now(KST).strftime("%Y%m%d_%H%M%S") + ".zip"

    return Response(
        data,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# DISABLED_BY_FORCE_REAL_ZIP_EXPORT
# @app.get("/admin/backup/export")
def admin_backup_export_force_real_zip(request: Request):
    r = require_admin(request)
    if r:
        return r

    from datetime import timezone, timedelta,  datetime

    data = _make_force_full_backup_zip()
    filename = "admin_backup_" + datetime.now(KST).strftime("%Y%m%d_%H%M%S") + ".zip"

    return Response(
        data,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )



# ============================================================
# FINAL BACKUP ROUTE OVERRIDE
# 기존 중복 /backup/export, /backup/import 라우트를 모두 제거하고
# 실제 ZIP 백업/복원 라우트만 다시 등록합니다.
# ============================================================

def _final_make_backup_zip_bytes():
    import json
    import zipfile
    from io import BytesIO
    from pathlib import Path
    from datetime import timezone, timedelta,  datetime

    mem = BytesIO()
    ts = datetime.now(KST).strftime("%Y%m%d_%H%M%S")

    roots = []
    try:
        roots.append(Path(STORE_DIR))
    except Exception:
        pass
    roots.append(Path("."))

    include_names = {
        "store_riders.json",
        "store_status.json",
        "store_delivery_status.json",
        "join_overrides.json",
        "login4_overrides.json",
        "prevplus_overrides.json",
        "plannedplus_overrides.json",
        "attendance_log.json",
        "sticker_overrides.json",
        "admin_settings.json",
        "roulette.json",
        "roulette_spins.json",
        "roulette_config.json",
        "roulette_settings.json",
        "roulette.db",
        "app.db",
        "grade.db",
    }

    added = set()

    with zipfile.ZipFile(mem, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("backup_info.json", json.dumps({
            "ok": True,
            "created_at": ts,
            "type": "riderwelfare_final_full_zip_backup",
            "store_dir": str(STORE_DIR),
        }, ensure_ascii=False, indent=2))

        for root in roots:
            if not root.exists():
                continue

            for path in root.rglob("*"):
                if not path.is_file():
                    continue

                name = path.name
                rel = str(path.relative_to(root)).replace("\\", "/")
                lower_rel = rel.lower()

                should_add = False
                if name in include_names:
                    should_add = True
                if name.endswith(".json"):
                    should_add = True
                if "roulette" in lower_rel:
                    should_add = True

                if not should_add:
                    continue

                key = str(path.resolve())
                if key in added:
                    continue

                arcname = f"{root.name}/{rel}" if root.name else rel

                try:
                    z.write(path, arcname)
                    added.add(key)
                except Exception:
                    pass

        try:
            rj = roulette_db.export_json()
            z.writestr("roulette_export.json", json.dumps(rj, ensure_ascii=False, indent=2))
        except Exception as e:
            z.writestr("roulette_export_error.txt", str(e))

        try:
            payload = get_backup_payload()
            z.writestr("full_backup_payload.json", json.dumps(payload, ensure_ascii=False, indent=2))
        except Exception as e:
            z.writestr("full_backup_payload_error.txt", str(e))

    mem.seek(0)
    return mem.getvalue()


def _final_restore_backup_bytes(data: bytes):
    import json
    import zipfile
    from io import BytesIO
    from pathlib import Path

    try:
        payload = json.loads(data.decode("utf-8-sig"))
        if isinstance(payload, dict):
            try:
                result = restore_backup_payload(payload)
                result["type"] = "json"
                return result
            except Exception:
                pass
    except Exception:
        pass

    zf = zipfile.ZipFile(BytesIO(data), "r")
    target_root = Path(STORE_DIR)
    target_root.mkdir(parents=True, exist_ok=True)

    restored = 0
    skipped = 0
    roulette_result = None
    roulette_payload = None
    roulette_spins = None

    allowed_ext = {".json", ".db", ".sqlite", ".sqlite3"}

    with zf as z:
        for info in z.infolist():
            if info.is_dir():
                continue

            raw_name = info.filename.replace("\\", "/")
            file_name = Path(raw_name).name
            suffix = Path(file_name).suffix.lower()
            raw = z.read(info.filename)

            if file_name == "backup_info.json":
                skipped += 1
                continue

            if file_name == "full_backup_payload.json":
                try:
                    payload = json.loads(raw.decode("utf-8-sig"))
                    if isinstance(payload, dict):
                        result = restore_backup_payload(payload)
                        result["type"] = "zip_payload"
                        return result
                except Exception:
                    pass

            if file_name in {"roulette_export.json", "roulette_spins.json", "roulette_history_all.json", "roulette_history.json"}:
                try:
                    loaded = json.loads(raw.decode("utf-8-sig"))
                    if isinstance(loaded, dict):
                        if "spins" in loaded or "settings" in loaded or "segments" in loaded:
                            roulette_payload = loaded
                        elif isinstance(loaded.get("roulette"), dict):
                            roulette_payload = loaded.get("roulette")
                    elif isinstance(loaded, list):
                        roulette_spins = loaded
                except Exception:
                    pass

            if suffix not in allowed_ext:
                skipped += 1
                continue

            try:
                out_path = target_root / file_name
                out_path.write_bytes(raw)
                restored += 1
            except Exception:
                skipped += 1

    try:
        if isinstance(roulette_payload, dict):
            roulette_result = roulette_db.import_json(roulette_payload)
        elif isinstance(roulette_spins, list):
            try:
                current = roulette_db.export_json()
            except Exception:
                current = {}
            if not isinstance(current, dict):
                current = {}
            current["spins"] = roulette_spins
            roulette_result = roulette_db.import_json(current)
    except Exception as e:
        roulette_result = {"ok": False, "error": str(e)}

    try:
        _riders_cache["ts"] = 0.0
        _riders_cache["data"] = None
    except Exception:
        pass
    try:
        _status_cache.clear()
    except Exception:
        pass
    try:
        _delivery_status_cache["ts"] = 0.0
        _delivery_status_cache["data"] = None
    except Exception:
        pass

    return {
        "ok": True,
        "type": "zip",
        "restored": restored,
        "skipped": skipped,
        "roulette_result": roulette_result,
    }


def _final_backup_export(request: Request):
    auth = _require_ingest(request)
    if auth:
        return auth

    from datetime import timezone, timedelta,  datetime

    data = _final_make_backup_zip_bytes()
    filename = "backup_" + datetime.now(KST).strftime("%Y%m%d_%H%M%S") + ".zip"
    return Response(
        data,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


def _final_admin_backup_export(request: Request):
    r = require_admin(request)
    if r:
        return r

    from datetime import timezone, timedelta,  datetime

    data = _final_make_backup_zip_bytes()
    filename = "admin_backup_" + datetime.now(KST).strftime("%Y%m%d_%H%M%S") + ".zip"
    return Response(
        data,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


async def _final_backup_import(request: Request, file: UploadFile = File(...)):
    auth = _require_ingest(request)
    if auth:
        return auth

    data = await file.read()
    try:
        return _final_restore_backup_bytes(data)
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=400)


async def _final_admin_backup_import(request: Request, file: UploadFile = File(...)):
    r = require_admin(request)
    if r:
        return r

    data = await file.read()
    try:
        result = _final_restore_backup_bytes(data)
        body = (
            "<div style='background:#fff;border:1px solid #e8e8e8;border-radius:16px;padding:16px;max-width:820px;margin:0 auto;'>"
            "<h3 style='margin-top:0;'>전체 백업 복원 완료</h3>"
            "<div style='color:#666;line-height:1.7;'>"
            "결과: <b>" + str(result) + "</b>"
            "</div>"
            "<div style='margin-top:12px;'>"
            "<a href='/admin/roulette' style='text-decoration:none;color:#111;'>룰렛 관리자 확인</a>"
            "&nbsp; | &nbsp;"
            "<a href='/dashboard' style='text-decoration:none;color:#111;'>대시보드 확인</a>"
            "&nbsp; | &nbsp;"
            "<a href='/health' style='text-decoration:none;color:#111;'>Health 확인</a>"
            "</div></div>"
        )
        return HTMLResponse(html_page("전체 백업 복원 완료", body))
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=400)


try:
    _backup_paths = {"/backup/export", "/admin/backup/export", "/backup/import", "/admin/backup/import"}
    app.router.routes = [
        route for route in app.router.routes
        if not (getattr(route, "path", None) in _backup_paths)
    ]

    app.add_api_route("/backup/export", _final_backup_export, methods=["GET"], name="Final Backup Export")
    app.add_api_route("/admin/backup/export", _final_admin_backup_export, methods=["GET"], name="Final Admin Backup Export")
    app.add_api_route("/backup/import", _final_backup_import, methods=["POST"], name="Final Backup Import")
    app.add_api_route("/admin/backup/import", _final_admin_backup_import, methods=["POST"], name="Final Admin Backup Import")
except Exception as _e:
    print("FINAL BACKUP ROUTE OVERRIDE ERROR:", _e)

# ====================================================================================
# Team management module (Rider Welfare)
# - Admin creates teams by selecting one leader and multiple riders.
# - Leader sees member-level complete/reject/cancel details.
# - Members see only team aggregate quota/complete/reject/cancel.
# - Weekly cycle: Wednesday 06:00 ~ next Wednesday 05:59.
# - Period statistics are accumulated from delivery-status snapshot deltas.
# ====================================================================================
TEAM_FILE = str(STORE_DIR / "teams.json")
TEAM_WEEKLY_SET_FILE = str(STORE_DIR / "team_weekly_sets.json")
TEAM_PERIOD_STATS_FILE = str(STORE_DIR / "team_period_stats.json")
_team_lock = threading.RLock()

TEAM_SET_QUOTA = {
    0: (21, 20, 30, 29),  # Monday
    1: (21, 20, 30, 29),  # Tuesday
    2: (21, 20, 30, 29),  # Wednesday
    3: (21, 20, 30, 29),  # Thursday
    4: (24, 21, 32, 33),  # Friday
    5: (31, 22, 36, 31),  # Saturday
    6: (33, 22, 35, 30),  # Sunday
}
TEAM_PERIODS = ("morning_lunch", "afternoon_offpeak", "dinner_peak", "late_night")
TEAM_PERIOD_LABELS = {
    "morning_lunch": "아침·점심피크",
    "afternoon_offpeak": "오후 논피크",
    "dinner_peak": "저녁피크",
    "late_night": "심야",
}


def _team_load(path: str, default: Any) -> Any:
    with _team_lock:
        try:
            p = Path(path)
            if not p.exists():
                return default
            data = json.loads(p.read_text(encoding="utf-8"))
            return data
        except Exception:
            return default


def _team_save(path: str, data: Any) -> None:
    with _team_lock:
        p = Path(path)
        p.parent.mkdir(parents=True, exist_ok=True)
        tmp = p.with_suffix(p.suffix + ".tmp")
        tmp.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        tmp.replace(p)


def load_teams() -> List[Dict[str, Any]]:
    data = _team_load(TEAM_FILE, [])
    return data if isinstance(data, list) else []


def save_teams(data: List[Dict[str, Any]]) -> None:
    _team_save(TEAM_FILE, data)


def load_team_weekly_sets() -> Dict[str, Any]:
    data = _team_load(TEAM_WEEKLY_SET_FILE, {})
    return data if isinstance(data, dict) else {}


def save_team_weekly_sets(data: Dict[str, Any]) -> None:
    _team_save(TEAM_WEEKLY_SET_FILE, data)


def load_team_period_stats() -> Dict[str, Any]:
    data = _team_load(TEAM_PERIOD_STATS_FILE, {"snapshots": {}, "stats": {}})
    if not isinstance(data, dict):
        data = {}
    data.setdefault("snapshots", {})
    data.setdefault("stats", {})
    return data


def save_team_period_stats(data: Dict[str, Any]) -> None:
    _team_save(TEAM_PERIOD_STATS_FILE, data)


def current_team_week_start(op_day: date | None = None) -> date:
    d = op_day or operation_date()
    # weekday: Monday=0, Wednesday=2
    return d - timedelta(days=(d.weekday() - 2) % 7)


def next_team_week_start(op_day: date | None = None) -> date:
    return current_team_week_start(op_day) + timedelta(days=7)


def team_period_for_datetime(now: datetime | None = None) -> str:
    now = now or now_kst()
    op_day = operation_date(now)
    hour = now.hour
    if hour >= 20 or hour < 6:
        return "late_night"
    if op_day.weekday() in (5, 6):
        if hour < 14:
            return "morning_lunch"
        if hour < 17:
            return "afternoon_offpeak"
        return "dinner_peak"
    if hour < 13:
        return "morning_lunch"
    if hour < 17:
        return "afternoon_offpeak"
    return "dinner_peak"


def previous_team_period(period: str) -> str:
    order = list(TEAM_PERIODS)
    try:
        idx = order.index(period)
    except ValueError:
        return period
    return order[idx - 1] if idx > 0 else "late_night"


def team_period_time_text(op_day: date, period: str) -> str:
    weekend = op_day.weekday() in (5, 6)
    if period == "morning_lunch":
        return "06:00~14:00" if weekend else "06:00~13:00"
    if period == "afternoon_offpeak":
        return "14:00~17:00" if weekend else "13:00~17:00"
    if period == "dinner_peak":
        return "17:00~20:00"
    return "20:00~다음 날 06:00"


def team_quota(op_day: date, set_count: int) -> Dict[str, int]:
    base = TEAM_SET_QUOTA.get(op_day.weekday(), (0, 0, 0, 0))
    sets = max(0, int(set_count or 0))
    return {p: int(base[i]) * sets for i, p in enumerate(TEAM_PERIODS)}


def rider_directory() -> Dict[str, Dict[str, Any]]:
    out: Dict[str, Dict[str, Any]] = {}
    for rr in fetch_riders_cached():
        name = rr.get("name") or ""
        login4, real4, _ = get_login4_for_rider(rr)
        if not name or not login4:
            continue
        key = f"{norm_name(name)}|{login4}"
        out[key] = {
            "key": key,
            "name": name,
            "login4": login4,
            "real4": real4,
            "phone": rr.get("phoneNumber") or "",
            "rider": rr,
        }
    return out


def get_team_member_current_grade(info: Dict[str, Any]) -> str:
    """팀장 화면에 표시할 현재등급(이전 평가기간 기준)을 계산합니다."""
    try:
        rider = info.get("rider") or {}
        name = str(info.get("name") or rider.get("name") or "")
        login4 = str(info.get("login4") or "")
        real4 = str(info.get("real4") or "")
        if not name or not login4 or not real4:
            return "-"

        today = operation_date()
        eff_join, _ = get_effective_join_date_by_login_key(rider, login4)
        cur_start, _ = current_period(eff_join, today)
        prev_end = cur_start - timedelta(days=1)
        prev_month = date(cur_start.year, cur_start.month, 1) + relativedelta(months=-1)
        prev_start = clamp_day(prev_month.year, prev_month.month, eff_join.day)
        prev_from, prev_to = period_to_from_to(prev_start, prev_end)

        api_key = f"{norm_name(name)}|{real4}"
        prev_completed = 0
        if prev_from <= prev_to:
            prev_completed = int(fetch_status_complete_map_cached(prev_from, prev_to).get(api_key, 0) or 0)

        login_key = f"{norm_name(name)}|{login4}"
        return grade_from_total(prev_completed + get_prevplus(login_key))
    except Exception:
        return "-"


def find_team_by_member(login_key: str) -> Optional[Dict[str, Any]]:
    for team in load_teams():
        if not team.get("active", True):
            continue
        if login_key == team.get("leader_key") or login_key in (team.get("member_keys") or []):
            return team
    return None


def is_team_leader(team: Dict[str, Any], login_key: str) -> bool:
    return bool(team and team.get("leader_key") == login_key)


def team_week_record(team_id: str, week_start: date | None = None) -> Dict[str, Any]:
    ws = week_start or current_team_week_start()
    key = f"{team_id}|{ws.isoformat()}"
    rec = load_team_weekly_sets().get(key) or {}
    if not isinstance(rec, dict):
        rec = {}
    return rec


def capture_team_snapshot(delivery_data: Dict[str, Any] | None = None) -> None:
    """배민 v4 응답의 기사별 구간 완료건수를 그대로 저장합니다.

    deliveryPeakTimeCount의 morning/afternoon/evening/midnight 값을 사용하므로
    Render 재배포 시점이나 최초 수집 시각과 관계없이 완료건수가 정확한 구간에 표시됩니다.
    거절·취소는 API에 구간별 값이 없으므로 이전 누적값 대비 증분만 현재 구간에 기록합니다.
    """
    try:
        ds = delivery_data if isinstance(delivery_data, dict) else fetch_delivery_status_cached()
        today_map = build_today_stats_map(ds)
        directory = rider_directory()
        now = now_kst()
        op_day = operation_date(now)
        current_period = team_period_for_datetime(now)
        payload = load_team_period_stats()
        snapshots = payload.setdefault("snapshots", {})
        stats = payload.setdefault("stats", {})

        for login_key, info in directory.items():
            api_key = f"{norm_name(info['name'])}|{info['real4']}"
            cur = today_map.get(api_key) or {}
            current = {
                "complete": int(cur.get("complete") or 0),
                "reject": int(cur.get("reject") or 0),
                "cancel": int(cur.get("cancel") or 0),
            }
            peak_complete = cur.get("peak_complete") or {}

            # 완료건수는 배민이 제공하는 구간별 누적값으로 매번 덮어씁니다.
            # 기존에 잘못 배정된 완료건수도 다음 수집 즉시 정상 구간으로 교정됩니다.
            for period in TEAM_PERIODS:
                stat_key = f"{op_day.isoformat()}|{period}|{login_key}"
                row = stats.get(stat_key) or {"complete": 0, "reject": 0, "cancel": 0}
                row["complete"] = max(0, int(peak_complete.get(period) or 0))
                row["updated_at"] = now.isoformat()
                row["complete_source"] = "deliveryPeakTimeCount"
                stats[stat_key] = row

            # 거절·취소는 구간별 API 값이 없으므로 증분만 현재 구간에 더합니다.
            snap_key = f"{op_day.isoformat()}|{login_key}"
            prev = snapshots.get(snap_key) if isinstance(snapshots.get(snap_key), dict) else {}
            stat_key = f"{op_day.isoformat()}|{current_period}|{login_key}"
            row = stats.get(stat_key) or {"complete": 0, "reject": 0, "cancel": 0}
            for field in ("reject", "cancel"):
                old = int(prev.get(field) or 0)
                new = int(current[field])
                delta = new - old if new >= old else new
                row[field] = int(row.get(field) or 0) + max(0, delta)
            row["updated_at"] = now.isoformat()
            stats[stat_key] = row

            snapshots[snap_key] = {
                **current,
                "updated_at": now.isoformat(),
                "period": current_period,
                "peak_api": True,
            }

        cutoff = op_day - timedelta(days=120)
        payload["stats"] = {
            k: v for k, v in stats.items()
            if safe_date_parse(k.split("|", 1)[0]) is None or safe_date_parse(k.split("|", 1)[0]) >= cutoff
        }
        payload["snapshots"] = {
            k: v for k, v in snapshots.items()
            if safe_date_parse(k.split("|", 1)[0]) is None or safe_date_parse(k.split("|", 1)[0]) >= cutoff
        }
        save_team_period_stats(payload)
    except Exception as e:
        print("TEAM SNAPSHOT ERROR:", e)


def team_member_day_stats(login_key: str, op_day: date) -> Dict[str, Dict[str, int]]:
    out = {p: {"complete": 0, "reject": 0, "cancel": 0} for p in TEAM_PERIODS}

    # 과거 날짜는 collector가 rider-delivery-status로 가져온 개인 기록을 우선 사용합니다.
    if op_day < operation_date():
        directory = rider_directory()
        info = directory.get(login_key) or {}
        name = info.get("name") or ""
        real4 = info.get("real4") or ""
        api_key = f"{norm_name(name)}|{real4}" if name and real4 else ""
        history = fetch_status_history_map(op_day, op_day)
        rec = history.get(api_key) if api_key else None
        if isinstance(rec, dict):
            peak = rec.get("peak") or {}
            out["morning_lunch"]["complete"] = int(peak.get("morning") or 0)
            out["afternoon_offpeak"]["complete"] = int(peak.get("afternoon") or 0)
            out["dinner_peak"]["complete"] = int(peak.get("evening") or 0)
            out["late_night"]["complete"] = int(peak.get("midnight") or 0)
            out["_meta"] = {  # type: ignore[assignment]
                "historical_api": 1,
                "daily_reject": int(rec.get("reject") or 0),
                "daily_cancel": int(rec.get("cancel") or 0),
                "daily_complete": int(rec.get("complete") or 0),
                "hourlyCompleted": rec.get("hourlyCompleted") or [],
            }
            return out

    # 오늘 또는 구형 저장자료는 기존 팀 스냅샷을 사용합니다.
    payload = load_team_period_stats()
    all_stats = payload.get("stats") or {}
    has_period_data = False
    for period in TEAM_PERIODS:
        row = all_stats.get(f"{op_day.isoformat()}|{period}|{login_key}") or {}
        if row:
            has_period_data = True
        out[period] = {
            "complete": int(row.get("complete") or 0),
            "reject": int(row.get("reject") or 0),
            "cancel": int(row.get("cancel") or 0),
        }

    # 구형 completeMap만 남아 있는 과거 자료에 대한 호환 처리
    if not has_period_data and op_day < operation_date():
        directory = rider_directory()
        info = directory.get(login_key) or {}
        name = info.get("name") or ""
        real4 = info.get("real4") or ""
        api_key = f"{norm_name(name)}|{real4}" if name and real4 else ""
        if api_key:
            daily_map = fetch_status_complete_map_cached(op_day, op_day)
            if api_key in daily_map:
                out["morning_lunch"]["complete"] = int(daily_map.get(api_key) or 0)
                out["_meta"] = {"historical_daily": 1}  # type: ignore[assignment]
    return out


def team_all_keys(team: Dict[str, Any]) -> List[str]:
    keys = [str(team.get("leader_key") or "")]
    keys.extend([str(x) for x in (team.get("member_keys") or [])])
    return list(dict.fromkeys([x for x in keys if x]))


def aggregate_team_day(team: Dict[str, Any], op_day: date) -> Tuple[Dict[str, Dict[str, int]], Dict[str, Dict[str, Dict[str, int]]], bool]:
    period_total = {p: {"complete": 0, "reject": 0, "cancel": 0} for p in TEAM_PERIODS}
    member_rows: Dict[str, Dict[str, Dict[str, int]]] = {}
    historical_daily = False
    for key in team_all_keys(team):
        member_stats = team_member_day_stats(key, op_day)
        meta = member_stats.get("_meta") if isinstance(member_stats.get("_meta"), dict) else {}
        if meta.get("historical_daily"):
            historical_daily = True
        member_rows[key] = member_stats
        for p in TEAM_PERIODS:
            for f in ("complete", "reject", "cancel"):
                period_total[p][f] += int(member_stats[p][f])
        # 과거 API는 거절/취소를 구간별로 주지 않으므로 일일 합계를 첫 구간에만 담아
        # 팀/개인 하루 합계 계산에는 포함시키고 구간 표시는 별도로 처리합니다.
        if meta.get("historical_api"):
            period_total["morning_lunch"]["reject"] += int(meta.get("daily_reject") or 0)
            period_total["morning_lunch"]["cancel"] += int(meta.get("daily_cancel") or 0)
    return period_total, member_rows, historical_daily


def _team_escape(s: Any) -> str:
    import html
    return html.escape(str(s or ""), quote=True)


def _team_nav(admin: bool = False) -> str:
    if admin:
        return "<div style='display:flex;gap:10px;flex-wrap:wrap;margin-bottom:14px;'><a href='/dashboard'>전체현황</a><a href='/admin/teams'><b>팀 구성 관리</b></a><a href='/admin/team-sets'>팀 세트 관리</a><a href='/admin-logout'>로그아웃</a></div>"
    return "<div style='display:flex;gap:10px;flex-wrap:wrap;margin-bottom:14px;'><a href='/'>등급조회</a><a href='/team'><b>팀 페이지</b></a><a href='/team-logout'>팀 로그아웃</a></div>"


@app.get("/admin/teams", response_class=HTMLResponse)
def admin_teams_page(request: Request, edit_id: str = ""):
    r = require_admin(request)
    if r:
        return r
    directory = rider_directory()
    teams = load_teams()
    assigned = {}
    for t in teams:
        if not t.get("active", True):
            continue
        for k in team_all_keys(t):
            assigned[k] = t.get("name") or "-"

    edit_team = next((t for t in teams if str(t.get("id")) == str(edit_id)), None)
    leader_selected = str((edit_team or {}).get("leader_key") or "")
    default_team_start = operation_date().isoformat()
    if edit_team:
        saved_start = safe_date_parse(str(edit_team.get("team_start_date") or ""))
        if saved_start:
            default_team_start = saved_start.isoformat()
        else:
            created_text = str(edit_team.get("created_at") or "")[:10]
            default_team_start = (safe_date_parse(created_text) or operation_date()).isoformat()
    member_selected = set((edit_team or {}).get("member_keys") or [])
    options = []
    checks = []
    for key, info in sorted(directory.items(), key=lambda x: x[1]["name"]):
        team_name = assigned.get(key, "소속 없음")
        disabled = bool(key in assigned and (not edit_team or key not in team_all_keys(edit_team)))
        dis = " disabled" if disabled else ""
        sel = " selected" if key == leader_selected else ""
        chk = " checked" if key in member_selected else ""
        muted = "color:#aaa;" if disabled else ""
        options.append(f"<option value='{_team_escape(key)}'{sel}{dis}>{_team_escape(info['name'])} · {info['login4']} · {team_name}</option>")
        checks.append(f"<label style='display:flex;align-items:center;gap:8px;padding:9px;border-bottom:1px solid #eee;{muted}'><input type='checkbox' name='member_keys' value='{_team_escape(key)}'{chk}{dis}> <b>{_team_escape(info['name'])}</b><span style='color:#777'>{info['login4']} · {team_name}</span></label>")

    team_cards = ""
    for t in teams:
        leader = directory.get(str(t.get("leader_key") or ""), {})
        members = [directory.get(k, {}).get("name", k) for k in (t.get("member_keys") or [])]
        team_cards += f"""
        <div style='border:1px solid #e5e5e5;border-radius:14px;padding:14px;background:#fff;margin-bottom:10px;'>
          <div style='display:flex;justify-content:space-between;gap:10px;align-items:center;'><div><b style='font-size:18px'>{_team_escape(t.get('name'))}</b><div style='color:#666;margin-top:5px'>팀장 {_team_escape(leader.get('name','-'))} · 총 {1+len(members)}명</div></div><a href='/admin/teams?edit_id={_team_escape(t.get('id'))}'>수정</a></div>
          <div style='margin-top:8px;color:#555'>팀 생성일: <b>{_team_escape(t.get('team_start_date') or str(t.get('created_at') or '')[:10] or '-')}</b></div>
          <div style='margin-top:8px;color:#555'>팀원: {_team_escape(', '.join(members) if members else '없음')}</div>
          <form method='post' action='/admin/team-delete' style='margin-top:10px' onsubmit="return confirm('팀을 해체하시겠습니까?')"><input type='hidden' name='team_id' value='{_team_escape(t.get('id'))}'><button style='border:1px solid #ddd;background:#fff;border-radius:8px;padding:7px 10px'>팀 해체</button></form>
        </div>"""

    body = f"""
    {_team_nav(True)}
    <div style='display:grid;grid-template-columns:minmax(320px,1fr) minmax(360px,1.3fr);gap:16px;align-items:start'>
      <div><h2 style='margin-top:0'>팀 구성 관리</h2>{team_cards or '<div>생성된 팀이 없습니다.</div>'}</div>
      <div style='background:#fff;border:1px solid #e5e5e5;border-radius:16px;padding:16px'>
        <h2 style='margin-top:0'>{'팀 수정' if edit_team else '새 팀 만들기'}</h2>
        <form method='post' action='/admin/team-save'>
          <input type='hidden' name='team_id' value='{_team_escape((edit_team or {}).get('id',''))}'>
          <label>팀명</label><input name='team_name' value='{_team_escape((edit_team or {}).get('name',''))}' required style='width:100%;box-sizing:border-box;padding:11px;border:1px solid #ddd;border-radius:10px;margin:6px 0 14px'>
          <label>팀 생성일</label><input type='date' name='team_start_date' value='{_team_escape(default_team_start)}' max='{operation_date().isoformat()}' required style='width:100%;box-sizing:border-box;padding:11px;border:1px solid #ddd;border-radius:10px;margin:6px 0 14px'>
          <div style='color:#777;font-size:13px;margin:-8px 0 14px'>선택한 날짜부터 팀 과거 운행기록을 조회합니다.</div>
          <label>팀장 선택</label><select name='leader_key' required style='width:100%;padding:11px;border:1px solid #ddd;border-radius:10px;margin:6px 0 14px'><option value=''>팀장 선택</option>{''.join(options)}</select>
          <div style='font-weight:700;margin-bottom:6px'>팀원 선택</div><div style='max-height:420px;overflow:auto;border:1px solid #ddd;border-radius:10px'>{''.join(checks)}</div>
          <button style='width:100%;margin-top:14px;padding:12px;border:none;border-radius:10px;background:#111;color:#fff;font-weight:800'>{'팀 수정 저장' if edit_team else '팀 생성'}</button>
        </form>
      </div>
    </div>"""
    return HTMLResponse(html_page("팀 구성 관리", body))


@app.post("/admin/team-save")
async def admin_team_save(request: Request):
    r = require_admin(request)
    if r:
        return r
    form = await request.form()
    team_id = str(form.get("team_id") or "").strip()
    team_name = str(form.get("team_name") or "").strip()
    team_start_date = safe_date_parse(str(form.get("team_start_date") or "").strip())
    leader_key = str(form.get("leader_key") or "").strip()
    member_keys = [str(x) for x in form.getlist("member_keys") if str(x)]
    member_keys = [x for x in dict.fromkeys(member_keys) if x != leader_key]
    directory = rider_directory()
    if not team_name or leader_key not in directory or team_start_date is None:
        return RedirectResponse("/admin/teams", status_code=303)
    if team_start_date > operation_date():
        team_start_date = operation_date()

    teams = load_teams()
    # Do not allow a rider to belong to multiple active teams.
    selected = set([leader_key] + member_keys)
    for t in teams:
        if team_id and str(t.get("id")) == team_id:
            continue
        if t.get("active", True) and selected.intersection(team_all_keys(t)):
            return HTMLResponse(html_page("팀 저장 실패", "<div style='max-width:600px;margin:40px auto;background:#fff;padding:20px;border-radius:14px'><h3>이미 다른 팀에 소속된 기사가 포함되어 있습니다.</h3><a href='/admin/teams'>돌아가기</a></div>"), status_code=400)

    now_text = now_kst().isoformat()
    if team_id:
        target = next((t for t in teams if str(t.get("id")) == team_id), None)
        if target:
            target.update({"name": team_name, "team_start_date": team_start_date.isoformat(), "leader_key": leader_key, "member_keys": member_keys, "active": True, "updated_at": now_text})
    else:
        next_id = str(max([int(t.get("id") or 0) for t in teams] + [0]) + 1)
        teams.append({"id": next_id, "name": team_name, "team_start_date": team_start_date.isoformat(), "leader_key": leader_key, "member_keys": member_keys, "active": True, "created_at": now_text, "updated_at": now_text})
    save_teams(teams)
    return RedirectResponse("/admin/teams", status_code=303)


@app.post("/admin/team-delete")
def admin_team_delete(request: Request, team_id: str = Form(...)):
    r = require_admin(request)
    if r:
        return r
    teams = load_teams()
    teams = [t for t in teams if str(t.get("id")) != str(team_id)]
    save_teams(teams)
    return RedirectResponse("/admin/teams", status_code=303)


@app.get("/admin/team-sets", response_class=HTMLResponse)
def admin_team_sets_page(request: Request):
    r = require_admin(request)
    if r:
        return r
    directory = rider_directory()
    sets = load_team_weekly_sets()
    ws = current_team_week_start()
    nws = next_team_week_start()
    rows = ""
    for t in load_teams():
        leader_name = directory.get(str(t.get("leader_key") or ""), {}).get("name", "-")
        for week in (ws, nws):
            key = f"{t.get('id')}|{week.isoformat()}"
            rec = sets.get(key) or {}
            rows += f"<tr><td>{_team_escape(t.get('name'))}</td><td>{_team_escape(leader_name)}</td><td>{week}~{week+timedelta(days=6)}</td><td>{int(rec.get('set_count') or 0)}</td><td>{_team_escape(rec.get('status') or '미신청')}</td><td><form method='post' action='/admin/team-set-save' style='display:flex;gap:6px'><input type='hidden' name='team_id' value='{_team_escape(t.get('id'))}'><input type='hidden' name='week_start' value='{week}'><input name='set_count' type='number' min='0' max='20' value='{int(rec.get('set_count') or 0)}' style='width:64px'><select name='status'><option value='approved'>승인</option><option value='pending'>대기</option><option value='rejected'>반려</option></select><button>저장</button></form></td></tr>"
    body = f"{_team_nav(True)}<h2>팀 세트 관리</h2><div style='overflow:auto;background:#fff;border:1px solid #ddd;border-radius:14px'><table style='width:100%;border-collapse:collapse'><thead><tr><th>팀</th><th>팀장</th><th>적용 주차</th><th>세트</th><th>상태</th><th>관리자 변경</th></tr></thead><tbody>{rows}</tbody></table></div><style>th,td{{padding:10px;border-bottom:1px solid #eee;text-align:center;white-space:nowrap}}</style>"
    return HTMLResponse(html_page("팀 세트 관리", body))


@app.post("/admin/team-set-save")
def admin_team_set_save(request: Request, team_id: str = Form(...), week_start: str = Form(...), set_count: int = Form(...), status: str = Form("approved")):
    r = require_admin(request)
    if r:
        return r
    ws = safe_date_parse(week_start)
    if not ws:
        return RedirectResponse("/admin/team-sets", status_code=303)
    data = load_team_weekly_sets()
    key = f"{team_id}|{ws.isoformat()}"
    old = data.get(key) or {}
    old.update({"team_id": team_id, "week_start": ws.isoformat(), "week_end": (ws+timedelta(days=6)).isoformat(), "set_count": max(0, int(set_count)), "status": status, "approved_at": now_kst().isoformat(), "approved_by": "admin"})
    data[key] = old
    save_team_weekly_sets(data)
    return RedirectResponse("/admin/team-sets", status_code=303)


@app.get("/team-login", response_class=HTMLResponse)
def team_login_page():
    body = """<div style='max-width:430px;margin:70px auto;background:#fff;border:1px solid #ddd;border-radius:16px;padding:20px'><h2>팀 페이지 로그인</h2><form method='post' action='/team-login'><input name='name' placeholder='이름' required style='width:100%;box-sizing:border-box;padding:12px;margin-bottom:10px'><input name='login4' pattern='\\d{4}' maxlength='4' placeholder='로그인용 뒷 4자리' required style='width:100%;box-sizing:border-box;padding:12px;margin-bottom:10px'><button style='width:100%;padding:12px;background:#111;color:#fff;border:none;border-radius:10px'>로그인</button></form><div style='margin-top:12px'><a href='/'>등급조회로</a></div></div>"""
    return HTMLResponse(html_page("팀 로그인", body))


@app.post("/team-login")
def team_login_action(request: Request, name: str = Form(...), login4: str = Form(...)):
    rider = find_rider_by_name_login4(name, login4)
    if not rider:
        return HTMLResponse(html_page("로그인 실패", "<div style='max-width:500px;margin:50px auto;background:#fff;padding:20px'><h3>기사 정보를 찾을 수 없습니다.</h3><a href='/team-login'>다시 로그인</a></div>"), status_code=400)
    key = f"{norm_name(rider.get('name') or '')}|{login4}"
    request.session["rider_login_key"] = key
    return RedirectResponse("/team", status_code=303)


@app.get("/team-logout")
def team_logout(request: Request):
    request.session.pop("rider_login_key", None)
    return RedirectResponse("/", status_code=303)


@app.post("/team/set-request")
def team_set_request(request: Request, week_start: str = Form(...), set_count: int = Form(...)):
    key = str(request.session.get("rider_login_key") or "")
    team = find_team_by_member(key)
    if not team or not is_team_leader(team, key):
        return RedirectResponse("/team-login", status_code=303)
    ws = safe_date_parse(week_start)
    if ws != next_team_week_start():
        return RedirectResponse("/team", status_code=303)
    data = load_team_weekly_sets()
    record_key = f"{team.get('id')}|{ws.isoformat()}"
    data[record_key] = {"team_id": str(team.get("id")), "week_start": ws.isoformat(), "week_end": (ws+timedelta(days=6)).isoformat(), "set_count": max(0, min(20, int(set_count))), "member_count": len(team_all_keys(team)), "status": "pending", "requested_by": key, "requested_at": now_kst().isoformat()}
    save_team_weekly_sets(data)
    return RedirectResponse("/team", status_code=303)


@app.get("/team", response_class=HTMLResponse)
def team_page(request: Request, day: str = ""):
    login_key = str(request.session.get("rider_login_key") or "")
    if not login_key:
        return RedirectResponse("/team-login", status_code=303)

    team = find_team_by_member(login_key)
    if not team:
        return HTMLResponse(html_page("팀 페이지", f"{_team_nav(False)}<div class='empty-card'><h3>소속된 팀이 없습니다.</h3></div>"))

    capture_team_snapshot()
    directory = rider_directory()
    requested_day = safe_date_parse(day) or operation_date()
    team_start = safe_date_parse(str(team.get("team_start_date") or ""))
    if team_start is None:
        team_start = safe_date_parse(str(team.get("created_at") or "")[:10]) or operation_date()

    leader_mode = is_team_leader(team, login_key)
    today_op = operation_date()
    # 팀장은 팀 생성일부터 전체 기록을 조회할 수 있습니다.
    # 일반 팀원은 오늘과 최근 7일 전까지(과거 일주일)만 조회할 수 있습니다.
    member_min_day = max(team_start, today_op - timedelta(days=7))
    allowed_min_day = team_start if leader_mode else member_min_day

    op_day = min(requested_day, today_op)
    if op_day < allowed_min_day:
        op_day = allowed_min_day
    ws = current_team_week_start(op_day)
    rec = team_week_record(str(team.get("id")), ws)
    set_count = int(rec.get("set_count") or 0) if rec.get("status") in ("approved", "pending") else 0
    quota = team_quota(op_day, set_count)
    totals, member_rows, historical_daily = aggregate_team_day(team, op_day)
    leader_name = directory.get(str(team.get("leader_key") or ""), {}).get("name", "-")

    period_rows = ""
    period_cards = ""
    day_complete = day_reject = day_cancel = day_quota = 0
    display_periods = ("morning_lunch",) if historical_daily else TEAM_PERIODS
    for p in display_periods:
        q = sum(quota.values()) if historical_daily else quota[p]
        c = sum(totals[x]["complete"] for x in TEAM_PERIODS) if historical_daily else totals[p]["complete"]
        rj = sum(totals[x]["reject"] for x in TEAM_PERIODS) if historical_daily else totals[p]["reject"]
        ca = sum(totals[x]["cancel"] for x in TEAM_PERIODS) if historical_daily else totals[p]["cancel"]
        label = "일일 전체 개인기록" if historical_daily else TEAM_PERIOD_LABELS[p]
        time_text = "배민 과거 개인기록 합산" if historical_daily else team_period_time_text(op_day,p)
        day_complete += c
        day_reject += rj
        day_cancel += ca
        day_quota += q
        rate = round(c / q * 100, 1) if q else 0
        period_rows += f"<tr><td>{label}</td><td>{time_text}</td><td><b>{c}</b> / {q}</td><td>{rate}%</td></tr>"
        period_cards += f"""
        <div class='period-card'>
          <div class='period-card-head'><b>{label}</b><span>{time_text}</span></div>
          <div class='period-main'><strong>{c}</strong><span>/ {q}건</span><em>{rate}%</em></div>
        </div>"""

    member_table = ""
    if leader_mode:
        mrows = ""
        member_cards = ""
        for k in team_all_keys(team):
            info = directory.get(k, {"name": k})
            parts = member_rows.get(k) or {}
            complete = sum(int((parts.get(p) or {}).get("complete") or 0) for p in TEAM_PERIODS)
            meta = parts.get("_meta") if isinstance(parts.get("_meta"), dict) else {}
            reject = int(meta.get("daily_reject") or 0) if meta.get("historical_api") else sum(int((parts.get(p) or {}).get("reject") or 0) for p in TEAM_PERIODS)
            cancel = int(meta.get("daily_cancel") or 0) if meta.get("historical_api") else sum(int((parts.get(p) or {}).get("cancel") or 0) for p in TEAM_PERIODS)
            ratio = round((reject + cancel) / complete * 100, 1) if complete else 0
            role = "팀장" if k == team.get("leader_key") else "팀원"
            current_grade = get_team_member_current_grade(info)
            phone_text = mask_phone(str(info.get("phone") or ""))
            if historical_daily:
                details = f"일일 전체 완료: {complete}건"
                detail_cards = f"<div><span>일일 전체</span><b>{complete}</b><small>배민 과거 개인기록</small></div>"
            elif meta.get("historical_api"):
                details = "<br>".join(
                    f"{TEAM_PERIOD_LABELS[p]}: 완료 {(parts.get(p) or {}).get('complete',0)}"
                    for p in TEAM_PERIODS
                )
                detail_cards = "".join(
                    f"<div><span>{TEAM_PERIOD_LABELS[p]}</span><b>{(parts.get(p) or {}).get('complete',0)}</b><small>배민 과거 구간기록</small></div>"
                    for p in TEAM_PERIODS
                )
            else:
                # 배민 API는 구간별 거절/취소를 제공하지 않으므로 구간 상세에는 완료만 표시합니다.
                details = "<br>".join(
                    f"{TEAM_PERIOD_LABELS[p]}: 완료 {(parts.get(p) or {}).get('complete',0)}"
                    for p in TEAM_PERIODS
                )
                detail_cards = "".join(
                    f"<div><span>{TEAM_PERIOD_LABELS[p]}</span><b>{(parts.get(p) or {}).get('complete',0)}</b><small>완료건수</small></div>"
                    for p in TEAM_PERIODS
                )
            mrows += f"<tr><td>{_team_escape(info.get('name'))}<div class='role'>{role} · 현재등급 {current_grade}<br>{_team_escape(phone_text)}</div></td><td>{complete}</td><td>{reject}</td><td>{cancel}</td><td>{ratio}%</td><td class='detail-cell'>{details}</td></tr>"
            member_cards += f"""
            <div class='member-card'>
              <div class='member-head'><div><b>{_team_escape(info.get('name'))}</b><span>{role} · 현재등급 {current_grade}</span><span>{_team_escape(phone_text)}</span></div><strong>{complete}건</strong></div>
              <div class='member-summary'><span>거절 <b>{reject}</b></span><span>취소 <b>{cancel}</b></span><span>비율 <b>{ratio}%</b></span></div>
              <details><summary>구간별 상세 보기</summary><div class='member-periods'>{detail_cards}</div></details>
            </div>"""

        member_table = f"""
        <section class='section-block'>
          <h3>팀원별 세부 운행기록</h3>
          <div class='desktop-table'><table><thead><tr><th>이름</th><th>완료</th><th>거절</th><th>취소</th><th>거절+취소율</th><th>구간별 상세</th></tr></thead><tbody>{mrows}</tbody></table></div>
          <div class='mobile-cards'>{member_cards}</div>
        </section>"""

    set_form = ""
    if leader_mode:
        nws = next_team_week_start()
        next_rec = team_week_record(str(team.get("id")), nws)
        set_form = f"""
        <section class='set-card'>
          <h3>다음 주 세트 신청</h3>
          <div class='set-meta'>{nws} ~ {nws + timedelta(days=6)}<br>현재 {len(team_all_keys(team))}명 · 신청상태 <b>{_team_escape(next_rec.get('status') or '미신청')}</b></div>
          <form method='post' action='/team/set-request'>
            <input type='hidden' name='week_start' value='{nws}'>
            <input type='number' name='set_count' min='0' max='20' value='{int(next_rec.get('set_count') or 1)}'>
            <button>신청/수정</button>
          </form>
        </section>"""

    prev_day = max(allowed_min_day, op_day - timedelta(days=1))
    next_day = min(today_op, op_day + timedelta(days=1))
    prev_disabled = op_day <= allowed_min_day
    next_disabled = op_day >= today_op
    progress = round(day_complete / day_quota * 100, 1) if day_quota else 0

    body = f"""
    <style>
      *{{box-sizing:border-box}}
      body{{margin:0}}
      .team-shell{{max-width:980px;margin:0 auto}}
      .team-nav{{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:12px}}
      .team-nav a{{text-decoration:none;padding:8px 11px;border-radius:9px;background:#f1f1f1;color:#222;font-size:14px}}
      .hero{{background:#fff;border:1px solid #e3e3e3;border-radius:18px;padding:18px;box-shadow:0 3px 14px rgba(0,0,0,.03)}}
      .hero h2{{margin:0;font-size:26px}}
      .sub{{color:#666;margin-top:7px;font-size:14px}}
      .date-nav{{display:grid;grid-template-columns:1fr auto 1fr;align-items:center;margin-top:18px;gap:10px}}
      .date-nav a{{text-decoration:none;color:#5b2a86;font-weight:800}}
      .date-nav a:last-child{{text-align:right}}
      .date-nav b{{font-size:18px}}
      .summary-grid{{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:10px;margin-top:16px}}
      .team-card{{padding:14px;border:1px solid #ececec;border-radius:14px;color:#777;min-width:0;background:#fcfcfc}}
      .team-card b{{display:block;font-size:24px;color:#111;margin-top:5px;white-space:nowrap}}
      .progress-wrap{{margin-top:14px}}
      .progress-label{{display:flex;justify-content:space-between;font-size:13px;color:#666;margin-bottom:6px}}
      .progress-bar{{height:10px;background:#eee;border-radius:999px;overflow:hidden}}
      .progress-bar i{{display:block;height:100%;width:{min(progress,100)}%;background:#111;border-radius:999px}}
      .section-block{{margin-top:24px}}
      .section-block h3,.set-card h3{{margin:0 0 12px;font-size:21px}}
      .desktop-table{{overflow:auto;background:#fff;border:1px solid #e3e3e3;border-radius:15px}}
      table{{width:100%;border-collapse:collapse}}
      th,td{{padding:12px 10px;border-bottom:1px solid #eee;text-align:center;white-space:nowrap}}
      th{{background:#fafafa;font-size:14px}}
      .detail-cell{{text-align:left;font-size:12px;line-height:1.7}}
      .role{{font-size:12px;color:#888;margin-top:4px}}
      .mobile-cards{{display:none}}
      .period-card,.member-card,.set-card{{background:#fff;border:1px solid #e3e3e3;border-radius:15px;padding:14px}}
      .period-card-head,.member-head{{display:flex;justify-content:space-between;gap:10px;align-items:flex-start}}
      .period-card-head span,.member-head span{{display:block;color:#777;font-size:12px;margin-top:3px}}
      .period-main{{display:flex;align-items:baseline;gap:5px;margin-top:13px}}
      .period-main strong{{font-size:30px}}
      .period-main span{{color:#777}}
      .period-main em{{margin-left:auto;font-style:normal;font-weight:900}}
      .period-sub,.member-summary{{display:flex;gap:10px;margin-top:12px}}
      .period-sub span,.member-summary span{{flex:1;background:#f7f7f7;padding:9px;border-radius:9px;text-align:center;font-size:13px}}
      .member-head strong{{font-size:22px}}
      details{{margin-top:12px}}
      summary{{cursor:pointer;font-weight:800;color:#5b2a86}}
      .member-periods{{display:grid;grid-template-columns:repeat(2,1fr);gap:8px;margin-top:10px}}
      .member-periods div{{background:#f8f8f8;padding:10px;border-radius:10px}}
      .member-periods span,.member-periods small{{display:block;color:#777;font-size:11px}}
      .member-periods b{{display:block;font-size:19px;margin:3px 0}}
      .set-card{{margin-top:20px}}
      .set-meta{{color:#666;line-height:1.6}}
      .set-card form{{display:flex;gap:8px;margin-top:12px}}
      .set-card input{{width:90px;padding:11px;border:1px solid #ddd;border-radius:10px;font-size:16px}}
      .set-card button{{flex:1;padding:11px 14px;background:#111;color:#fff;border:none;border-radius:10px;font-weight:900}}
      .empty-card{{max-width:600px;margin:40px auto;background:#fff;padding:20px;border-radius:14px}}
      @media(max-width:700px){{
        body{{padding:10px!important}}
        .team-shell{{width:100%}}
        .team-nav{{overflow-x:auto;flex-wrap:nowrap;padding-bottom:2px}}
        .team-nav a{{white-space:nowrap;flex:0 0 auto}}
        .hero{{padding:15px;border-radius:16px}}
        .hero h2{{font-size:24px}}
        .sub{{font-size:13px;line-height:1.5}}
        .date-nav{{margin-top:15px}}
        .date-nav b{{font-size:16px}}
        .summary-grid{{grid-template-columns:repeat(2,minmax(0,1fr));gap:8px}}
        .team-card{{padding:12px}}
        .team-card b{{font-size:21px}}
        .desktop-table{{display:none}}
        .mobile-cards{{display:grid;gap:10px}}
        .section-block{{margin-top:20px}}
        .section-block h3,.set-card h3{{font-size:20px}}
      }}
      @media(max-width:380px){{
        .summary-grid{{grid-template-columns:1fr 1fr}}
        .team-card b{{font-size:18px}}
        .member-periods{{grid-template-columns:1fr}}
      }}
    </style>
    <div class='team-shell'>
      <div class='team-nav'><a href='/'>등급조회</a><a href='/team'><b>팀 페이지</b></a><a href='/team-logout'>로그아웃</a></div>
      <section class='hero'>
        <h2>{_team_escape(team.get('name'))}</h2>
        <div class='sub'>팀장 {_team_escape(leader_name)} · 총 {len(team_all_keys(team))}명 · 팀 생성일 {team_start} · {'팀장 전체기간 조회' if leader_mode else '팀원 최근 7일 조회'}</div>
        <div class='date-nav'><a href='/team?day={prev_day}' style='{"pointer-events:none;opacity:.35" if prev_disabled else ""}'>← 이전날</a><b>{op_day}</b><a href='/team?day={next_day}' style='{"pointer-events:none;opacity:.35" if next_disabled else ""}'>다음날 →</a></div>
        <div class='summary-grid'>
          <div class='team-card'>신청 세트<b>{set_count}세트</b></div>
          <div class='team-card'>완료<b>{day_complete}/{day_quota}</b></div>
          <div class='team-card'>거절<b>{day_reject}</b></div>
          <div class='team-card'>취소<b>{day_cancel}</b></div>
        </div>
        <div class='progress-wrap'><div class='progress-label'><span>오늘 팀 달성률</span><b>{progress}%</b></div><div class='progress-bar'><i></i></div></div>
      </section>

      <section class='section-block'>
        <h3>팀 전체 구간 현황</h3>
        <div class='desktop-table'><table><thead><tr><th>구간</th><th>시간</th><th>완료/할당량</th><th>달성률</th></tr></thead><tbody>{period_rows}</tbody></table></div>
        <div class='mobile-cards'>{period_cards}</div>
      </section>

      {member_table}
      {set_form}
    </div>
    """
    return HTMLResponse(html_page("팀 페이지", body))
